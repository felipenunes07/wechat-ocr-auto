#!/usr/bin/env python3
"""
WeChat receipt ingestion daemon (Windows-friendly).

Goal:
- Detect new image files continuously (no manual save click loop).
- Process with OCR.
- Extract date, time, beneficiary and amount.
- Append results to the configured sink with idempotency.
- Avoid missing files via periodic reconciliation scan.

Notes:
- There is no official webhook from WeChat Desktop local storage.
- This script emulates webhook behavior using filesystem events + durable queue.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import logging
import os
import re
import sqlite3
import subprocess
import sys
import threading
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_FLOOR
from pathlib import Path, PureWindowsPath
from typing import Any, Optional
from urllib.parse import parse_qs, urlparse

from PIL import Image, ImageFilter, ImageOps, ImageStat
from openpyxl import Workbook, load_workbook

try:
    from watchdog.events import FileSystemEventHandler
    from watchdog.observers import Observer

    WATCHDOG_AVAILABLE = True
except Exception:
    WATCHDOG_AVAILABLE = False
    class FileSystemEventHandler:  # type: ignore[override]
        pass
    class Observer:  # type: ignore[override]
        pass

try:
    from wechat_ui_force_download import UIMessageCandidate, WeChatUIForceDownloader

    UI_FORCE_DOWNLOADER_AVAILABLE = True
    UI_FORCE_DOWNLOADER_IMPORT_ERROR: Optional[str] = None
except Exception as exc:
    UIMessageCandidate = Any  # type: ignore[misc,assignment]
    WeChatUIForceDownloader = None  # type: ignore[assignment]
    UI_FORCE_DOWNLOADER_AVAILABLE = False
    UI_FORCE_DOWNLOADER_IMPORT_ERROR = f"{type(exc).__name__}: {exc}"


IMG_HEADERS: dict[str, tuple[int, int]] = {
    "jpg": (0xFF, 0xD8),
    "png": (0x89, 0x50),
    "gif": (0x47, 0x49),
    "webp": (0x52, 0x49),
}

IMG_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif", ".dat"}
PLAIN_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif"}
LANCZOS_FILTER = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
BASE_LANC_HEADERS = [
    "CLIENTE",
    "DATA",
    "HORA",
    "BANCO",
    "VALOR",
]
DEFAULT_VERIFICATION_COLUMN_NAME = "STATUS_VERIFICACAO"
UI_FORCE_RUNTIME_META_KEY = "ui_force_runtime_enabled"
MANUAL_SESSION_META_KEY = "manual_session_started_at"
MANUAL_SESSION_ID_META_KEY = "manual_session_id"
REALTIME_SOURCE_EVENTS = {"created", "modified", "ui-force"}
MANUAL_OPEN_SOURCE_KINDS = {"msgattach_image_dat", "msgattach_image_plain", "temp_image"}
MANUAL_OPEN_ONLY_IGNORE_REASON = "IGNORED_MANUAL_OPEN_ONLY"
SESSION_PENDING_OPEN_STATE = "SESSION_PENDING_OPEN"
IGNORED_SESSION_ROLLOVER_STATE = "IGNORED_SESSION_ROLLOVER"
IGNORED_STALE_MANUAL_SESSION_STATE = "IGNORED_STALE_MANUAL_SESSION"
IGNORED_BY_USER_STATE = "IGNORED_BY_USER"
MANUAL_SESSION_TERMINAL_STATES = {
    "RESOLVED",
    "THUMB_FALLBACK",
    "EXCEPTION",
    IGNORED_SESSION_ROLLOVER_STATE,
    IGNORED_STALE_MANUAL_SESSION_STATE,
    IGNORED_BY_USER_STATE,
}
MANUAL_OPEN_ONLY_WAIT_REASONS = (
    "MANUAL_WAIT_ORIGINAL",
    "WAITING_ORIGINAL_MEDIA",
    "WAITING_TEMP_CONTEXT",
)
MANUAL_SESSION_FILE_HOLD_PREFIXES = (
    "WAITING_SESSION_PRIOR_MESSAGE_ORDER:",
    "WAITING_PRIOR_SINK_SESSION_MESSAGE:",
    "WAITING_PRIOR_SINK_RECEIPT:",
)


def is_candidate(path: Path, thumb_candidates_enabled: bool) -> bool:
    if not path.is_file():
        return False
    if path.suffix.lower() not in IMG_SUFFIXES:
        return False

    s = str(path).lower().replace("/", "\\")

    if "\\msgattach\\" in s and "\\image\\" in s and path.suffix.lower() == ".dat":
        return True

    # WeChat can store full images in plain formats (.png/.jpg) under MsgAttach/Image.
    if "\\msgattach\\" in s and "\\image\\" in s and path.suffix.lower() in PLAIN_IMAGE_SUFFIXES:
        return True

    # Optional fallback lane when thumbnail-only processing is desired.
    if thumb_candidates_enabled and "\\msgattach\\" in s and "\\thumb\\" in s and path.suffix.lower() == ".dat":
        return True

    if "\\filestorage\\temp\\" in s and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".bmp", ".webp"}:
        return True

    return False


def should_refresh_manual_session(source_kind: str, source_event: str) -> bool:
    return source_kind in MANUAL_OPEN_SOURCE_KINDS and str(source_event or "").strip().lower() in REALTIME_SOURCE_EVENTS


def is_message_job_terminal_state(state: Optional[str]) -> bool:
    return str(state or "").strip().upper() in MANUAL_SESSION_TERMINAL_STATES


def wall_duration_ms(start_ts: float | None, end_ts: float | None = None) -> Optional[float]:
    if start_ts is None or float(start_ts) <= 0:
        return None
    end_value = time.time() if end_ts is None else float(end_ts)
    return max(0.0, (end_value - float(start_ts)) * 1000.0)


def perf_duration_ms(start: float, end: float | None = None) -> float:
    end_value = time.perf_counter() if end is None else float(end)
    return max(0.0, (end_value - float(start)) * 1000.0)


def format_ms(value: Optional[float]) -> str:
    return "-" if value is None else f"{value:.0f}"


def hold_retry_delay_seconds(now: float, deadline: float, minimum: int = 2, maximum: int = 5) -> int:
    remaining = max(0.0, float(deadline) - float(now))
    if remaining <= 0:
        return max(1, int(minimum))
    return max(int(minimum), min(int(maximum), int(remaining) if remaining >= 1.0 else 1))


def runtime_media_resolver(media_resolver: Optional["WeChatDBResolver"]) -> Optional["WeChatDBResolver"]:
    if media_resolver is None:
        return None
    if media_resolver.last_error:
        return None
    return media_resolver


def candidate_initial_delay_seconds(source_kind: str, settle_seconds: int, thumb_candidates_enabled: bool) -> int:
    base_delay = max(1, int(settle_seconds))
    if source_kind == "temp_image" and not thumb_candidates_enabled:
        return max(3, base_delay)
    return base_delay


def detect_source_kind(path: Path) -> str:
    s = str(path).lower().replace("/", "\\")
    if "\\msgattach\\" in s and "\\image\\" in s and path.suffix.lower() == ".dat":
        return "msgattach_image_dat"
    if "\\msgattach\\" in s and "\\thumb\\" in s and path.suffix.lower() == ".dat":
        return "msgattach_thumb_dat"
    if "\\filestorage\\temp\\" in s:
        return "temp_image"
    if "\\msgattach\\" in s and "\\image\\" in s:
        return "msgattach_image_plain"
    return "other"


def normalize_windows_text(value: str) -> str:
    return str(value or "").replace("/", "\\").strip().lower()


def path_to_normalized_windows(path: Path) -> str:
    return normalize_windows_text(str(path))


def build_lanc_headers(verification_column_name: str) -> list[str]:
    return [*BASE_LANC_HEADERS]


def sheet_header_range(headers: list[str]) -> str:
    last_col = chr(ord("A") + max(0, len(headers) - 1))
    return f"A1:{last_col}1"


def sheet_table_range(headers: list[str]) -> str:
    last_col = chr(ord("A") + max(0, len(headers) - 1))
    return f"A:{last_col}"


def sheet_row_range(headers: list[str], row_idx: int) -> str:
    last_col = chr(ord("A") + max(0, len(headers) - 1))
    row_value = max(1, int(row_idx))
    return f"A{row_value}:{last_col}{row_value}"


def build_sink_row_values(row_payload: dict[str, Any]) -> list[Any]:
    return [
        row_payload.get("client"),
        row_payload.get("txn_date"),
        row_payload.get("txn_time"),
        row_payload.get("bank"),
        row_payload.get("amount"),
    ]


def resolve_full_image_from_thumb_path(thumb_path: Path) -> Optional[Path]:
    """Try to map MsgAttach/Thumb/<month>/<hash>_t.dat -> MsgAttach/Image/<month>/<hash>.(dat|jpg|png...)."""
    s = str(thumb_path).replace("/", "\\")
    if "\\msgattach\\" not in s.lower() or "\\thumb\\" not in s.lower():
        return None

    img_loc = s.replace("\\Thumb\\", "\\Image\\").replace("\\thumb\\", "\\Image\\")
    img_path = Path(img_loc)

    stem = img_path.stem
    base = stem[:-2] if stem.lower().endswith("_t") else stem
    candidates: list[Path] = []
    for ext in (".dat", ".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"):
        candidates.append(img_path.with_name(f"{base}{ext}"))
    for ext in (".dat", ".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"):
        candidates.append(img_path.with_name(f"{stem}{ext}"))

    for c in candidates:
        if c.exists() and c.is_file():
            return c
    return None


def expected_full_image_from_thumb_path(thumb_path: Path) -> Optional[Path]:
    s = str(thumb_path).replace("/", "\\")
    if "\\msgattach\\" not in s.lower() or "\\thumb\\" not in s.lower():
        return None

    img_loc = s.replace("\\Thumb\\", "\\Image\\").replace("\\thumb\\", "\\Image\\")
    img_path = Path(img_loc)
    stem = img_path.stem
    base = stem[:-2] if stem.lower().endswith("_t") else stem
    return img_path.with_name(f"{base}.dat")


def sha256_bytes(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


def decode_wechat_dat(raw: bytes) -> tuple[bytes, str, int]:
    if len(raw) < 2:
        raise ValueError("Empty or invalid .dat file")

    for ext, (h0, h1) in IMG_HEADERS.items():
        k0 = raw[0] ^ h0
        k1 = raw[1] ^ h1
        if k0 != k1:
            continue
        key = k0
        decoded = bytes(b ^ key for b in raw)
        try:
            with Image.open(io.BytesIO(decoded)) as im:
                im.verify()
            return decoded, ext, key
        except Exception:
            continue

    raise ValueError("Unable to decode .dat as known image format")


def open_image_from_file(path: Path) -> tuple[Image.Image, bytes, str, Optional[int]]:
    raw = path.read_bytes()
    if path.suffix.lower() == ".dat":
        decoded, ext, key = decode_wechat_dat(raw)
        with Image.open(io.BytesIO(decoded)) as im:
            return im.convert("RGB"), decoded, ext, key
    with Image.open(io.BytesIO(raw)) as im:
        return im.convert("RGB"), raw, path.suffix.lower().lstrip("."), None


def quality_score(img: Image.Image) -> float:
    w, h = img.size
    gray = img.convert("L")
    var = float(ImageStat.Stat(gray).var[0])
    long_side = max(w, h)
    short_side = min(w, h)

    res_component = min(1.0, long_side / 1400.0) * 0.65 + min(1.0, short_side / 700.0) * 0.20
    contrast_component = min(1.0, var / 1800.0) * 0.15
    score = res_component + contrast_component
    return round(max(0.0, min(1.0, score)), 4)


class OCREngine:
    name = "none"

    def extract(self, img: Image.Image) -> tuple[str, float]:
        raise NotImplementedError


class RapidOCREngine(OCREngine):
    name = "rapidocr"

    def __init__(self) -> None:
        from rapidocr_onnxruntime import RapidOCR  # type: ignore

        self._ocr = RapidOCR()

    def extract(self, img: Image.Image) -> tuple[str, float]:
        import numpy as np  # type: ignore

        arr = np.array(img.convert("RGB"))
        result, _ = self._ocr(arr)
        if not result:
            return "", 0.0
        texts: list[str] = []
        confs: list[float] = []
        for item in result:
            if len(item) >= 3:
                texts.append(str(item[1]))
                try:
                    confs.append(float(item[2]))
                except Exception:
                    pass
        text = "\n".join(t for t in texts if t.strip())
        conf = (sum(confs) / len(confs)) if confs else 0.5
        return text, round(max(0.0, min(1.0, conf)), 4)


class TesseractOCREngine(OCREngine):
    name = "tesseract"

    def __init__(self) -> None:
        import pytesseract  # type: ignore

        cmd = os.getenv("TESSERACT_CMD", "").strip()
        if cmd:
            pytesseract.pytesseract.tesseract_cmd = cmd
        self._pytesseract = pytesseract
        self._lang = os.getenv("OCR_LANG", "por+eng+chi_sim")

    def extract(self, img: Image.Image) -> tuple[str, float]:
        text = self._pytesseract.image_to_string(img, lang=self._lang)
        text = text.strip()
        if not text:
            return "", 0.0
        return text, 0.55


def build_ocr_engine() -> OCREngine:
    try:
        return RapidOCREngine()
    except Exception:
        pass
    try:
        return TesseractOCREngine()
    except Exception:
        pass
    raise RuntimeError(
        "No OCR engine available. Install one:\n"
        "- pip install rapidocr-onnxruntime\n"
        "or\n"
        "- pip install pytesseract and install Tesseract OCR binary"
    )


def warmup_ocr_engine(ocr: OCREngine) -> None:
    started = time.perf_counter()
    try:
        blank = Image.new("RGB", (64, 64), "white")
        ocr.extract(blank)
    except Exception as exc:
        print(f"[WARN] ocr_warmup_failed | err={type(exc).__name__}: {exc}")
        return
    print(f"[OCR] warmup_complete | ms={perf_duration_ms(started):.0f}")


DATE_PATTERNS = [
    re.compile(r"(?<!\d)(\d{1,2}/\d{1,2}/\d{4})"),
    re.compile(r"(?<!\d)(\d{4}-\d{2}-\d{2})"),
    re.compile(r"(?<!\d)(\d{1,2}-\d{1,2}-\d{4})"),
    re.compile(r"(?<!\d)(\d{1,2}/\d{1,2}/\d{2})(?!\d)"),
]
ALPHA_MONTH_DATE_PATTERN = re.compile(
    r"(?<!\d)(\d{1,2})(?:\s*de\s*|\s*[,\/\-.]?\s*)([a-z]{3,12})\.?(?:\s*de\s*|\s*[,\/\-.]?\s*)(\d{4})(?!\d)",
    re.IGNORECASE,
)
TIME_PATTERN = re.compile(r"(?<!\d)(\d{1,2}\s*(?::|h)\s*\d{2}(?:\s*(?::|h)\s*\d{2})?)(?!\d)", re.IGNORECASE)
AMOUNT_CURRENCY_PATTERN = re.compile(
    r"(?<![A-Z0-9])(R\$|RS|US\$|USD|BRL|CNY|RMB|¥|￥|R(?=\s))\s*([0-9][0-9\.,]{0,20})",
    re.IGNORECASE,
)
AMOUNT_FALLBACK_PATTERN = re.compile(
    r"(?<!\d)([0-9]{1,3}(?:[\.,][0-9]{3})+(?:[\.,][0-9]{1,2})?|[0-9]+[\.,][0-9]{1,2})(?!\d)"
)
AMOUNT_DIRECT_HINTS = (
    "valor",
    "pagamento",
    "transfer",
    "pix",
    "enviado",
    "recebido",
    "total",
)
AMOUNT_NEGATIVE_HINTS = (
    "tarifa",
    "taxa",
    "juros",
    "autentic",
    "documento",
    "agencia",
    "conta",
    "chave",
    "cnpj",
    "cpf",
    "telefone",
    "ouvidoria",
    "codigo",
    "protocolo",
    "id",
)
DOCUMENT_NUMBER_CONTEXT_HINTS = (
    "cnpj",
    "cpf",
    "chavepix",
    "iddatransacao",
    "transacaoid",
    "protocolo",
    "telefone",
    "ouvidoria",
    "agencia",
    "conta",
)
MONTH_TOKEN_MAP = {
    "JAN": 1,
    "JANEIRO": 1,
    "ENERO": 1,
    "JANUARY": 1,
    "FEB": 2,
    "FEV": 2,
    "FEVEREIRO": 2,
    "FEBRERO": 2,
    "FEBRUARY": 2,
    "MAR": 3,
    "MARCO": 3,
    "MARZO": 3,
    "MARCH": 3,
    "APR": 4,
    "ABR": 4,
    "ABRIL": 4,
    "APRIL": 4,
    "MAY": 5,
    "MAI": 5,
    "MAIO": 5,
    "MAYO": 5,
    "JUN": 6,
    "JUNHO": 6,
    "JUNIO": 6,
    "JUNE": 6,
    "JUL": 7,
    "JULHO": 7,
    "JULIO": 7,
    "JULY": 7,
    "AUG": 8,
    "AGO": 8,
    "AGOSTO": 8,
    "AUGUST": 8,
    "SEP": 9,
    "SET": 9,
    "SETEMBRO": 9,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "SEPTEMBER": 9,
    "OCT": 10,
    "OUT": 10,
    "OUTUBRO": 10,
    "OCTUBRE": 10,
    "OCTOBER": 10,
    "NOV": 11,
    "NOVEMBRO": 11,
    "NOVIEMBRE": 11,
    "NOVEMBER": 11,
    "DEC": 12,
    "DEZ": 12,
    "DEZEMBRO": 12,
    "DICIEMBRE": 12,
    "DECEMBER": 12,
}
DATE_CONTEXT_HINTS = (
    "data",
    "horario",
    "hora",
    "realizadaem",
    "realizadoem",
    "transferidoem",
    "transferido",
    "pagamento",
    "datadopagamento",
    "comprovante",
    "pix",
    "geracao",
)
TIME_CONTEXT_HINTS = (
    "horario",
    "hora",
    "as",
    "realizadaem",
    "realizadoem",
    "transferidoem",
    "pagamento",
    "datadopagamento",
    "comprovante",
    "pix",
    "geracao",
)
COMPACT_AMOUNT_CONTEXT_HINTS = (
    "comprovante",
    "valor",
    "pagamento",
    "transfer",
    "pix",
    "realizada",
    "realizado",
    "transferido",
    "transferencia",
)

BENEFICIARY_KEYS = [
    "favorecido",
    "beneficiario",
    "beneficiario",
    "destinatario",
    "destinatario",
    "recebedor",
    "recebedora",
    "para",
    "recebido por",
    "para:",
    "destino",
    "收款方",
    "收款人",
    "对方",
]

BENEFICIARY_SKIP_LABELS = {
    "nome",
    "origem",
    "destino",
    "cpf",
    "cnpj",
    "instituicao",
    "instituição",
    "chave pix",
    "chavepix",
    "id",
}

BANK_ALLOWED = ("AMD", "DIAMOND", "CLEEND")
CLIENT_LABEL_SPECIAL_CASES = {
    "PP": "6",
}
IGNORED_SENDER_USERNAMES = {
    "wxid_wml3ftd6qpea12",
    "wxid_jhb1tt23of8422",
    "wxid_5sd4qzz1lyhl12",
    "jinshuo2004",
}


def normalize_text_for_match(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", "", value)
    return value


def detect_bank(text: str, beneficiary: Optional[str]) -> Optional[str]:
    material = f"{text}\n{beneficiary or ''}"
    compact = normalize_text_for_match(material)
    if "DIAMOND" in compact:
        return "DIAMOND"
    if any(token in compact for token in ("CLEEND", "CLEND", "CUEEND", "GLEEND")):
        return "CLEEND"
    if "AMD" in compact:
        return "AMD"
    return None


def normalize_client_label(value: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    label = str(value or "").strip()
    if not label:
        return (None, None)

    normalized = unicodedata.normalize("NFKC", label).upper().replace("—", "-")
    normalized_without_year = re.sub(r"\b2026\b", " ", normalized)
    match = re.search(r"\d+(?:-\d+)*(?:[A-Z](?![A-Z]))?", normalized_without_year)
    if match:
        return (re.sub(r"[^A-Z0-9]", "", match.group(0)), None)

    compact_letters = re.sub(r"[^A-Z]", "", normalized_without_year)
    for key, mapped in CLIENT_LABEL_SPECIAL_CASES.items():
        if key in compact_letters:
            return (mapped, None)

    if not re.search(r"[A-Z0-9]", normalized):
        return (None, "IGNORED_CLIENT_LABEL_EMPTY")
    if not re.search(r"[A-Z0-9]", normalized_without_year):
        return (None, "IGNORED_CLIENT_LABEL_DECORATIVE")
    return (label, None)


def should_ignore_sender(msg_ref: Optional["WeChatMessageRef"]) -> bool:
    if msg_ref is None:
        return False
    sender = str(msg_ref.sender_user_name or "").strip().lower()
    talker = str(msg_ref.talker or "").strip().lower()
    return sender in IGNORED_SENDER_USERNAMES or talker in IGNORED_SENDER_USERNAMES


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_ocr_text_for_parsing(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value or "")
    normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
    normalized = (
        normalized.replace("，", ",")
        .replace("。", ".")
        .replace("：", ":")
        .replace("；", ";")
        .replace("／", "/")
        .replace("—", "-")
        .replace("–", "-")
        .replace("·", " ")
        .replace("•", " ")
        .replace("\u00a0", " ")
    )
    normalized = strip_accents(normalized).lower()
    normalized = re.sub(r"[^\S\n]+", " ", normalized)
    return normalized.strip()


def extract_beneficiary_name(lines: list[str]) -> Optional[str]:
    normalized_lines = [normalize_ocr_text_for_parsing(line) for line in lines]
    for idx, low in enumerate(normalized_lines):
        if not any(key in low for key in BENEFICIARY_KEYS):
            continue

        original = lines[idx].strip()
        if ":" in original:
            right = original.split(":", 1)[1].strip()
            if right and normalize_ocr_text_for_parsing(right) not in BENEFICIARY_SKIP_LABELS:
                return right

        probe_idx = idx + 1
        while probe_idx < len(lines):
            candidate = lines[probe_idx].strip()
            candidate_low = normalized_lines[probe_idx]
            if not candidate:
                probe_idx += 1
                continue
            if candidate_low in BENEFICIARY_SKIP_LABELS:
                probe_idx += 1
                continue
            return candidate
    return None


def today_local_date_str() -> str:
    return datetime.now().strftime("%d/%m/%Y")


def normalize_date_for_excel(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    v = value.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            dt = datetime.strptime(v, fmt)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            continue
    compact = re.sub(r"[^A-Za-z0-9]", "", strip_accents(v)).upper()
    m = re.fullmatch(r"(\d{1,2})([A-Z]{3,12})(\d{4})", compact)
    if m:
        token = m.group(2)
        month = MONTH_TOKEN_MAP.get(token)
        if month is not None:
            try:
                dt = datetime(int(m.group(3)), int(month), int(m.group(1)))
                return dt.strftime("%d/%m/%Y")
            except Exception:
                pass
    return None


def normalize_currency_code(value: str) -> Optional[str]:
    cur = (value or "").strip().upper()
    if cur in {"R$", "RS", "R"}:
        return "BRL"
    if cur == "US$":
        return "USD"
    return cur or None


def normalize_time_for_excel(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = re.sub(r"\s+", "", str(value).strip().lower()).replace("h", ":")
    parts = cleaned.split(":")
    if len(parts) < 2:
        return None
    try:
        h = int(parts[0])
        m = int(parts[1])
    except Exception:
        return None
    if h < 0 or h > 23 or m < 0 or m > 59:
        return None
    return f"{h:02d}:{m:02d}"


def _iter_date_candidates(text: str) -> list[tuple[str, int, int]]:
    candidates: list[tuple[str, int, int]] = []
    for pat in DATE_PATTERNS:
        for match in pat.finditer(text):
            normalized = normalize_date_for_excel(match.group(1))
            if normalized:
                candidates.append((normalized, match.start(1), match.end(1)))

    for match in ALPHA_MONTH_DATE_PATTERN.finditer(text):
        token = f"{match.group(1)}{match.group(2)}{match.group(3)}"
        normalized = normalize_date_for_excel(token)
        if normalized:
            candidates.append((normalized, match.start(1), match.end(3)))
    return candidates


def _iter_time_candidates(text: str) -> list[tuple[str, int, int]]:
    candidates: list[tuple[str, int, int]] = []
    for match in TIME_PATTERN.finditer(text):
        normalized = normalize_time_for_excel(match.group(1))
        if normalized:
            candidates.append((normalized, match.start(1), match.end(1)))
    return candidates


def _pick_best_date_candidate(lines: list[str]) -> Optional[str]:
    candidates: list[tuple[int, int, int, str]] = []
    for idx, line in enumerate(lines):
        if not line:
            continue
        prev_low = lines[idx - 1] if idx > 0 else ""
        next_low = lines[idx + 1] if idx + 1 < len(lines) else ""
        context_low = " ".join(part for part in (prev_low, line, next_low) if part)
        for value, start, _end in _iter_date_candidates(line):
            score = 0
            if any(token in context_low for token in DATE_CONTEXT_HINTS):
                score += 12
            if _iter_time_candidates(line):
                score += 8
            if idx == 0 and len(line.strip()) <= 24:
                score -= 2
            candidates.append((-score, idx, start, value))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][3]


def _pick_best_time_candidate(lines: list[str]) -> Optional[str]:
    candidates: list[tuple[int, int, int, str]] = []
    for idx, line in enumerate(lines):
        if not line:
            continue
        prev_low = lines[idx - 1] if idx > 0 else ""
        next_low = lines[idx + 1] if idx + 1 < len(lines) else ""
        context_low = " ".join(part for part in (prev_low, line, next_low) if part)
        stripped = line.strip()
        line_dates = _iter_date_candidates(line)
        for value, start, _end in _iter_time_candidates(line):
            score = 0
            if any(token in context_low for token in TIME_CONTEXT_HINTS):
                score += 12
            if line_dates:
                score += 10
            if idx == 0 and re.fullmatch(r"\d{1,2}(?::|h)\d{2}", stripped):
                score -= 14
            if len(stripped) <= 8 and re.fullmatch(r"\d{1,2}(?::|h)\d{2}", stripped):
                score -= 8
            candidates.append((-score, idx, start, value))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][3]


def extract_datetime_values(text: str) -> tuple[Optional[str], Optional[str]]:
    normalized = normalize_ocr_text_for_parsing(text)
    date_candidates = _iter_date_candidates(normalized)
    time_candidates = _iter_time_candidates(normalized)

    paired: list[tuple[int, int, int, str, str]] = []
    for date_value, date_start, date_end in date_candidates:
        suffix_match = re.match(
            r"^\s*(?:as)?\s*,?\s*(\d{1,2}\s*(?::|h)\s*\d{2}(?:\s*(?::|h)\s*\d{2})?)",
            normalized[date_end : date_end + 24],
        )
        if suffix_match:
            suffix_time = normalize_time_for_excel(suffix_match.group(1))
            if suffix_time:
                bridge = normalized[date_end : date_end + suffix_match.end(1)]
                score = 110
                if "as" in bridge:
                    score += 6
                paired.append((-score, date_start, date_end, date_value, suffix_time))
        for time_value, time_start, _time_end in time_candidates:
            gap = time_start - date_end
            if gap < 0 or gap > 20:
                continue
            bridge = normalized[date_end:time_start]
            score = 100 - gap
            if "as" in bridge:
                score += 10
            if "," in bridge or "." in bridge:
                score += 2
            paired.append((-score, date_start, time_start, date_value, time_value))

    if paired:
        paired.sort()
        _score, _date_start, _time_start, date_value, time_value = paired[0]
        return date_value, time_value

    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    return _pick_best_date_candidate(lines), _pick_best_time_candidate(lines)


def extract_first_date_value(text: str) -> Optional[str]:
    txn_date, _txn_time = extract_datetime_values(text)
    return txn_date


def _count_date_matches(text: str) -> int:
    normalized = normalize_ocr_text_for_parsing(text)
    return len(_iter_date_candidates(normalized))


def looks_like_single_receipt(text: str) -> tuple[bool, str]:
    low = text.lower()
    compact_low = re.sub(r"\s+", "", low)
    date_count = _count_date_matches(text)
    time_count = len(_iter_time_candidates(normalize_ocr_text_for_parsing(text)))
    amount_count = len(AMOUNT_FALLBACK_PATTERN.findall(text)) + len(AMOUNT_CURRENCY_PATTERN.findall(text))

    has_strong_kw = any(
        kw in low
        for kw in (
            "comprovante",
            "pix",
            "transferência",
            "transferencia",
            "pagamento",
            "recibo",
            "receipt",
            "收款",
            "转账",
            "付款",
            "交易",
        )
    )

    has_table_header = (
        ("data" in low and "hora" in low and "banco" in low and "transfer" in low)
        or ("horario" in low and "banco" in low and "transfer" in low)
    )
    has_balance_summary = (
        ("saldo antigo" in low or "saldoantigo" in compact_low)
        and ("saldo atual" in low or "saldoatual" in compact_low)
        and any(
            token in low or token.replace(" ", "") in compact_low
            for token in ("cheque", "dinheiro", "chq devolvido", "tx cheque", "no caiu")
        )
    )

    if has_balance_summary:
        return (False, "BALANCE_SUMMARY")
    if has_table_header and ("total" in low or not has_strong_kw):
        return (False, "TABULAR_TRANSFER_LIST")
    if has_table_header and (date_count >= 3 or time_count >= 3):
        return (False, "TABULAR_TRANSFER_LIST")
    if date_count >= 4 and amount_count >= 6:
        return (False, "MULTI_TRANSACTION_LIST")
    if not has_strong_kw and date_count >= 2 and amount_count >= 4:
        return (False, "WEAK_RECEIPT_SIGNAL")

    return (True, "OK")


def normalize_amount(value: str) -> Optional[float]:
    s = re.sub(r"[^\d,\.]", "", value.strip())
    if not s:
        return None

    grouped_thousands_comma = bool(re.fullmatch(r"\d{1,3}(?:,\d{3})+", s))
    grouped_thousands_dot = bool(re.fullmatch(r"\d{1,3}(?:\.\d{3})+", s))
    grouped_thousands_comma_compact_cent = bool(re.fullmatch(r"\d{1,3}(?:,\d{3})+\d{2}", s))
    grouped_thousands_dot_compact_cent = bool(re.fullmatch(r"\d{1,3}(?:\.\d{3})+\d{2}", s))
    repeated_dot_decimal = bool(re.fullmatch(r"\d{1,3}(?:\.\d{3})+\.\d{1,2}", s))
    repeated_comma_decimal = bool(re.fullmatch(r"\d{1,3}(?:,\d{3})+,\d{1,2}", s))

    if grouped_thousands_comma_compact_cent:
        s = s.replace(",", "")
        s = f"{s[:-2]}.{s[-2:]}"
    elif grouped_thousands_dot_compact_cent:
        s = s.replace(".", "")
        s = f"{s[:-2]}.{s[-2:]}"
    elif repeated_dot_decimal:
        integer_part, decimal_part = s.rsplit(".", 1)
        s = f"{integer_part.replace('.', '')}.{decimal_part}"
    elif repeated_comma_decimal:
        integer_part, decimal_part = s.rsplit(",", 1)
        s = f"{integer_part.replace(',', '')}.{decimal_part}"
    elif "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        if grouped_thousands_comma:
            s = s.replace(",", "")
        elif re.search(r",\d{1,2}$", s):
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "." in s:
        if grouped_thousands_dot:
            s = s.replace(".", "")
        elif re.search(r"\.\d{1,2}$", s):
            pass
        elif re.search(r"\.\d{3}$", s):
            s = s.replace(".", "")
    try:
        return round(float(s), 2)
    except Exception:
        return None


def round_amount_for_output(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    dec = Decimal(str(value)).quantize(Decimal("0.01"))
    integer = dec.to_integral_value(rounding=ROUND_FLOOR)
    if (dec - integer) >= Decimal("0.50"):
        integer += 1
    return float(integer)


def should_apply_compact_cent_fix(raw_value: str, currency: Optional[str], context_low: str) -> bool:
    digits = re.sub(r"\D", "", raw_value or "")
    if not digits or "." in raw_value or "," in raw_value:
        return False
    if currency != "BRL":
        return False
    if len(digits) < 4 or len(digits) > 6:
        return False
    if digits.endswith("00"):
        return False
    return any(token in context_low for token in COMPACT_AMOUNT_CONTEXT_HINTS)


@dataclass(frozen=True)
class AmountParseResult:
    value: Optional[float]
    rounded_value: Optional[float]
    currency: Optional[str]
    raw_value: Optional[str]
    source: str
    used_compact_cent_fix: bool = False


def extract_best_amount(lines: list[str]) -> AmountParseResult:
    candidates: list[tuple[int, int, int, float, Optional[str], str, str, bool]] = []
    order = 0
    for idx, line in enumerate(lines):
        prev2_low = lines[idx - 2].lower() if idx > 1 else ""
        prev_low = lines[idx - 1].lower() if idx > 0 else ""
        line_low = line.lower()
        next_low = lines[idx + 1].lower() if idx + 1 < len(lines) else ""
        next2_low = lines[idx + 2].lower() if idx + 2 < len(lines) else ""
        context_low = " ".join(part for part in (prev2_low, prev_low, line_low, next_low, next2_low) if part)
        compact_prev_low = re.sub(r"[^a-z0-9]+", "", prev_low)
        compact_line_low = re.sub(r"[^a-z0-9]+", "", line_low)
        compact_next_low = re.sub(r"[^a-z0-9]+", "", next_low)
        has_direct_hint = any(token in part for part in (prev_low, line_low, next_low) for token in AMOUNT_DIRECT_HINTS)
        has_negative_hint = any(token in context_low for token in AMOUNT_NEGATIVE_HINTS)
        line_has_document_hint = any(token in compact_line_low for token in DOCUMENT_NUMBER_CONTEXT_HINTS)
        adjacent_document_hint = any(
            token in compact_part
            for compact_part in (compact_prev_low, compact_next_low)
            for token in DOCUMENT_NUMBER_CONTEXT_HINTS
        )

        def score_candidate(raw_value: str, currency: Optional[str], source: str, used_compact_fix: bool) -> int:
            score = 30 if source == "currency" else 18
            if any(token in prev_low for token in AMOUNT_DIRECT_HINTS):
                score += 18
            if any(token in line_low for token in AMOUNT_DIRECT_HINTS):
                score += 10
            if any(token in next_low for token in AMOUNT_DIRECT_HINTS):
                score += 4
            if any(token in context_low for token in AMOUNT_NEGATIVE_HINTS):
                score -= 14
            if re.search(r"[.,]\d{1,2}$", raw_value):
                score += 4
            if re.fullmatch(r"\d{1,3}(?:[.,]\d{3})+(?:[.,]\d{1,2})?", raw_value):
                score += 3
            if currency == "BRL":
                score += 2
            if "." not in raw_value and "," not in raw_value:
                score -= 3
            if used_compact_fix:
                score += 7
            return score

        def should_skip_candidate(raw_value: str, source: str, currency: Optional[str]) -> bool:
            del currency
            if source != "fallback":
                return False
            embedded_document_number = bool(re.search(rf"{re.escape(raw_value)}\s*[/:-]\s*\d", line_low))
            digits = re.sub(r"\D", "", raw_value or "")
            has_fractional_tail = bool(re.search(r"[.,]\d{1,2}$", raw_value))
            if embedded_document_number and line_has_document_hint:
                return True
            if line_has_document_hint and len(digits) >= 6 and not has_fractional_tail:
                return True
            if adjacent_document_hint and len(digits) >= 10 and not has_fractional_tail:
                return True
            if not has_negative_hint or has_direct_hint:
                return False
            if len(digits) < 6:
                return False
            if has_fractional_tail:
                return False
            return True

        for m in AMOUNT_CURRENCY_PATTERN.finditer(line):
            raw_value = m.group(2)
            currency = normalize_currency_code(m.group(1))
            used_compact_fix = False
            value = normalize_amount(raw_value)
            source = "currency"
            if should_apply_compact_cent_fix(raw_value, currency, context_low):
                compact_value = normalize_amount(f"{raw_value[:-2]},{raw_value[-2:]}")
                if compact_value is not None:
                    value = compact_value
                    used_compact_fix = True
                    source = "currency_compact_cent_fix"
            if value is None:
                continue
            candidates.append((score_candidate(raw_value, currency, "currency", used_compact_fix), idx, order, value, currency, raw_value, source, used_compact_fix))
            order += 1

        for m in AMOUNT_FALLBACK_PATTERN.finditer(line):
            raw_value = m.group(1)
            if should_skip_candidate(raw_value, "fallback", None):
                continue
            value = normalize_amount(raw_value)
            if value is None:
                continue
            candidates.append((score_candidate(raw_value, None, "fallback", False), idx, order, value, None, raw_value, "fallback", False))
            order += 1

    if not candidates:
        return AmountParseResult(value=None, rounded_value=None, currency=None, raw_value=None, source="missing")

    candidates.sort(key=lambda item: (-item[0], item[1], item[2]))
    best_score, best_idx, _best_order, best_value, best_currency, best_raw_value, best_source, best_used_compact_fix = candidates[0]
    if best_currency is None:
        for score, idx, _order, value, currency, _raw_value, _source, _used_fix in candidates:
            if score != best_score or idx != best_idx or value != best_value:
                continue
            if currency is not None:
                best_currency = currency
                break
    return AmountParseResult(
        value=best_value,
        rounded_value=round_amount_for_output(best_value),
        currency=best_currency,
        raw_value=best_raw_value,
        source=best_source,
        used_compact_cent_fix=best_used_compact_fix,
    )


def prepare_image_for_ocr(img: Image.Image, source_kind: str) -> Image.Image:
    out = img.convert("RGB")
    w, h = out.size
    is_thumb_like = source_kind == "msgattach_thumb_dat" or max(w, h) <= 420
    if not is_thumb_like:
        max_side = max(w, h)
        if max_side > 1600:
            scale = 1600.0 / float(max_side)
            out = out.resize((max(1, int(w * scale)), max(1, int(h * scale))), LANCZOS_FILTER)
        return out

    # Miniatures are tiny and blurry; upscale + contrast helps OCR.
    if max(w, h) <= 260:
        scale = 4
    elif max(w, h) <= 420:
        scale = 3
    else:
        scale = 2

    out = out.resize((w * scale, h * scale), LANCZOS_FILTER)
    gray = out.convert("L")
    gray = ImageOps.autocontrast(gray, cutoff=2)
    gray = gray.filter(ImageFilter.MedianFilter(size=3))
    gray = gray.filter(ImageFilter.SHARPEN)
    return gray.convert("RGB")


def parse_receipt_fields(text: str, ocr_conf: float, q_score: float) -> dict[str, Any]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)

    parsed_txn_date, parsed_txn_time = extract_datetime_values(raw)
    txn_date_source = "parsed" if parsed_txn_date else "fallback_today"
    txn_time_source = "parsed" if parsed_txn_time else "fallback_dash"
    txn_date = parsed_txn_date or today_local_date_str()
    txn_time = parsed_txn_time or "-"

    amount_result = extract_best_amount(lines)
    amount = amount_result.value
    currency = amount_result.currency
    if amount is not None and currency is None:
        currency = "BRL"

    beneficiary = extract_beneficiary_name(lines)

    bank = detect_bank(raw, beneficiary)

    has_receipt_keyword = any(
        kw in raw.lower()
        for kw in [
            "pix",
            "comprovante",
            "transfer",
            "pagamento",
            "valor",
            "favorecido",
            "destinat",
            "recibo",
            "收款",
            "转账",
            "付款",
            "金额",
        ]
    )

    parse_conf = 0.0
    parse_conf += min(0.20, max(0.0, ocr_conf) * 0.20)
    parse_conf += 0.35 if amount is not None else 0.0
    parse_conf += 0.20 if parsed_txn_date else 0.0
    parse_conf += 0.10 if parsed_txn_time else 0.0
    parse_conf += 0.15 if beneficiary else 0.0
    parse_conf += 0.10 if bank else 0.0
    parse_conf += 0.10 if has_receipt_keyword else 0.0
    parse_conf += min(0.10, q_score * 0.10)
    parse_conf = round(min(1.0, parse_conf), 4)

    return {
        "txn_date": txn_date,
        "txn_date_source": txn_date_source,
        "txn_time": txn_time,
        "txn_time_source": txn_time_source,
        "beneficiary": beneficiary,
        "bank": bank,
        "amount": amount,
        "amount_raw": amount_result.raw_value,
        "amount_rounded": amount_result.rounded_value,
        "amount_source": amount_result.source,
        "amount_used_compact_cent_fix": amount_result.used_compact_cent_fix,
        "currency": currency,
        "parse_conf": parse_conf,
        "has_receipt_keyword": has_receipt_keyword,
    }


def uses_thumb_fallback_resolution(resolution_source: Optional[str]) -> bool:
    low = str(resolution_source or "").strip().lower()
    return "thumb" in low and "fallback" in low


def compute_review_needed(
    fields: dict[str, Any],
    bank: Optional[str],
    quality_score_value: float,
    verification_status: Optional[str],
    min_confidence: float,
    resolution_source: Optional[str],
) -> bool:
    using_thumb_fallback = uses_thumb_fallback_resolution(resolution_source)
    quality_floor = 0.20 if using_thumb_fallback else 0.38
    conf_floor = max(min_confidence, 0.70) if using_thumb_fallback else min_confidence
    return (
        fields.get("amount") is None
        or bank is None
        or fields.get("parse_conf", 0.0) < conf_floor
        or quality_score_value < quality_floor
        or verification_status != "CONFIRMADO"
        or fields.get("txn_date_source") != "parsed"
        or fields.get("txn_time_source") != "parsed"
        or fields.get("amount_source") == "currency_compact_cent_fix"
    )


def build_sheet_payload_from_receipt(
    receipt_payload: dict[str, Any],
    existing_payload: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    payload = dict(existing_payload or {})
    payload.update(
        {
            "file_id": receipt_payload.get("file_id"),
            "client": receipt_payload.get("client"),
            "txn_date": receipt_payload.get("txn_date"),
            "txn_time": receipt_payload.get("txn_time"),
            "bank": receipt_payload.get("bank"),
            "amount": (
                receipt_payload.get("amount_rounded")
                if receipt_payload.get("amount_rounded") is not None
                else receipt_payload.get("amount")
            ),
            "verification_status": receipt_payload.get("verification_status"),
            "msg_svr_id": receipt_payload.get("msg_svr_id"),
            "talker": receipt_payload.get("talker"),
        }
    )
    return payload


@dataclass
class QueueItem:
    file_id: str
    path: str
    source_kind: str
    ext: str
    size: int
    mtime: float
    first_seen: float
    attempts: int
    msg_svr_id: Optional[str] = None
    talker: Optional[str] = None
    msg_create_time: float = 0.0
    manual_session_id: Optional[str] = None
    session_release_at: float = 0.0


@dataclass
class WeChatMessageRef:
    msg_svr_id: Optional[str]
    talker: Optional[str]
    create_time: float
    sender_user_name: Optional[str]
    sender_display: Optional[str]
    image_rel_path: Optional[str]
    thumb_rel_path: Optional[str]
    image_abs_path: Optional[Path]
    thumb_abs_path: Optional[Path]

    def preferred_context_path(self) -> Optional[Path]:
        if self.image_abs_path is not None:
            return self.image_abs_path
        return self.thumb_abs_path

    def group_hash(self) -> Optional[str]:
        ctx = self.preferred_context_path()
        if ctx is None:
            return None
        return extract_group_id_from_path(ctx)


@dataclass
class MediaResolution:
    original_source_path: Path
    original_source_kind: str
    resolved_path: Path
    resolved_source_kind: str
    client_source_path: Path
    resolution_source: str
    verification_status: str
    msg_ref: Optional[WeChatMessageRef]
    using_thumb_fallback: bool = False


def extract_group_id_from_path(path: Path) -> Optional[str]:
    parts = path.parts
    for idx, part in enumerate(parts):
        if part.lower() == "msgattach" and idx + 1 < len(parts):
            return parts[idx + 1]
    return None


class ClientResolver:
    def __init__(self, map_path: Path) -> None:
        self.map_path = map_path
        self._mtime: float = -1.0
        self._map: dict[str, str] = {}
        self._raw_map: dict[str, str] = {}
        self.reload_if_needed(force=True)

    def _normalize_keys(self, data: dict[str, Any]) -> dict[str, str]:
        out: dict[str, str] = {}
        for k, v in data.items():
            key = str(k).strip().lower()
            client, _ignore_reason = normalize_client_label(v)
            if key and client:
                out[key] = client
        return out

    def ignore_reason(self, source_path: Path) -> Optional[str]:
        self.reload_if_needed()
        gid = extract_group_id_from_path(source_path)
        if not gid:
            return None
        raw_value = self._raw_map.get(str(gid).strip().lower())
        if raw_value is None:
            return None
        _client, ignore_reason = normalize_client_label(raw_value)
        return ignore_reason

    def reload_if_needed(self, force: bool = False) -> None:
        if not self.map_path.exists():
            if force or self._map:
                self._map = {}
                self._raw_map = {}
                self._mtime = -1.0
            return
        mtime = self.map_path.stat().st_mtime
        if not force and mtime == self._mtime:
            return
        try:
            raw = self.map_path.read_text(encoding="utf-8")
            data = json.loads(raw)
            if isinstance(data, dict):
                self._raw_map = {
                    str(k).strip().lower(): str(v).strip()
                    for k, v in data.items()
                    if str(k).strip() and str(v).strip()
                }
                self._map = self._normalize_keys(data)
            else:
                self._raw_map = {}
                self._map = {}
            self._mtime = mtime
        except Exception:
            self._raw_map = {}
            self._map = {}
            self._mtime = mtime

    def resolve(self, source_path: Path) -> Optional[str]:
        self.reload_if_needed()
        gid = extract_group_id_from_path(source_path)
        if not gid:
            return None
        key = gid.strip().lower()
        if key in self._map:
            return self._map[key]
        return None


class WeChatDBResolver:
    _MERGE_RESULT_PREFIX = "__WXMERGE__"

    def __init__(self, watch_roots: list[Path], merge_path: Path, refresh_seconds: int = 10) -> None:
        self.watch_roots = [p.resolve() for p in watch_roots]
        self.wx_dirs = [p.parent.resolve() for p in self.watch_roots]
        self.wechat_root = self.wx_dirs[0].parent if self.wx_dirs else None
        self.merge_path = merge_path.resolve()
        self.refresh_seconds = max(5, int(refresh_seconds))
        self.merge_timeout_seconds = 60
        self.failure_backoff_seconds = max(30, self.refresh_seconds, self.merge_timeout_seconds * 2)
        self._pywxdump: Any = None
        self._decode_bytes_extra: Any = None
        self._wx_key: Optional[str] = None
        self._wx_dir: Optional[Path] = None
        self._last_refresh = 0.0
        self._last_failure = 0.0
        self._last_error: Optional[str] = None
        self._lock = threading.Lock()
        self._load_dependencies()
        self._load_account_info(force=True)

    @property
    def available(self) -> bool:
        return self._pywxdump is not None and self._decode_bytes_extra is not None and self._wx_key is not None and self._wx_dir is not None

    @property
    def last_error(self) -> Optional[str]:
        return self._last_error

    @property
    def selected_wx_dir(self) -> Optional[Path]:
        return self._wx_dir

    def _load_dependencies(self) -> None:
        try:
            import pywxdump  # type: ignore
            from pywxdump.db.dbMSG import get_BytesExtra  # type: ignore

            # pywxdump emits a warning for unsupported account/nickname offsets even when key/wx_dir are usable.
            # Keep errors visible, but suppress warning noise that confuses operators.
            logging.getLogger("wx_core").setLevel(logging.ERROR)
            self._pywxdump = pywxdump
            self._decode_bytes_extra = get_BytesExtra
        except Exception as exc:
            self._last_error = f"pywxdump_unavailable:{type(exc).__name__}:{exc}"

    def _load_account_info(self, force: bool = False) -> bool:
        if not force and self._wx_key and self._wx_dir is not None:
            return True
        if self._pywxdump is None:
            return False
        try:
            infos = self._pywxdump.get_wx_info(is_print=False)
        except Exception as exc:
            self._last_error = f"wx_info_failed:{type(exc).__name__}:{exc}"
            return False
        if not infos:
            self._last_error = "wx_info_empty"
            return False

        chosen: Optional[dict[str, Any]] = None
        normalized_dirs = {path_to_normalized_windows(p) for p in self.wx_dirs}
        for info in infos:
            wx_dir_raw = str(info.get("wx_dir") or "").strip()
            if not wx_dir_raw:
                continue
            if normalize_windows_text(wx_dir_raw) in normalized_dirs:
                chosen = info
                break
        if chosen is None:
            chosen = infos[0]

        key = str(chosen.get("key") or "").strip()
        wx_dir_raw = str(chosen.get("wx_dir") or "").strip()
        if not key or not wx_dir_raw:
            self._last_error = "wx_info_missing_key_or_dir"
            return False

        self._wx_key = key
        self._wx_dir = Path(wx_dir_raw)
        if self.wechat_root is None:
            self.wechat_root = self._wx_dir.parent
        self._last_error = None
        return True

    @classmethod
    def _parse_merge_runner_output(cls, output: str) -> Optional[dict[str, Any]]:
        prefix = cls._MERGE_RESULT_PREFIX
        for line in reversed(str(output or "").splitlines()):
            if not line.startswith(prefix):
                continue
            raw_payload = line[len(prefix) :].strip()
            try:
                payload = json.loads(raw_payload)
            except Exception:
                return None
            if isinstance(payload, dict):
                return payload
        return None

    def _terminate_process_tree(self, proc: subprocess.Popen[str]) -> None:
        if proc.poll() is not None:
            return
        try:
            if os.name == "nt":
                subprocess.run(
                    ["taskkill", "/PID", str(proc.pid), "/T", "/F"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=5,
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                )
            else:
                proc.kill()
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass

    def _merge_real_time_db_with_timeout(self) -> tuple[bool, str]:
        assert self._wx_key is not None
        assert self._wx_dir is not None

        runner = (
            "import json, sys\n"
            "import pywxdump\n"
            "code, ret = pywxdump.all_merge_real_time_db(sys.argv[1], sys.argv[2], sys.argv[3])\n"
            f"print('{self._MERGE_RESULT_PREFIX}' + json.dumps({{'code': bool(code), 'ret': ret if isinstance(ret, str) else repr(ret)}}, ensure_ascii=False))\n"
        )
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        try:
            proc = subprocess.Popen(
                [
                    sys.executable,
                    "-X",
                    "utf8",
                    "-c",
                    runner,
                    self._wx_key,
                    str(self._wx_dir),
                    str(self.merge_path),
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=creationflags,
            )
            stdout, stderr = proc.communicate(timeout=self.merge_timeout_seconds)
        except subprocess.TimeoutExpired as exc:
            self._terminate_process_tree(proc)
            try:
                stdout, stderr = proc.communicate(timeout=5)
            except Exception:
                stdout = exc.stdout or ""
                stderr = exc.stderr or ""
            out = str(stdout or "").strip()
            err = str(stderr or "").strip()
            detail = f"merge_runner_timeout:{self.merge_timeout_seconds}s"
            if out:
                detail += f"|out={out[-300:]}"
            if err:
                detail += f"|err={err[-300:]}"
            return False, detail
        except Exception as exc:
            return False, f"merge_runner_failed:{type(exc).__name__}:{exc}"

        payload = self._parse_merge_runner_output(stdout)
        if payload is None:
            stdout_tail = str(stdout or "").strip()[-300:]
            stderr_tail = str(stderr or "").strip()[-300:]
            return (
                False,
                f"merge_runner_invalid_output:exit={proc.returncode}|out={stdout_tail}|err={stderr_tail}",
            )
        return bool(payload.get("code")), str(payload.get("ret"))

    def _mark_refresh_failure(self, now: float, detail: str) -> bool:
        self._last_failure = now
        self._last_error = detail
        return False

    def refresh_if_due(self, force: bool = False) -> bool:
        with self._lock:
            now = time.time()
            if not force and self.merge_path.exists() and (now - self._last_refresh) < self.refresh_seconds:
                return True
            if not force and self._last_failure > 0 and (now - self._last_failure) < self.failure_backoff_seconds:
                return False
            if not self._load_account_info(force=force):
                self._last_failure = now
                return False
            self.merge_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                assert self._pywxdump is not None
                code, ret = self._merge_real_time_db_with_timeout()
            except Exception as exc:
                return self._mark_refresh_failure(now, f"merge_failed:{type(exc).__name__}:{exc}")
            if not code:
                return self._mark_refresh_failure(now, f"merge_failed:{ret}")
            self._last_refresh = now
            self._last_failure = 0.0
            self._last_error = None
            return True

    def _absolute_path_from_rel(self, rel_path: Optional[str]) -> Optional[Path]:
        if not rel_path or self.wechat_root is None:
            return None
        parts = [part for part in PureWindowsPath(rel_path).parts if part not in ("\\", "/")]
        if not parts:
            return None
        return self.wechat_root.joinpath(*parts)

    def _extract_media_paths(self, bytes_extra: Any) -> tuple[Optional[str], Optional[str], Optional[str]]:
        image_rel: Optional[str] = None
        thumb_rel: Optional[str] = None
        sender_user_name: Optional[str] = None
        try:
            decoded = self._decode_bytes_extra(bytes_extra) if self._decode_bytes_extra else {}
        except Exception:
            decoded = {}

        items = decoded.get("3") if isinstance(decoded, dict) else None
        if isinstance(items, list):
            for item in items:
                if not isinstance(item, dict):
                    continue
                key = str(item.get("1") or "").strip()
                value = str(item.get("2") or "").strip()
                if key == "1" and sender_user_name is None and value:
                    sender_user_name = value
                if "filestorage" not in value.lower():
                    continue
                if key == "4" and image_rel is None:
                    image_rel = value
                elif key == "3" and thumb_rel is None:
                    thumb_rel = value

        if image_rel or thumb_rel:
            return image_rel, thumb_rel, sender_user_name

        raw_text = str(decoded)
        matches = re.findall(r"(wxid_[^\\']+\\FileStorage\\[^']+)", raw_text)
        for match in matches:
            lowered = match.lower()
            if "\\image\\" in lowered and image_rel is None:
                image_rel = match
            elif "\\thumb\\" in lowered and thumb_rel is None:
                thumb_rel = match
        return image_rel, thumb_rel, sender_user_name

    def _recent_messages(
        self,
        pivot_ts: float,
        lookback_sec: int,
        lookahead_sec: int,
        limit: int = 80,
    ) -> list[WeChatMessageRef]:
        if not self.refresh_if_due():
            return []
        if not self.merge_path.exists():
            return []

        lower = max(0, int(pivot_ts) - max(5, int(lookback_sec)))
        upper = int(pivot_ts) + max(1, int(lookahead_sec))
        conn = sqlite3.connect(str(self.merge_path))
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                """
                SELECT MsgSvrID, StrTalker, CreateTime, BytesExtra
                FROM MSG
                WHERE Type=3
                  AND CreateTime BETWEEN ? AND ?
                ORDER BY ABS(CreateTime - ?) ASC, CreateTime DESC
                LIMIT ?
                """,
                (lower, upper, int(pivot_ts), int(max(1, limit))),
            ).fetchall()
        finally:
            conn.close()

        out: list[WeChatMessageRef] = []
        for row in rows:
            image_rel, thumb_rel, sender_user_name = self._extract_media_paths(row["BytesExtra"])
            if not image_rel and not thumb_rel:
                continue
            out.append(
                WeChatMessageRef(
                    msg_svr_id=str(row["MsgSvrID"]) if row["MsgSvrID"] is not None else None,
                    talker=str(row["StrTalker"]) if row["StrTalker"] is not None else None,
                    create_time=float(row["CreateTime"]),
                    sender_user_name=sender_user_name,
                    sender_display=None,
                    image_rel_path=image_rel,
                    thumb_rel_path=thumb_rel,
                    image_abs_path=self._absolute_path_from_rel(image_rel),
                    thumb_abs_path=self._absolute_path_from_rel(thumb_rel),
                )
            )
        return out

    def _candidate_norms(self, path: Path) -> set[str]:
        norms = {path_to_normalized_windows(path)}
        if self.wechat_root is not None:
            try:
                rel = path.resolve().relative_to(self.wechat_root.resolve())
                norms.add(normalize_windows_text(str(PureWindowsPath(*rel.parts))))
            except Exception:
                pass
        return norms

    def find_message_for_path(self, path: Path, pivot_ts: float) -> Optional[WeChatMessageRef]:
        path_norms = self._candidate_norms(path)
        for msg in self._recent_messages(pivot_ts, lookback_sec=180, lookahead_sec=45, limit=120):
            msg_paths = {normalize_windows_text(p) for p in (msg.image_rel_path, msg.thumb_rel_path) if p}
            if msg.image_abs_path is not None:
                msg_paths.add(path_to_normalized_windows(msg.image_abs_path))
            if msg.thumb_abs_path is not None:
                msg_paths.add(path_to_normalized_windows(msg.thumb_abs_path))
            if path_norms & msg_paths:
                return msg
        return None

    def find_unique_message_for_group(self, group_hash: str, pivot_ts: float, window_sec: int) -> Optional[WeChatMessageRef]:
        target = str(group_hash or "").strip().lower()
        if not target:
            return None
        matches = [
            msg
            for msg in self._recent_messages(pivot_ts, lookback_sec=window_sec, lookahead_sec=5, limit=80)
            if (msg.group_hash() or "").strip().lower() == target
        ]
        if not matches:
            return None
        unique: dict[str, WeChatMessageRef] = {}
        for msg in matches:
            key = msg.msg_svr_id or f"{msg.talker}|{msg.create_time}|{msg.thumb_rel_path}|{msg.image_rel_path}"
            unique.setdefault(key, msg)
        if len(unique) != 1:
            return None
        return next(iter(unique.values()))

    def list_image_messages_for_talker(
        self,
        talker: Optional[str],
        start_create_time: float,
        end_create_time: float,
        limit: int = 240,
    ) -> list[WeChatMessageRef]:
        talker_value = str(talker or "").strip()
        if not talker_value:
            return []
        lower = int(min(float(start_create_time or 0.0), float(end_create_time or 0.0)))
        upper = int(max(float(start_create_time or 0.0), float(end_create_time or 0.0)))
        if upper <= 0:
            return []
        if not self.refresh_if_due():
            return []
        if not self.merge_path.exists():
            return []

        conn = sqlite3.connect(str(self.merge_path))
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                """
                SELECT MsgSvrID, StrTalker, CreateTime, BytesExtra
                FROM MSG
                WHERE Type=3
                  AND StrTalker=?
                  AND CreateTime BETWEEN ? AND ?
                ORDER BY CreateTime ASC, MsgSvrID ASC
                LIMIT ?
                """,
                (talker_value, lower, upper, int(max(1, limit))),
            ).fetchall()
        finally:
            conn.close()

        out: list[WeChatMessageRef] = []
        for row in rows:
            image_rel, thumb_rel, sender_user_name = self._extract_media_paths(row["BytesExtra"])
            if not image_rel and not thumb_rel:
                continue
            out.append(
                WeChatMessageRef(
                    msg_svr_id=str(row["MsgSvrID"]) if row["MsgSvrID"] is not None else None,
                    talker=str(row["StrTalker"]) if row["StrTalker"] is not None else None,
                    create_time=float(row["CreateTime"]),
                    sender_user_name=sender_user_name,
                    sender_display=None,
                    image_rel_path=image_rel,
                    thumb_rel_path=thumb_rel,
                    image_abs_path=self._absolute_path_from_rel(image_rel),
                    thumb_abs_path=self._absolute_path_from_rel(thumb_rel),
                )
            )
        return out

    def resolve_contact_display_name(self, username: Optional[str]) -> Optional[str]:
        username = str(username or "").strip()
        if not username:
            return None
        if not self.refresh_if_due():
            return username
        if not self.merge_path.exists():
            return username

        conn = sqlite3.connect(str(self.merge_path))
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                """
                SELECT Remark, NickName, Alias
                FROM Contact
                WHERE UserName=?
                LIMIT 1
                """,
                (username,),
            ).fetchone()
        finally:
            conn.close()

        if row is None:
            return username
        for key in ("Remark", "NickName", "Alias"):
            value = str(row[key] or "").strip()
            if value:
                return value
        return username

    def resolve_talker_display_name(self, talker: str) -> Optional[str]:
        return self.resolve_contact_display_name(talker)


class StateDB:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self._conn.row_factory = sqlite3.Row
        self._lock = threading.RLock()
        self._init_schema()

    def _init_schema(self) -> None:
        with self._lock:
            cur = self._conn.cursor()
            cur.executescript(
                """
                PRAGMA journal_mode=WAL;
                PRAGMA synchronous=NORMAL;

                CREATE TABLE IF NOT EXISTS files (
                    file_id TEXT PRIMARY KEY,
                    path TEXT NOT NULL,
                    source_kind TEXT NOT NULL,
                    ext TEXT NOT NULL,
                    size INTEGER NOT NULL,
                    mtime REAL NOT NULL,
                    ctime REAL NOT NULL,
                    status TEXT NOT NULL,
                    attempts INTEGER NOT NULL DEFAULT 0,
                    next_attempt REAL NOT NULL DEFAULT 0,
                    first_seen REAL NOT NULL,
                    last_seen REAL NOT NULL,
                    msg_svr_id TEXT,
                    talker TEXT,
                    msg_create_time REAL,
                    manual_session_id TEXT,
                    session_release_at REAL NOT NULL DEFAULT 0,
                    processed_at REAL,
                    sha256 TEXT,
                    last_error TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_files_status_next ON files(status, next_attempt);
                CREATE INDEX IF NOT EXISTS idx_files_path ON files(path);

                CREATE TABLE IF NOT EXISTS receipts (
                    file_id TEXT PRIMARY KEY,
                    source_path TEXT NOT NULL,
                    source_kind TEXT NOT NULL,
                    ingested_at REAL NOT NULL,
                    sha256 TEXT NOT NULL,
                    txn_date TEXT,
                    txn_time TEXT,
                    client TEXT,
                    bank TEXT,
                    beneficiary TEXT,
                    amount REAL,
                    currency TEXT,
                    parse_conf REAL NOT NULL,
                    quality_score REAL NOT NULL,
                    ocr_engine TEXT NOT NULL,
                    ocr_conf REAL NOT NULL,
                    ocr_chars INTEGER NOT NULL,
                    review_needed INTEGER NOT NULL,
                    ocr_text TEXT,
                    parser_json TEXT,
                    msg_svr_id TEXT,
                    talker TEXT,
                    msg_create_time REAL,
                    manual_session_id TEXT,
                    resolved_media_path TEXT,
                    resolution_source TEXT,
                    verification_status TEXT,
                    sheet_status TEXT,
                    sheet_payload_json TEXT,
                    sheet_next_attempt REAL,
                    sheet_last_error TEXT,
                    sheet_committed_at REAL,
                    excel_sheet TEXT,
                    excel_row INTEGER
                );
                CREATE TABLE IF NOT EXISTS message_jobs (
                    msg_svr_id TEXT PRIMARY KEY,
                    talker TEXT NOT NULL,
                    talker_display TEXT,
                    thumb_path TEXT,
                    expected_image_path TEXT,
                    create_time REAL NOT NULL,
                    state TEXT NOT NULL,
                    first_seen_at REAL NOT NULL,
                    last_seen_at REAL NOT NULL,
                    ui_force_requested_at REAL,
                    ui_force_completed_at REAL,
                    ui_force_attempts INTEGER NOT NULL DEFAULT 0,
                    next_ui_attempt_at REAL NOT NULL DEFAULT 0,
                    last_ui_result TEXT,
                    batch_id TEXT,
                    manual_session_id TEXT
                );
                CREATE TABLE IF NOT EXISTS manual_sessions (
                    session_id TEXT PRIMARY KEY,
                    talker TEXT NOT NULL,
                    started_at REAL NOT NULL,
                    last_seen_at REAL NOT NULL,
                    release_at REAL NOT NULL,
                    max_release_at REAL NOT NULL,
                    min_create_time REAL,
                    max_create_time REAL,
                    range_seeded INTEGER NOT NULL DEFAULT 0,
                    state TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at REAL NOT NULL
                );
                """
            )
            self._ensure_column_exists(cur, "receipts", "client", "TEXT")
            self._ensure_column_exists(cur, "receipts", "bank", "TEXT")
            self._ensure_column_exists(cur, "receipts", "msg_svr_id", "TEXT")
            self._ensure_column_exists(cur, "receipts", "talker", "TEXT")
            self._ensure_column_exists(cur, "receipts", "msg_create_time", "REAL")
            self._ensure_column_exists(cur, "receipts", "manual_session_id", "TEXT")
            self._ensure_column_exists(cur, "receipts", "resolved_media_path", "TEXT")
            self._ensure_column_exists(cur, "receipts", "resolution_source", "TEXT")
            self._ensure_column_exists(cur, "receipts", "verification_status", "TEXT")
            self._ensure_column_exists(cur, "receipts", "txn_date_source", "TEXT")
            self._ensure_column_exists(cur, "receipts", "txn_time_source", "TEXT")
            self._ensure_column_exists(cur, "receipts", "amount_raw", "TEXT")
            self._ensure_column_exists(cur, "receipts", "amount_rounded", "REAL")
            self._ensure_column_exists(cur, "receipts", "amount_source", "TEXT")
            self._ensure_column_exists(cur, "receipts", "sheet_status", "TEXT")
            self._ensure_column_exists(cur, "receipts", "sheet_payload_json", "TEXT")
            self._ensure_column_exists(cur, "receipts", "sheet_next_attempt", "REAL")
            self._ensure_column_exists(cur, "receipts", "sheet_last_error", "TEXT")
            self._ensure_column_exists(cur, "receipts", "sheet_committed_at", "REAL")
            self._ensure_column_exists(cur, "files", "msg_svr_id", "TEXT")
            self._ensure_column_exists(cur, "files", "talker", "TEXT")
            self._ensure_column_exists(cur, "files", "msg_create_time", "REAL")
            self._ensure_column_exists(cur, "files", "manual_session_id", "TEXT")
            self._ensure_column_exists(cur, "files", "session_release_at", "REAL NOT NULL DEFAULT 0")
            self._ensure_column_exists(cur, "message_jobs", "activation_seen_at", "REAL NOT NULL DEFAULT 0")
            self._ensure_column_exists(cur, "message_jobs", "manual_session_id", "TEXT")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_sha256 ON receipts(sha256)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_msg_svr_id ON receipts(msg_svr_id)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_sheet_status_next ON receipts(sheet_status, sheet_next_attempt, msg_create_time)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_talker_msg_order ON receipts(talker, msg_create_time, msg_svr_id)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_manual_session_order ON receipts(manual_session_id, talker, msg_create_time, msg_svr_id)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_message_jobs_state_next ON message_jobs(state, next_ui_attempt_at, first_seen_at)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_message_jobs_talker_state ON message_jobs(talker, state, create_time DESC)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_message_jobs_expected_path ON message_jobs(expected_image_path)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_message_jobs_manual_session ON message_jobs(manual_session_id, talker, create_time, msg_svr_id)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_files_manual_session_order ON files(manual_session_id, talker, msg_create_time, next_attempt)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_manual_sessions_state_release ON manual_sessions(state, release_at, last_seen_at)")
            cur.execute(
                """
                UPDATE receipts
                SET sheet_status=CASE
                        WHEN sheet_status IS NULL OR sheet_status='' THEN
                            CASE
                                WHEN excel_row IS NOT NULL THEN 'SINK_COMMITTED'
                                ELSE 'SINK_PENDING'
                            END
                        ELSE sheet_status
                    END,
                    sheet_next_attempt=COALESCE(sheet_next_attempt, 0),
                    sheet_committed_at=CASE
                        WHEN sheet_committed_at IS NULL AND excel_row IS NOT NULL THEN ingested_at
                        ELSE sheet_committed_at
                    END
                """
            )
            cur.execute(
                """
                UPDATE message_jobs
                SET activation_seen_at=CASE
                        WHEN COALESCE(activation_seen_at, 0) > 0 THEN activation_seen_at
                        ELSE COALESCE(first_seen_at, create_time, 0)
                    END
                """
            )
            self._conn.commit()

    @staticmethod
    def _ensure_column_exists(cur: sqlite3.Cursor, table_name: str, column_name: str, column_type: str) -> None:
        existing = {str(row[1]).lower() for row in cur.execute(f"PRAGMA table_info({table_name})").fetchall()}
        if column_name.lower() not in existing:
            cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")

    @staticmethod
    def compute_file_id(path: Path, stat: os.stat_result) -> str:
        payload = f"{str(path).lower()}|{stat.st_size}|{stat.st_mtime_ns}"
        return hashlib.sha1(payload.encode("utf-8")).hexdigest()

    def upsert_candidate(
        self,
        path: Path,
        settle_seconds: int,
        source_event: str,
        thumb_candidates_enabled: bool,
    ) -> Optional[str]:
        if not is_candidate(path, thumb_candidates_enabled=thumb_candidates_enabled):
            return None
        try:
            st = path.stat()
        except FileNotFoundError:
            return None
        if st.st_size <= 0:
            return None

        now = time.time()
        file_id = self.compute_file_id(path, st)
        source_kind = detect_source_kind(path)
        ext = path.suffix.lower()
        next_attempt = now + candidate_initial_delay_seconds(source_kind, settle_seconds, thumb_candidates_enabled)
        refresh_manual_session = should_refresh_manual_session(source_kind, source_event)
        candidate_id: Optional[str] = None

        with self._lock:
            cur = self._conn.cursor()
            existing = cur.execute(
                """
                SELECT file_id, status
                FROM files
                WHERE path=?
                ORDER BY last_seen DESC
                LIMIT 1
                """,
                (str(path),),
                ).fetchone()
            if existing is not None and existing["status"] in ("pending", "retry", "processing"):
                cur.execute(
                    """
                    UPDATE files
                    SET last_seen=?, mtime=?, size=?, next_attempt=MIN(next_attempt, ?)
                    WHERE file_id=?
                    """,
                    (
                        float(now),
                        float(st.st_mtime),
                        int(st.st_size),
                        float(next_attempt),
                        existing["file_id"],
                    ),
                )
                self._conn.commit()
                candidate_id = str(existing["file_id"])
            else:
                cur.execute(
                    """
                    INSERT INTO files(file_id, path, source_kind, ext, size, mtime, ctime, status, attempts, next_attempt, first_seen, last_seen)
                    VALUES(?, ?, ?, ?, ?, ?, ?, 'pending', 0, ?, ?, ?)
                    ON CONFLICT(file_id) DO UPDATE SET
                        last_seen=excluded.last_seen,
                        mtime=excluded.mtime,
                        size=excluded.size,
                        next_attempt=CASE
                            WHEN files.status IN ('done', 'duplicate') THEN files.next_attempt
                            ELSE MIN(files.next_attempt, excluded.next_attempt)
                        END
                    """,
                    (
                        file_id,
                        str(path),
                        source_kind,
                        ext,
                        int(st.st_size),
                        float(st.st_mtime),
                        float(st.st_ctime),
                        float(next_attempt),
                        float(now),
                        float(now),
                    ),
                )
                self._conn.commit()
                candidate_id = file_id

        if refresh_manual_session:
            self.start_manual_session(now)
        return candidate_id

    def get_file(self, file_id: Optional[str]) -> Optional[sqlite3.Row]:
        file_id_value = str(file_id or "").strip()
        if not file_id_value:
            return None
        with self._lock:
            return self._conn.execute(
                """
                SELECT *
                FROM files
                WHERE file_id=?
                LIMIT 1
                """,
                (file_id_value,),
            ).fetchone()

    def update_file_message_context(
        self,
        file_id: Optional[str],
        *,
        msg_svr_id: Optional[str],
        talker: Optional[str],
        msg_create_time: float,
        manual_session_id: Optional[str],
        session_release_at: float,
    ) -> None:
        file_id_value = str(file_id or "").strip()
        if not file_id_value:
            return
        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET msg_svr_id=COALESCE(NULLIF(?, ''), msg_svr_id),
                    talker=COALESCE(NULLIF(?, ''), talker),
                    msg_create_time=CASE
                        WHEN ? > 0 THEN ?
                        ELSE msg_create_time
                    END,
                    manual_session_id=COALESCE(NULLIF(?, ''), manual_session_id),
                    session_release_at=CASE
                        WHEN ? > COALESCE(session_release_at, 0) THEN ?
                        ELSE COALESCE(session_release_at, 0)
                    END
                WHERE file_id=?
                """,
                (
                    str(msg_svr_id or "").strip(),
                    str(talker or "").strip(),
                    float(msg_create_time or 0.0),
                    float(msg_create_time or 0.0),
                    str(manual_session_id or "").strip(),
                    float(session_release_at or 0.0),
                    float(session_release_at or 0.0),
                    file_id_value,
                ),
            )
            self._conn.commit()

    def get_current_manual_session_id(self) -> Optional[str]:
        raw = self.get_meta(MANUAL_SESSION_ID_META_KEY)
        value = str(raw or "").strip()
        return value or None

    def _set_current_manual_session_id_locked(self, cur: sqlite3.Cursor, session_id: Optional[str], now: float) -> None:
        cur.execute(
            """
            INSERT INTO meta(key, value, updated_at)
            VALUES(?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
            """,
            (MANUAL_SESSION_ID_META_KEY, str(session_id or "").strip(), float(now)),
        )

    def _release_manual_session_file_holds_locked(self, cur: sqlite3.Cursor, session_id: str, reason: str, now: float) -> int:
        cur.execute(
            """
            UPDATE files
            SET status='retry',
                next_attempt=?,
                last_error=?
            WHERE manual_session_id=?
              AND status IN ('pending', 'retry')
              AND (
                    last_error LIKE 'WAITING_SESSION_PRIOR_MESSAGE_ORDER:%'
                 OR last_error LIKE 'WAITING_PRIOR_SINK_SESSION_MESSAGE:%'
                 OR last_error LIKE 'WAITING_PRIOR_SINK_RECEIPT:%'
              )
            """,
            (float(now), reason[:1200], str(session_id)),
        )
        return int(cur.rowcount or 0)

    def _rollover_manual_sessions_locked(
        self,
        cur: sqlite3.Cursor,
        session_ids: list[str],
        reason: str,
        now: float,
    ) -> tuple[int, int]:
        ignored_placeholders = 0
        released_files = 0
        for session_id in session_ids:
            if not session_id:
                continue
            cur.execute(
                """
                UPDATE manual_sessions
                SET state='ROLLED', last_seen_at=?
                WHERE session_id=?
                """,
                (float(now), str(session_id)),
            )
            cur.execute(
                """
                UPDATE message_jobs
                SET state=?,
                    last_seen_at=?,
                    last_ui_result=?
                WHERE manual_session_id=?
                  AND state=?
                """,
                (reason, float(now), reason[:1200], str(session_id), SESSION_PENDING_OPEN_STATE),
            )
            ignored_placeholders += int(cur.rowcount or 0)
            released_files += self._release_manual_session_file_holds_locked(cur, str(session_id), reason, now)
        return ignored_placeholders, released_files

    def get_manual_session(self, session_id: Optional[str]) -> Optional[sqlite3.Row]:
        session_value = str(session_id or "").strip()
        if not session_value:
            return None
        with self._lock:
            return self._conn.execute(
                """
                SELECT *
                FROM manual_sessions
                WHERE session_id=?
                LIMIT 1
                """,
                (session_value,),
            ).fetchone()

    def start_or_extend_manual_order_session(
        self,
        *,
        talker: Optional[str],
        create_time: float,
        event_ts: float,
        burst_gap_seconds: int,
        burst_max_seconds: int,
        preferred_session_id: Optional[str] = None,
    ) -> Optional[sqlite3.Row]:
        talker_value = str(talker or "").strip()
        if not talker_value:
            return None

        now = float(event_ts or time.time())
        gap_seconds = max(1, int(burst_gap_seconds))
        max_seconds = max(gap_seconds, int(burst_max_seconds))
        preferred_value = str(preferred_session_id or "").strip()
        result_session_id: Optional[str] = None

        with self._lock:
            cur = self._conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            session_row: Optional[sqlite3.Row] = None
            if preferred_value:
                session_row = cur.execute(
                    """
                    SELECT *
                    FROM manual_sessions
                    WHERE session_id=? AND state='ACTIVE'
                    LIMIT 1
                    """,
                    (preferred_value,),
                ).fetchone()

            if session_row is None:
                active_same = cur.execute(
                    """
                    SELECT *
                    FROM manual_sessions
                    WHERE talker=? AND state='ACTIVE'
                    ORDER BY started_at DESC
                    LIMIT 1
                    """,
                    (talker_value,),
                ).fetchone()
                if active_same is not None:
                    still_extendable = (
                        int(active_same["range_seeded"] or 0) == 0
                        and now <= float(active_same["max_release_at"] or 0.0)
                        and now <= float(active_same["last_seen_at"] or 0.0) + float(gap_seconds)
                    )
                    session_row = active_same if still_extendable else None
                    if active_same is not None and session_row is None:
                        self._rollover_manual_sessions_locked(
                            cur,
                            [str(active_same["session_id"])],
                            IGNORED_SESSION_ROLLOVER_STATE,
                            now,
                        )

            if session_row is None:
                other_active_ids = [
                    str(row["session_id"])
                    for row in cur.execute(
                        """
                        SELECT session_id
                        FROM manual_sessions
                        WHERE state='ACTIVE'
                        """
                    ).fetchall()
                ]
                if other_active_ids:
                    self._rollover_manual_sessions_locked(cur, other_active_ids, IGNORED_SESSION_ROLLOVER_STATE, now)

                session_id = hashlib.sha1(f"{talker_value}|{now:.6f}".encode("utf-8")).hexdigest()[:16]
                release_at = now + float(gap_seconds)
                max_release_at = now + float(max_seconds)
                min_create = float(create_time) if float(create_time or 0.0) > 0 else None
                max_create = float(create_time) if float(create_time or 0.0) > 0 else None
                cur.execute(
                    """
                    INSERT INTO manual_sessions(
                        session_id, talker, started_at, last_seen_at, release_at, max_release_at,
                        min_create_time, max_create_time, range_seeded, state
                    )
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, 0, 'ACTIVE')
                    """,
                    (
                        session_id,
                        talker_value,
                        float(now),
                        float(now),
                        float(release_at),
                        float(max_release_at),
                        min_create,
                        max_create,
                    ),
                )
                self._set_current_manual_session_id_locked(cur, session_id, now)
                self._conn.commit()
                result_session_id = session_id
            else:
                session_id = str(session_row["session_id"])
                release_at = float(session_row["release_at"] or 0.0)
                if int(session_row["range_seeded"] or 0) == 0:
                    release_at = min(
                        float(session_row["max_release_at"] or now),
                        max(float(session_row["release_at"] or 0.0), now + float(gap_seconds)),
                    )
                min_create = float(session_row["min_create_time"] or 0.0)
                max_create = float(session_row["max_create_time"] or 0.0)
                create_value = float(create_time or 0.0)
                if create_value > 0:
                    min_create = create_value if min_create <= 0 else min(min_create, create_value)
                    max_create = create_value if max_create <= 0 else max(max_create, create_value)
                cur.execute(
                    """
                    UPDATE manual_sessions
                    SET last_seen_at=?,
                        release_at=?,
                        min_create_time=?,
                        max_create_time=?
                    WHERE session_id=?
                    """,
                    (
                        float(now),
                        float(release_at),
                        min_create if min_create > 0 else None,
                        max_create if max_create > 0 else None,
                        session_id,
                    ),
                )
                self._set_current_manual_session_id_locked(cur, session_id, now)
                self._conn.commit()
                result_session_id = session_id

        return self.get_manual_session(result_session_id)

    def list_manual_sessions_ready_for_seed(self, now: Optional[float] = None) -> list[sqlite3.Row]:
        moment = time.time() if now is None else float(now)
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT *
                FROM manual_sessions
                WHERE state='ACTIVE'
                  AND range_seeded=0
                  AND release_at <= ?
                ORDER BY started_at ASC
                """,
                (moment,),
            ).fetchall()
        return list(rows)

    def mark_manual_session_seeded(self, session_id: Optional[str]) -> None:
        session_value = str(session_id or "").strip()
        if not session_value:
            return
        with self._lock:
            self._conn.execute(
                """
                UPDATE manual_sessions
                SET range_seeded=1
                WHERE session_id=?
                """,
                (session_value,),
            )
            self._conn.commit()

    def ignore_stale_manual_sessions(self, max_age_sec: int = 1800) -> tuple[int, int]:
        threshold = time.time() - max(60, int(max_age_sec))
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            session_ids = [
                str(row["session_id"])
                for row in cur.execute(
                    """
                    SELECT session_id
                    FROM manual_sessions
                    WHERE state='ACTIVE'
                      AND last_seen_at <= ?
                    """,
                    (float(threshold),),
                ).fetchall()
            ]
            ignored_placeholders, released_files = self._rollover_manual_sessions_locked(
                cur,
                session_ids,
                IGNORED_STALE_MANUAL_SESSION_STATE,
                now,
            )
            self._conn.commit()
            return ignored_placeholders, released_files

    def get_meta(self, key: str) -> Optional[str]:
        with self._lock:
            row = self._conn.execute("SELECT value FROM meta WHERE key=? LIMIT 1", (key,)).fetchone()
            return None if row is None else str(row["value"])

    def get_meta_float(self, key: str) -> Optional[float]:
        raw = self.get_meta(key)
        if raw is None:
            return None
        try:
            return float(raw)
        except Exception:
            return None

    def set_meta(self, key: str, value: str) -> None:
        now = time.time()
        with self._lock:
            self._conn.execute(
                """
                INSERT INTO meta(key, value, updated_at)
                VALUES(?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value=excluded.value,
                    updated_at=excluded.updated_at
                """,
                (key, value, float(now)),
            )
            self._conn.commit()

    def get_manual_session_started_at(self) -> Optional[float]:
        return self.get_meta_float(MANUAL_SESSION_META_KEY)

    def start_manual_session(self, started_at: Optional[float] = None) -> float:
        ts = float(started_at if started_at is not None else time.time())
        self.set_meta(MANUAL_SESSION_META_KEY, f"{ts:.6f}")
        return ts

    @staticmethod
    def _parse_bool_text(value: Any, default: bool = False) -> bool:
        if value is None:
            return default
        text = str(value).strip().lower()
        if not text:
            return default
        if text in {"1", "true", "yes", "y", "on"}:
            return True
        if text in {"0", "false", "no", "n", "off"}:
            return False
        return default

    def is_ui_force_runtime_enabled(self, default_enabled: bool = False) -> bool:
        if not default_enabled:
            return False
        raw = self.get_meta(UI_FORCE_RUNTIME_META_KEY)
        if raw is None:
            return True
        return self._parse_bool_text(raw, default=False)

    def set_ui_force_runtime_enabled(self, enabled: bool, release_waiting: bool = True) -> tuple[int, int]:
        now = time.time()
        released_jobs = 0
        requeued_files = 0
        with self._lock:
            cur = self._conn.cursor()
            cur.execute(
                """
                INSERT INTO meta(key, value, updated_at)
                VALUES(?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value=excluded.value,
                    updated_at=excluded.updated_at
                """,
                (UI_FORCE_RUNTIME_META_KEY, "1" if enabled else "0", float(now)),
            )
            if not enabled:
                cur.execute(
                    """
                    INSERT INTO meta(key, value, updated_at)
                    VALUES(?, ?, ?)
                    ON CONFLICT(key) DO UPDATE SET
                        value=excluded.value,
                        updated_at=excluded.updated_at
                    """,
                    (MANUAL_SESSION_META_KEY, f"{float(now):.6f}", float(now)),
                )
            if not enabled and release_waiting:
                released_jobs = int(
                    cur.execute(
                        """
                        UPDATE message_jobs
                        SET state='WAITING_ORIGINAL',
                            batch_id=NULL,
                            next_ui_attempt_at=0,
                            last_seen_at=?,
                            last_ui_result=CASE
                                WHEN last_ui_result IS NULL OR last_ui_result=''
                                THEN 'UI_FORCE_DISABLED_MANUAL_MODE'
                                ELSE last_ui_result
                            END
                        WHERE state IN ('UI_FORCE_PENDING', 'UI_FORCE_RUNNING')
                        """,
                        (float(now),),
                    ).rowcount
                    or 0
                )
                requeued_files = int(
                    cur.execute(
                        """
                        UPDATE files
                        SET status=CASE
                                WHEN status='processing' THEN 'processing'
                                ELSE 'retry'
                            END,
                            next_attempt=CASE
                                WHEN next_attempt > ? THEN ?
                                ELSE next_attempt
                            END,
                            last_error='WAITING_ORIGINAL_MEDIA'
                        WHERE status IN ('pending', 'retry', 'processing')
                          AND last_error='WAITING_UI_FORCE_DOWNLOAD'
                        """,
                        (float(now + 3), float(now + 3)),
                    ).rowcount
                    or 0
                )
            self._conn.commit()
        return released_jobs, requeued_files

    def ignore_stale_queue(self, older_than_mtime: float) -> int:
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            count = int(
                cur.execute(
                    """
                    SELECT COUNT(*)
                    FROM files
                    WHERE status IN ('pending', 'retry', 'processing')
                      AND mtime < ?
                    """,
                    (float(older_than_mtime),),
                ).fetchone()[0]
            )
            if count <= 0:
                return 0
            cur.execute(
                """
                UPDATE files
                SET status='ignored',
                    processed_at=?,
                    next_attempt=0,
                    last_error=CASE
                        WHEN last_error IS NULL OR last_error=''
                            THEN 'IGNORED_OLD_BACKLOG'
                        ELSE last_error
                    END
                WHERE status IN ('pending', 'retry', 'processing')
                  AND mtime < ?
                """,
                (float(now), float(older_than_mtime)),
            )
            self._conn.commit()
            return count

    def ignore_manual_open_only_waits(self) -> int:
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            count = int(
                cur.execute(
                    """
                    SELECT COUNT(*)
                    FROM files
                    WHERE status IN ('pending', 'retry', 'processing')
                      AND source_kind IN ('msgattach_thumb_dat', 'temp_image')
                      AND (
                        last_error IN (?, ?, ?)
                        OR last_error LIKE 'WAITING_PRIOR_MESSAGE_ORDER:%'
                        OR last_error LIKE 'WAITING_SESSION_PRIOR_MESSAGE_ORDER:%'
                      )
                    """,
                    MANUAL_OPEN_ONLY_WAIT_REASONS,
                ).fetchone()[0]
            )
            if count <= 0:
                return 0
            cur.execute(
                """
                UPDATE files
                SET status='ignored',
                    processed_at=?,
                    next_attempt=0,
                    last_error=?
                WHERE status IN ('pending', 'retry', 'processing')
                  AND source_kind IN ('msgattach_thumb_dat', 'temp_image')
                  AND (
                    last_error IN (?, ?, ?)
                    OR last_error LIKE 'WAITING_PRIOR_MESSAGE_ORDER:%'
                    OR last_error LIKE 'WAITING_SESSION_PRIOR_MESSAGE_ORDER:%'
                  )
                """,
                (float(now), MANUAL_OPEN_ONLY_IGNORE_REASON, *MANUAL_OPEN_ONLY_WAIT_REASONS),
            )
            self._conn.commit()
            return count

    def find_recent_msgattach_context_path(
        self,
        pivot_mtime: float,
        lookback_sec: int = 90,
        lookahead_sec: int = 15,
        limit: int = 24,
    ) -> Optional[str]:
        lower = float(pivot_mtime) - max(5, int(lookback_sec))
        upper = float(pivot_mtime) + max(1, int(lookahead_sec))
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT path, mtime
                FROM files
                WHERE source_kind IN ('msgattach_image_dat', 'msgattach_thumb_dat', 'msgattach_image_plain')
                  AND mtime BETWEEN ? AND ?
                ORDER BY ABS(mtime - ?) ASC, mtime DESC
                LIMIT ?
                """,
                (lower, upper, float(pivot_mtime), int(max(1, limit))),
            ).fetchall()

        if not rows:
            return None

        unique_groups: dict[str, str] = {}
        for row in rows:
            candidate = str(row["path"])
            gid = extract_group_id_from_path(Path(candidate))
            if not gid:
                continue
            key = gid.strip().lower()
            unique_groups.setdefault(key, candidate)

        if len(unique_groups) == 1:
            return next(iter(unique_groups.values()))
        return None

    def get_latest_file_row_by_path(self, path: Path | str) -> Optional[sqlite3.Row]:
        with self._lock:
            row = self._conn.execute(
                """
                SELECT file_id, path, source_kind, status, attempts, first_seen, mtime, ctime, next_attempt, last_error
                FROM files
                WHERE path=?
                ORDER BY mtime DESC, last_seen DESC
                LIMIT 1
                """,
                (str(path),),
            ).fetchone()
            return row

    def find_recent_unresolved_msgattach_context_path(
        self,
        max_age_sec: int = 1800,
        limit: int = 40,
    ) -> Optional[str]:
        threshold = time.time() - max(60, int(max_age_sec))
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT path, mtime, last_error
                FROM files
                WHERE source_kind IN ('msgattach_image_dat', 'msgattach_thumb_dat', 'msgattach_image_plain')
                  AND mtime >= ?
                  AND (
                    status IN ('retry', 'processing')
                    OR status='exception'
                    OR last_error IN ('WAITING_ORIGINAL_MEDIA', 'EXCEPTION_MISSING_CORE_FIELDS')
                  )
                ORDER BY mtime DESC
                LIMIT ?
                """,
                (float(threshold), int(max(1, limit))),
            ).fetchall()

        if not rows:
            return None

        unique_groups: dict[str, str] = {}
        for row in rows:
            candidate = str(row["path"])
            gid = extract_group_id_from_path(Path(candidate))
            if not gid:
                continue
            key = gid.strip().lower()
            unique_groups.setdefault(key, candidate)

        if len(unique_groups) == 1:
            return next(iter(unique_groups.values()))
        return None

    def claim_next(
        self,
        manual_session_started_at: Optional[float] = None,
        manual_session_id: Optional[str] = None,
    ) -> Optional[QueueItem]:
        now = time.time()
        session_floor = float(manual_session_started_at or 0.0)
        session_id_value = str(manual_session_id or "").strip()
        with self._lock:
            cur = self._conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            row = cur.execute(
                """
                SELECT f.file_id, f.path, f.source_kind, f.ext, f.size, f.mtime, f.first_seen, f.attempts,
                       f.msg_svr_id, f.talker, f.msg_create_time, f.manual_session_id, f.session_release_at
                FROM files AS f
                WHERE f.status IN ('pending', 'retry')
                  AND f.next_attempt <= ?
                  AND COALESCE(f.session_release_at, 0) <= ?
                ORDER BY
                    CASE
                        WHEN ? <> '' AND COALESCE(f.manual_session_id, '') = ? THEN 0
                        WHEN ? <> '' THEN 1
                        ELSE 0
                    END ASC,
                    CASE
                        WHEN ? > 0 AND COALESCE(f.first_seen, 0) >= ? THEN 0
                        WHEN ? > 0 THEN 1
                        ELSE 0
                    END ASC,
                    CASE f.source_kind
                        WHEN 'msgattach_image_dat' THEN 0
                        WHEN 'msgattach_image_plain' THEN 1
                        WHEN 'temp_image' THEN 2
                        WHEN 'msgattach_thumb_dat' THEN 3
                        ELSE 4
                    END ASC,
                    CASE WHEN f.msg_create_time IS NOT NULL AND f.msg_create_time > 0 THEN 0 ELSE 1 END ASC,
                    CASE WHEN f.msg_create_time IS NOT NULL AND f.msg_create_time > 0 THEN f.msg_create_time END ASC,
                    CASE WHEN f.msg_create_time IS NOT NULL AND f.msg_create_time > 0 THEN f.msg_svr_id END ASC,
                    CASE WHEN f.msg_create_time IS NULL OR f.msg_create_time <= 0 THEN f.first_seen END ASC,
                    CASE WHEN f.msg_create_time IS NULL OR f.msg_create_time <= 0 THEN f.mtime END ASC,
                    f.next_attempt ASC
                LIMIT 1
                """,
                (
                    float(now),
                    float(now),
                    session_id_value,
                    session_id_value,
                    session_id_value,
                    session_floor,
                    session_floor,
                    session_floor,
                ),
            ).fetchone()
            if row is None:
                self._conn.commit()
                return None
            next_attempt_count = int(row["attempts"]) + 1
            cur.execute(
                """
                UPDATE files
                SET status='processing', attempts=?, last_error=NULL
                WHERE file_id=?
                """,
                (next_attempt_count, row["file_id"]),
            )
            self._conn.commit()
            return QueueItem(
                file_id=row["file_id"],
                path=row["path"],
                source_kind=row["source_kind"],
                ext=row["ext"],
                size=int(row["size"]),
                mtime=float(row["mtime"]),
                first_seen=float(row["first_seen"]),
                attempts=next_attempt_count,
                msg_svr_id=str(row["msg_svr_id"] or "").strip() or None,
                talker=str(row["talker"] or "").strip() or None,
                msg_create_time=float(row["msg_create_time"] or 0.0),
                manual_session_id=str(row["manual_session_id"] or "").strip() or None,
                session_release_at=float(row["session_release_at"] or 0.0),
            )

    def mark_done(self, file_id: str, sha256: str, processed_at: float, note: Optional[str] = None) -> None:
        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status='done', processed_at=?, sha256=?, last_error=?
                WHERE file_id=?
                """,
                (processed_at, sha256, note[:1200] if note else None, file_id),
            )
            self._conn.commit()

    def resolve_related_file_paths(
        self,
        source_path: Path | str,
        exclude_file_id: str,
        sha256: str = "",
        reason: str = "RESOLVED_BY_LATER_SUCCESS",
    ) -> int:
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            cur.execute(
                """
                UPDATE files
                SET status='done',
                    processed_at=?,
                    sha256=CASE WHEN sha256 IS NULL OR sha256='' THEN ? ELSE sha256 END,
                    last_error=?
                WHERE path=?
                  AND file_id<>?
                  AND status IN ('pending', 'retry', 'processing', 'exception')
                """,
                (float(now), sha256, reason[:1200], str(source_path), str(exclude_file_id)),
            )
            changed = int(cur.rowcount or 0)
            self._conn.commit()
            return changed

    def cleanup_stale_temp_orphans(self, max_age_sec: int = 600) -> int:
        threshold = time.time() - max(60, int(max_age_sec))
        now = time.time()
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT file_id, path
                FROM files
                WHERE source_kind='temp_image'
                  AND status IN ('pending', 'retry', 'processing', 'exception')
                  AND last_seen <= ?
                ORDER BY last_seen ASC
                """,
                (float(threshold),),
            ).fetchall()

            stale_ids: list[str] = []
            for row in rows:
                if not Path(str(row["path"])).exists():
                    stale_ids.append(str(row["file_id"]))

            if not stale_ids:
                return 0

            self._conn.executemany(
                """
                UPDATE files
                SET status='done',
                    processed_at=?,
                    last_error='STALE_TEMP_ORPHAN'
                WHERE file_id=?
                """,
                [(float(now), file_id) for file_id in stale_ids],
            )
            self._conn.commit()
            return len(stale_ids)

    def recover_stale_processing(self, max_age_sec: int = 180) -> tuple[int, int]:
        threshold = time.time() - max(30, int(max_age_sec))
        now = time.time()
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT f.file_id,
                       f.processed_at,
                       r.file_id AS receipt_file_id,
                       r.ingested_at
                FROM files AS f
                LEFT JOIN receipts AS r
                  ON r.file_id = f.file_id
                WHERE f.status='processing'
                  AND COALESCE(f.last_seen, 0) <= ?
                ORDER BY f.last_seen ASC
                """,
                (float(threshold),),
            ).fetchall()

            if not rows:
                return (0, 0)

            done_rows: list[tuple[float, str, str]] = []
            retry_rows: list[tuple[float, str, str]] = []
            for row in rows:
                file_id = str(row["file_id"])
                if row["receipt_file_id"] is not None:
                    processed_at = float(row["processed_at"] or row["ingested_at"] or now)
                    done_rows.append((processed_at, "RECOVERED_PROCESSING_WITH_RECEIPT", file_id))
                else:
                    retry_rows.append((float(now), "RECOVERED_STALE_PROCESSING", file_id))

            if done_rows:
                self._conn.executemany(
                    """
                    UPDATE files
                    SET status='done',
                        processed_at=?,
                        last_error=?
                    WHERE file_id=?
                    """,
                    done_rows,
                )
            if retry_rows:
                self._conn.executemany(
                    """
                    UPDATE files
                    SET status='retry',
                        next_attempt=?,
                        last_error=?
                    WHERE file_id=?
                    """,
                    retry_rows,
                )
            self._conn.commit()
            return (len(retry_rows), len(done_rows))

    def mark_retry(
        self,
        file_id: str,
        attempts: int,
        retry_base_sec: int,
        err: str,
        max_retries: int,
        delay_override_sec: Optional[int] = None,
    ) -> None:
        if max_retries > 0 and attempts >= max_retries:
            status = "failed"
            next_attempt = 0.0
        else:
            status = "retry"
            if delay_override_sec is not None:
                backoff = max(3, int(delay_override_sec))
            else:
                backoff = min(retry_base_sec * (2 ** max(0, attempts - 1)), 3600)
            next_attempt = time.time() + float(backoff)

        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status=?, next_attempt=?, last_error=?
                WHERE file_id=?
                """,
                (status, next_attempt, err[:1200], file_id),
            )
            self._conn.commit()

    def mark_hold(self, file_id: str, reason: str, delay_sec: int = 120) -> None:
        next_attempt = time.time() + max(3, delay_sec)
        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status='retry', next_attempt=?, last_error=?
                WHERE file_id=?
                """,
                (next_attempt, reason[:1200], file_id),
            )
            self._conn.commit()

    def mark_exception(self, file_id: str, reason: str) -> None:
        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status='exception', next_attempt=0, processed_at=?, last_error=?
                WHERE file_id=?
                """,
                (float(time.time()), reason[:1200], file_id),
            )
            self._conn.commit()

    def receipt_exists(self, file_id: str) -> bool:
        with self._lock:
            row = self._conn.execute("SELECT 1 FROM receipts WHERE file_id=? LIMIT 1", (file_id,)).fetchone()
            return row is not None

    def receipt_msg_exists(self, msg_svr_id: Optional[str]) -> bool:
        if not msg_svr_id:
            return False
        with self._lock:
            row = self._conn.execute("SELECT 1 FROM receipts WHERE msg_svr_id=? LIMIT 1", (str(msg_svr_id),)).fetchone()
            return row is not None

    def requeue_mapped_missing_client(
        self,
        resolver: "ClientResolver",
        max_age_hours: int = 3,
        limit: int = 1200,
    ) -> int:
        """
        Requeue files that were previously marked with MISSING_CLIENT_MAP but now
        have a valid client mapping.
        """
        threshold = time.time() - max(1, int(max_age_hours)) * 3600
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            rows = cur.execute(
                """
                SELECT file_id, path
                FROM files
                WHERE status='done'
                  AND last_error LIKE 'MISSING_CLIENT_MAP:%'
                  AND mtime >= ?
                ORDER BY mtime DESC
                LIMIT ?
                """,
                (float(threshold), int(max(1, limit))),
            ).fetchall()

            to_requeue: list[str] = []
            for row in rows:
                p = Path(row["path"])
                if resolver.resolve(p):
                    to_requeue.append(str(row["file_id"]))

            if not to_requeue:
                return 0

            cur.executemany(
                """
                UPDATE files
                SET status='retry', attempts=0, next_attempt=?, processed_at=NULL, last_error=NULL
                WHERE file_id=?
                """,
                [(float(now), fid) for fid in to_requeue],
            )
            self._conn.commit()
            return len(to_requeue)

    def backfill_receipt_context(self, resolver: "ClientResolver", limit: int = 5000) -> int:
        with self._lock:
            cur = self._conn.cursor()
            rows = cur.execute(
                """
                SELECT file_id, source_path, client, bank, beneficiary, ocr_text
                FROM receipts
                WHERE client IS NULL OR bank IS NULL
                LIMIT ?
                """,
                (int(max(1, limit)),),
            ).fetchall()

            updates: list[tuple[Optional[str], Optional[str], str]] = []
            for row in rows:
                source_path = Path(str(row["source_path"]))
                client = row["client"] if row["client"] else resolver.resolve(source_path)
                bank = row["bank"]
                if not bank:
                    bank = detect_bank(f"{row['ocr_text'] or ''}\n{client or ''}", row["beneficiary"])
                if client != row["client"] or bank != row["bank"]:
                    updates.append((client, bank, str(row["file_id"])))

            if not updates:
                return 0

            cur.executemany(
                """
                UPDATE receipts
                SET client=?, bank=?
                WHERE file_id=?
                """,
                updates,
            )
            self._conn.commit()
            return len(updates)

    def receipt_sha_exists(self, sha256: str) -> bool:
        if not sha256:
            return False
        with self._lock:
            row = self._conn.execute("SELECT 1 FROM receipts WHERE sha256=? LIMIT 1", (sha256,)).fetchone()
            return row is not None

    def insert_receipt(self, payload: dict[str, Any]) -> None:
        with self._lock:
            self._conn.execute(
                """
                INSERT OR REPLACE INTO receipts(
                    file_id, source_path, source_kind, ingested_at, sha256,
                    txn_date, txn_time, client, bank, beneficiary, amount, currency,
                    parse_conf, quality_score, ocr_engine, ocr_conf, ocr_chars,
                    review_needed, ocr_text, parser_json, msg_svr_id, talker, msg_create_time, manual_session_id,
                    resolved_media_path, resolution_source, verification_status,
                    txn_date_source, txn_time_source, amount_raw, amount_rounded, amount_source,
                    sheet_status, sheet_payload_json, sheet_next_attempt, sheet_last_error, sheet_committed_at,
                    excel_sheet, excel_row
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    payload["file_id"],
                    payload["source_path"],
                    payload["source_kind"],
                    payload["ingested_at"],
                    payload["sha256"],
                    payload.get("txn_date"),
                    payload.get("txn_time"),
                    payload.get("client"),
                    payload.get("bank"),
                    payload.get("beneficiary"),
                    payload.get("amount"),
                    payload.get("currency"),
                    payload["parse_conf"],
                    payload["quality_score"],
                    payload["ocr_engine"],
                    payload["ocr_conf"],
                    payload["ocr_chars"],
                    1 if payload["review_needed"] else 0,
                    payload.get("ocr_text"),
                    payload.get("parser_json"),
                    payload.get("msg_svr_id"),
                    payload.get("talker"),
                    payload.get("msg_create_time"),
                    payload.get("manual_session_id"),
                    payload.get("resolved_media_path"),
                    payload.get("resolution_source"),
                    payload.get("verification_status"),
                    payload.get("txn_date_source"),
                    payload.get("txn_time_source"),
                    payload.get("amount_raw"),
                    payload.get("amount_rounded"),
                    payload.get("amount_source"),
                    payload.get("sheet_status"),
                    payload.get("sheet_payload_json"),
                    payload.get("sheet_next_attempt"),
                    payload.get("sheet_last_error"),
                    payload.get("sheet_committed_at"),
                    payload.get("excel_sheet"),
                    payload.get("excel_row"),
                ),
            )
            self._conn.commit()

    @staticmethod
    def _receipt_message_sort_key(msg_svr_id: Optional[str], file_id: str) -> str:
        msg_value = str(msg_svr_id or "").strip()
        if msg_value:
            return msg_value
        return f"file:{str(file_id)}"

    def find_prior_pending_sink_receipt(
        self,
        talker: Optional[str],
        msg_create_time: float,
        msg_svr_id: Optional[str],
        file_id: str,
        manual_session_started_at: Optional[float] = None,
        manual_session_id: Optional[str] = None,
    ) -> Optional[sqlite3.Row]:
        talker_value = str(talker or "").strip()
        if not talker_value or msg_create_time <= 0:
            return None
        sort_key = self._receipt_message_sort_key(msg_svr_id, file_id)
        session_floor = float(manual_session_started_at or 0.0)
        session_id_value = str(manual_session_id or "").strip()
        with self._lock:
            return self._conn.execute(
                """
                SELECT file_id, msg_svr_id, msg_create_time, sheet_status, ingested_at, manual_session_id
                FROM receipts
                WHERE talker=?
                  AND file_id<>?
                  AND msg_create_time IS NOT NULL
                  AND msg_create_time > 0
                  AND COALESCE(sheet_status, '') NOT IN ('SINK_COMMITTED', 'SINK_SKIPPED_TERMINAL')
                  AND (? = '' OR COALESCE(manual_session_id, '') = ?)
                  AND (? <= 0 OR COALESCE(ingested_at, 0) >= ?)
                  AND (
                    msg_create_time < ?
                    OR (
                      msg_create_time = ?
                      AND COALESCE(NULLIF(msg_svr_id, ''), 'file:' || file_id) < ?
                    )
                  )
                ORDER BY msg_create_time ASC, COALESCE(NULLIF(msg_svr_id, ''), 'file:' || file_id) ASC
                LIMIT 1
                """,
                (
                    talker_value,
                    str(file_id),
                    session_id_value,
                    session_id_value,
                    session_floor,
                    session_floor,
                    float(msg_create_time),
                    float(msg_create_time),
                    sort_key,
                ),
            ).fetchone()

    def claim_next_sink_receipt(
        self,
        sheet_order_scope: str = "per_talker",
        commit_order: str = "asc",
        manual_session_started_at: Optional[float] = None,
        manual_session_id: Optional[str] = None,
    ) -> Optional[dict[str, Any]]:
        now = time.time()
        scope = str(sheet_order_scope or "per_talker").strip().lower() or "per_talker"
        order = str(commit_order or "asc").strip().lower() or "asc"
        order_by = "r.msg_create_time ASC, COALESCE(NULLIF(r.msg_svr_id, ''), 'file:' || r.file_id) ASC"
        if order == "desc":
            order_by = "r.msg_create_time DESC, COALESCE(NULLIF(r.msg_svr_id, ''), 'file:' || r.file_id) DESC"
        session_floor = float(manual_session_started_at or 0.0)
        session_id_value = str(manual_session_id or "").strip()

        with self._lock:
            cur = self._conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            rows = cur.execute(
                f"""
                SELECT r.file_id, r.source_path, r.ingested_at, r.review_needed, r.msg_svr_id, r.talker,
                       r.msg_create_time, r.manual_session_id, r.client, r.txn_date, r.txn_time, r.bank, r.amount, r.amount_rounded, r.verification_status,
                       r.sheet_payload_json, r.sheet_status, f.first_seen AS source_first_seen
                FROM receipts AS r
                LEFT JOIN files AS f
                  ON f.file_id = r.file_id
                WHERE COALESCE(r.sheet_status, '') IN ('SINK_PENDING', 'SINK_BLOCKED_PRIOR_MSG', 'SINK_RETRY')
                  AND COALESCE(r.sheet_next_attempt, 0) <= ?
                ORDER BY
                    CASE
                        WHEN ? <> '' AND COALESCE(r.manual_session_id, '') = ? THEN 0
                        WHEN ? <> '' THEN 1
                        ELSE 0
                    END ASC,
                    CASE
                        WHEN ? > 0 AND COALESCE(r.ingested_at, 0) >= ? THEN 0
                        WHEN ? > 0 THEN 1
                        ELSE 0
                    END ASC,
                    CASE
                        WHEN r.talker IS NOT NULL AND r.talker <> '' AND r.msg_create_time IS NOT NULL AND r.msg_create_time > 0
                            THEN 0
                        ELSE 1
                    END ASC,
                    {order_by},
                    r.ingested_at ASC
                """,
                (
                    float(now),
                    session_id_value,
                    session_id_value,
                    session_id_value,
                    session_floor,
                    session_floor,
                    session_floor,
                ),
            ).fetchall()

            for row in rows:
                file_id = str(row["file_id"])
                msg_svr_id = str(row["msg_svr_id"] or "").strip() or None
                talker = str(row["talker"] or "").strip() or None
                msg_create_time = float(row["msg_create_time"] or 0.0)
                receipt_session_id = str(row["manual_session_id"] or "").strip() or None
                blocker_note: Optional[str] = None
                if scope == "per_talker" and talker and msg_create_time > 0:
                    prior_sink = self.find_prior_pending_sink_receipt(
                        talker=talker,
                        msg_create_time=msg_create_time,
                        msg_svr_id=msg_svr_id,
                        file_id=file_id,
                        manual_session_started_at=session_floor,
                        manual_session_id=receipt_session_id,
                    )
                    if prior_sink is not None:
                        blocker_id = self._receipt_message_sort_key(prior_sink["msg_svr_id"], str(prior_sink["file_id"]))
                        blocker_note = f"WAITING_PRIOR_SINK_RECEIPT:{blocker_id}"
                    elif msg_svr_id:
                        prior_job = self.find_prior_pending_message_job(
                            talker=talker,
                            create_time=msg_create_time,
                            msg_svr_id=msg_svr_id,
                            manual_session_started_at=session_floor,
                            manual_session_id=receipt_session_id,
                        )
                        if prior_job is not None:
                            blocker_note = f"WAITING_PRIOR_SINK_SESSION_MESSAGE:{str(prior_job['msg_svr_id'])}"

                if blocker_note:
                    cur.execute(
                        """
                        UPDATE receipts
                        SET sheet_status='SINK_BLOCKED_PRIOR_MSG',
                            sheet_last_error=?
                        WHERE file_id=?
                        """,
                        (blocker_note[:1200], file_id),
                    )
                    continue

                cur.execute(
                    """
                    UPDATE receipts
                    SET sheet_status='SINK_RUNNING',
                        sheet_last_error=NULL
                    WHERE file_id=?
                    """,
                    (file_id,),
                )
                self._conn.commit()

                payload_json = str(row["sheet_payload_json"] or "").strip()
                try:
                    payload: dict[str, Any] = json.loads(payload_json) if payload_json else {}
                except Exception:
                    payload = {}
                if not payload:
                    payload = {
                        "file_id": file_id,
                        "client": row["client"],
                        "txn_date": row["txn_date"],
                        "txn_time": row["txn_time"],
                        "bank": row["bank"],
                        "amount": row["amount_rounded"] if row["amount_rounded"] is not None else row["amount"],
                        "verification_status": row["verification_status"],
                        "msg_svr_id": msg_svr_id,
                        "talker": talker,
                    }
                return {
                    "file_id": file_id,
                    "source_path": str(row["source_path"]),
                    "ingested_at": float(row["ingested_at"] or 0.0),
                    "source_first_seen": float(row["source_first_seen"] or 0.0),
                    "review_needed": bool(row["review_needed"]),
                    "msg_svr_id": msg_svr_id,
                    "talker": talker,
                    "msg_create_time": msg_create_time,
                    "manual_session_id": receipt_session_id,
                    "row_payload": payload,
                }

            self._conn.commit()
            return None

    def mark_receipt_sink_committed(self, file_id: str, sheet_name: str, row_idx: int, committed_at: float) -> None:
        with self._lock:
            self._conn.execute(
                """
                UPDATE receipts
                SET sheet_status='SINK_COMMITTED',
                    sheet_committed_at=?,
                    sheet_next_attempt=0,
                    sheet_last_error=NULL,
                    excel_sheet=?,
                    excel_row=?
                WHERE file_id=?
                """,
                (float(committed_at), str(sheet_name), int(row_idx), str(file_id)),
            )
            self._conn.commit()

    def mark_receipt_sink_retry(self, file_id: str, err: str, delay_sec: int) -> None:
        next_attempt = time.time() + max(5, int(delay_sec))
        with self._lock:
            self._conn.execute(
                """
                UPDATE receipts
                SET sheet_status='SINK_RETRY',
                    sheet_next_attempt=?,
                    sheet_last_error=?
                WHERE file_id=?
                """,
                (float(next_attempt), err[:1200], str(file_id)),
            )
            self._conn.commit()

    def ensure_message_job(
        self,
        msg_svr_id: Optional[str],
        talker: Optional[str],
        talker_display: Optional[str],
        thumb_path: Optional[Path],
        expected_image_path: Optional[Path],
        create_time: float,
        first_seen_at: float,
        manual_session_id: Optional[str] = None,
        state: str = "NEW",
        activation_seen_at: Optional[float] = None,
    ) -> None:
        msg_svr_id = str(msg_svr_id or "").strip()
        talker = str(talker or "").strip()
        if not msg_svr_id or not talker:
            return

        now = time.time()
        expected_str = str(expected_image_path) if expected_image_path else None
        thumb_str = str(thumb_path) if thumb_path else None
        display_str = str(talker_display or "").strip() or None
        first_seen = float(first_seen_at or now)
        requested_state = str(state or "NEW").strip() or "NEW"
        with self._lock:
            row = self._conn.execute(
                """
                SELECT first_seen_at, activation_seen_at, state, manual_session_id
                FROM message_jobs
                WHERE msg_svr_id=?
                LIMIT 1
                """,
                (msg_svr_id,),
            ).fetchone()
            preserved_first_seen = float(row["first_seen_at"]) if row is not None else first_seen
            if activation_seen_at is None:
                activation_seen = max(float(row["activation_seen_at"] or 0.0), first_seen) if row is not None else first_seen
            else:
                activation_seen = max(float(row["activation_seen_at"] or 0.0), float(activation_seen_at)) if row is not None else float(activation_seen_at)
            existing_state = str(row["state"] or "").strip() if row is not None else ""
            session_value = str(manual_session_id or "").strip() or None
            if is_message_job_terminal_state(existing_state):
                effective_state = existing_state
            elif existing_state == SESSION_PENDING_OPEN_STATE and requested_state != SESSION_PENDING_OPEN_STATE:
                effective_state = requested_state
            elif requested_state == SESSION_PENDING_OPEN_STATE and existing_state:
                effective_state = existing_state
            else:
                effective_state = requested_state
            effective_manual_session_id = session_value or (str(row["manual_session_id"] or "").strip() if row is not None else "")
            self._conn.execute(
                """
                INSERT INTO message_jobs(
                    msg_svr_id, talker, talker_display, thumb_path, expected_image_path,
                    create_time, state, first_seen_at, activation_seen_at, last_seen_at, next_ui_attempt_at, manual_session_id
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
                ON CONFLICT(msg_svr_id) DO UPDATE SET
                    talker=excluded.talker,
                    talker_display=COALESCE(NULLIF(excluded.talker_display, ''), message_jobs.talker_display),
                    thumb_path=COALESCE(NULLIF(excluded.thumb_path, ''), message_jobs.thumb_path),
                    expected_image_path=COALESCE(NULLIF(excluded.expected_image_path, ''), message_jobs.expected_image_path),
                    create_time=CASE
                        WHEN excluded.create_time > 0 THEN excluded.create_time
                        ELSE message_jobs.create_time
                    END,
                    activation_seen_at=CASE
                        WHEN excluded.activation_seen_at > COALESCE(message_jobs.activation_seen_at, 0)
                            THEN excluded.activation_seen_at
                        ELSE COALESCE(message_jobs.activation_seen_at, 0)
                    END,
                    state=CASE
                        WHEN message_jobs.state IN ('RESOLVED', 'THUMB_FALLBACK', 'EXCEPTION', 'IGNORED_SESSION_ROLLOVER', 'IGNORED_STALE_MANUAL_SESSION', 'IGNORED_BY_USER')
                            THEN message_jobs.state
                        WHEN message_jobs.state = ? AND excluded.state <> ?
                            THEN excluded.state
                        WHEN excluded.state = ? AND COALESCE(message_jobs.state, '') <> ''
                            THEN message_jobs.state
                        ELSE excluded.state
                    END,
                    manual_session_id=COALESCE(NULLIF(excluded.manual_session_id, ''), message_jobs.manual_session_id),
                    last_seen_at=excluded.last_seen_at
                """,
                (
                    msg_svr_id,
                    talker,
                    display_str,
                    thumb_str,
                    expected_str,
                    float(create_time),
                    effective_state,
                    preserved_first_seen,
                    activation_seen,
                    float(now),
                    effective_manual_session_id,
                    SESSION_PENDING_OPEN_STATE,
                    SESSION_PENDING_OPEN_STATE,
                    SESSION_PENDING_OPEN_STATE,
                ),
            )
            self._conn.commit()

    def list_receipts_needing_parser_backfill(self, limit: int = 5000) -> list[sqlite3.Row]:
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT file_id, ocr_text, ocr_conf, quality_score, client, bank, beneficiary,
                       amount, amount_raw, amount_rounded, amount_source, currency,
                       txn_date, txn_time, txn_date_source, txn_time_source,
                       parse_conf, review_needed, parser_json, verification_status,
                       sheet_status, sheet_payload_json, excel_sheet, excel_row,
                       msg_svr_id, talker, resolution_source
                FROM receipts
                WHERE txn_date IS NULL OR TRIM(txn_date)=''
                   OR txn_time IS NULL OR TRIM(txn_time)=''
                   OR txn_date_source IS NULL OR TRIM(txn_date_source)=''
                   OR txn_time_source IS NULL OR TRIM(txn_time_source)=''
                   OR amount_raw IS NULL OR TRIM(amount_raw)=''
                   OR amount_rounded IS NULL
                   OR amount_source IS NULL OR TRIM(amount_source)=''
                ORDER BY ingested_at ASC
                LIMIT ?
                """,
                (int(max(1, limit)),),
            ).fetchall()
        return list(rows)

    def update_receipt_parser_backfill(
        self,
        file_id: str,
        *,
        txn_date: Optional[str],
        txn_time: Optional[str],
        txn_date_source: Optional[str],
        txn_time_source: Optional[str],
        amount: Optional[float],
        amount_raw: Optional[str],
        amount_rounded: Optional[float],
        amount_source: Optional[str],
        currency: Optional[str],
        bank: Optional[str],
        parse_conf: Optional[float],
        review_needed: bool,
        parser_json: Optional[str],
        sheet_payload_json: Optional[str],
    ) -> None:
        with self._lock:
            self._conn.execute(
                """
                UPDATE receipts
                SET txn_date=?,
                    txn_time=?,
                    txn_date_source=?,
                    txn_time_source=?,
                    amount=?,
                    amount_raw=?,
                    amount_rounded=?,
                    amount_source=?,
                    currency=?,
                    bank=?,
                    parse_conf=COALESCE(?, parse_conf),
                    review_needed=?,
                    parser_json=?,
                    sheet_payload_json=?
                WHERE file_id=?
                """,
                (
                    txn_date,
                    txn_time,
                    txn_date_source,
                    txn_time_source,
                    amount,
                    amount_raw,
                    amount_rounded,
                    amount_source,
                    currency,
                    bank,
                    parse_conf,
                    1 if review_needed else 0,
                    parser_json,
                    sheet_payload_json,
                    str(file_id),
                ),
            )
            self._conn.commit()

    def get_message_job(self, msg_svr_id: Optional[str]) -> Optional[sqlite3.Row]:
        msg_svr_id = str(msg_svr_id or "").strip()
        if not msg_svr_id:
            return None
        with self._lock:
            return self._conn.execute(
                """
                SELECT *
                FROM message_jobs
                WHERE msg_svr_id=?
                LIMIT 1
                """,
                (msg_svr_id,),
            ).fetchone()

    def get_message_job_by_expected_path(self, expected_image_path: Path | str) -> Optional[sqlite3.Row]:
        with self._lock:
            return self._conn.execute(
                """
                SELECT *
                FROM message_jobs
                WHERE expected_image_path=?
                ORDER BY last_seen_at DESC
                LIMIT 1
                """,
                (str(expected_image_path),),
            ).fetchone()

    def find_prior_pending_message_job(
        self,
        talker: Optional[str],
        create_time: float,
        msg_svr_id: Optional[str],
        manual_session_started_at: Optional[float] = None,
        manual_session_id: Optional[str] = None,
    ) -> Optional[sqlite3.Row]:
        talker_value = str(talker or "").strip()
        msg_value = str(msg_svr_id or "").strip()
        if not talker_value or create_time <= 0 or not msg_value:
            return None
        session_floor = float(manual_session_started_at or 0.0)
        session_id_value = str(manual_session_id or "").strip()
        with self._lock:
            return self._conn.execute(
                """
                SELECT msg_svr_id, create_time, state, expected_image_path, thumb_path, activation_seen_at, manual_session_id
                FROM message_jobs
                WHERE talker=?
                  AND msg_svr_id<>?
                  AND state NOT IN ('RESOLVED', 'THUMB_FALLBACK', 'EXCEPTION', 'IGNORED_SESSION_ROLLOVER', 'IGNORED_STALE_MANUAL_SESSION', 'IGNORED_BY_USER')
                  AND (? = '' OR COALESCE(manual_session_id, '') = ?)
                  AND (? <= 0 OR COALESCE(activation_seen_at, first_seen_at, create_time, 0) >= ?)
                  AND (
                    create_time < ?
                    OR (create_time = ? AND msg_svr_id < ?)
                  )
                ORDER BY create_time ASC, msg_svr_id ASC
                LIMIT 1
                """,
                (
                    talker_value,
                    msg_value,
                    session_id_value,
                    session_id_value,
                    session_floor,
                    session_floor,
                    float(create_time),
                    float(create_time),
                    msg_value,
                ),
            ).fetchone()

    def set_message_job_state(
        self,
        msg_svr_id: Optional[str],
        state: str,
        note: Optional[str] = None,
        next_ui_attempt_at: Optional[float] = None,
        batch_id: Optional[str] = None,
        reset_batch: bool = False,
        touch_ui_requested: bool = False,
        touch_ui_completed: bool = False,
    ) -> None:
        msg_svr_id = str(msg_svr_id or "").strip()
        if not msg_svr_id:
            return

        now = time.time()
        with self._lock:
            row = self._conn.execute(
                """
                SELECT state, next_ui_attempt_at, batch_id, last_ui_result
                FROM message_jobs
                WHERE msg_svr_id=?
                LIMIT 1
                """,
                (msg_svr_id,),
            ).fetchone()
            if row is None:
                return

            current_state = str(row["state"] or "")
            current_note = str(row["last_ui_result"] or "").strip()
            effective_state = state
            if is_message_job_terminal_state(current_state) and not is_message_job_terminal_state(state):
                effective_state = current_state
            elif current_state == "THUMB_FALLBACK" and state != "EXCEPTION":
                effective_state = "THUMB_FALLBACK"

            effective_next = float(next_ui_attempt_at) if next_ui_attempt_at is not None else float(row["next_ui_attempt_at"] or 0.0)
            effective_batch = None if reset_batch else (batch_id if batch_id is not None else row["batch_id"])
            if note:
                note_value = note[:1200]
                if current_note.startswith("ui_") and not note_value.startswith("ui_") and is_message_job_terminal_state(effective_state):
                    effective_note = current_note
                else:
                    effective_note = note_value
            else:
                effective_note = current_note

            self._conn.execute(
                """
                UPDATE message_jobs
                SET state=?,
                    last_seen_at=?,
                    next_ui_attempt_at=?,
                    batch_id=?,
                    last_ui_result=?,
                    ui_force_requested_at=CASE WHEN ? THEN ? ELSE ui_force_requested_at END,
                    ui_force_completed_at=CASE WHEN ? THEN ? ELSE ui_force_completed_at END
                WHERE msg_svr_id=?
                """,
                (
                    effective_state,
                    float(now),
                    effective_next,
                    effective_batch,
                    effective_note,
                    1 if touch_ui_requested else 0,
                    float(now),
                    1 if touch_ui_completed else 0,
                    float(now),
                    msg_svr_id,
                ),
            )
            self._conn.commit()

    def mark_message_job_resolved(self, msg_svr_id: Optional[str], note: Optional[str] = None) -> None:
        self.set_message_job_state(
            msg_svr_id,
            state="RESOLVED",
            note=note,
            next_ui_attempt_at=0.0,
            reset_batch=True,
            touch_ui_completed=True,
        )

    def mark_message_job_thumb_fallback(self, msg_svr_id: Optional[str], note: Optional[str] = None) -> None:
        self.set_message_job_state(
            msg_svr_id,
            state="THUMB_FALLBACK",
            note=note,
            next_ui_attempt_at=0.0,
            reset_batch=True,
        )

    def mark_message_job_exception(self, msg_svr_id: Optional[str], note: Optional[str] = None) -> None:
        self.set_message_job_state(
            msg_svr_id,
            state="EXCEPTION",
            note=note,
            next_ui_attempt_at=0.0,
            reset_batch=True,
            touch_ui_completed=True,
        )

    def mark_message_job_ignored(self, msg_svr_id: Optional[str], state: str, note: Optional[str] = None) -> None:
        self.set_message_job_state(
            msg_svr_id,
            state=state,
            note=note,
            next_ui_attempt_at=0.0,
            reset_batch=True,
        )

    def message_job_is_terminal(self, msg_svr_id: Optional[str]) -> bool:
        row = self.get_message_job(msg_svr_id)
        if row is None:
            return False
        return is_message_job_terminal_state(row["state"])

    def resolve_message_job_paths(self, msg_svr_id: Optional[str], exclude_file_id: str, sha256: str = "") -> int:
        row = self.get_message_job(msg_svr_id)
        if row is None:
            return 0

        paths = [str(row["thumb_path"] or "").strip(), str(row["expected_image_path"] or "").strip()]
        paths = [path for path in paths if path]
        if not paths:
            return 0

        now = time.time()
        placeholders = ",".join("?" for _ in paths)
        params: list[Any] = [float(now), sha256, "RESOLVED_BY_LATER_SUCCESS"]
        params.extend(paths)
        params.extend([str(exclude_file_id)])
        with self._lock:
            cur = self._conn.cursor()
            cur.execute(
                f"""
                UPDATE files
                SET status='done',
                    processed_at=?,
                    sha256=CASE WHEN sha256 IS NULL OR sha256='' THEN ? ELSE sha256 END,
                    last_error=?
                WHERE path IN ({placeholders})
                  AND file_id<>?
                  AND status IN ('pending', 'retry', 'processing', 'exception')
                """,
                params,
            )
            changed = int(cur.rowcount or 0)
            self._conn.commit()
            return changed

    def set_msg_cursor(self, create_time: float, msg_svr_id: Optional[str]) -> None:
        if create_time <= 0:
            return
        self.set_meta("last_msg_cursor", f"{int(create_time)}|{str(msg_svr_id or '').strip()}")

    def claim_ui_batch(self, materialization_order: str = "desc") -> tuple[Optional[str], list[dict[str, Any]]]:
        now = time.time()
        order = str(materialization_order or "desc").strip().lower() or "desc"
        order_by = "create_time DESC, msg_svr_id DESC"
        if order == "asc":
            order_by = "create_time ASC, msg_svr_id ASC"
        with self._lock:
            cur = self._conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            talker_row = cur.execute(
                """
                SELECT talker
                FROM message_jobs
                WHERE state='UI_FORCE_PENDING'
                  AND next_ui_attempt_at <= ?
                  AND talker_display IS NOT NULL
                  AND expected_image_path IS NOT NULL
                  AND thumb_path IS NOT NULL
                ORDER BY first_seen_at ASC, create_time DESC
                LIMIT 1
                """,
                (float(now),),
            ).fetchone()
            if talker_row is None:
                self._conn.commit()
                return None, []

            talker = str(talker_row["talker"])
            rows = cur.execute(
                f"""
                SELECT msg_svr_id, talker, talker_display, expected_image_path, thumb_path, create_time, ui_force_attempts
                FROM message_jobs
                WHERE talker=?
                  AND state='UI_FORCE_PENDING'
                  AND next_ui_attempt_at <= ?
                  AND talker_display IS NOT NULL
                  AND expected_image_path IS NOT NULL
                  AND thumb_path IS NOT NULL
                ORDER BY {order_by}
                """,
                (talker, float(now)),
            ).fetchall()
            if not rows:
                self._conn.commit()
                return None, []

            batch_id = hashlib.sha1(f"{talker}|{now}".encode("utf-8")).hexdigest()[:16]
            msg_ids = [str(row["msg_svr_id"]) for row in rows]
            cur.executemany(
                """
                UPDATE message_jobs
                SET state='UI_FORCE_RUNNING',
                    batch_id=?,
                    ui_force_requested_at=?,
                    ui_force_attempts=ui_force_attempts + 1,
                    last_seen_at=?
                WHERE msg_svr_id=?
                """,
                [(batch_id, float(now), float(now), msg_id) for msg_id in msg_ids],
            )
            self._conn.commit()

        jobs: list[dict[str, Any]] = []
        for row in rows:
            jobs.append(
                {
                    "msg_svr_id": str(row["msg_svr_id"]),
                    "talker": str(row["talker"]),
                    "talker_display": str(row["talker_display"]),
                    "expected_image_path": str(row["expected_image_path"]),
                    "thumb_path": str(row["thumb_path"]),
                    "create_time": float(row["create_time"]),
                    "ui_force_attempts": int(row["ui_force_attempts"]) + 1,
                }
            )
        return batch_id, jobs

    def finish_ui_batch(
        self,
        batch_id: Optional[str],
        resolved_msg_ids: list[str],
        note: str,
        backoff_seconds: list[int],
        resolved_notes_by_msg_id: Optional[dict[str, str]] = None,
    ) -> None:
        batch_id = str(batch_id or "").strip()
        if not batch_id:
            return

        now = time.time()
        resolved_set = {str(msg_id) for msg_id in resolved_msg_ids if str(msg_id).strip()}
        resolved_notes = {str(key): str(value)[:1200] for key, value in (resolved_notes_by_msg_id or {}).items() if str(key).strip() and str(value).strip()}
        with self._lock:
            rows = self._conn.execute(
                """
                SELECT msg_svr_id, ui_force_attempts
                FROM message_jobs
                WHERE batch_id=?
                """,
                (batch_id,),
            ).fetchall()
            cur = self._conn.cursor()
            for row in rows:
                msg_svr_id = str(row["msg_svr_id"])
                if msg_svr_id in resolved_set:
                    success_note = resolved_notes.get(msg_svr_id, note[:1200])
                    cur.execute(
                        """
                        UPDATE message_jobs
                        SET state='WAITING_ORIGINAL',
                            batch_id=NULL,
                            last_seen_at=?,
                            next_ui_attempt_at=0,
                            last_ui_result=?,
                            ui_force_completed_at=?
                        WHERE msg_svr_id=?
                        """,
                        (float(now), success_note, float(now), msg_svr_id),
                    )
                    continue

                attempts = int(row["ui_force_attempts"] or 0)
                idx = min(max(0, attempts - 1), max(0, len(backoff_seconds) - 1))
                delay = int(backoff_seconds[idx]) if backoff_seconds else 10
                cur.execute(
                    """
                    UPDATE message_jobs
                    SET state='UI_FORCE_PENDING',
                        batch_id=NULL,
                        last_seen_at=?,
                        next_ui_attempt_at=?,
                        last_ui_result=?
                    WHERE msg_svr_id=?
                    """,
                    (float(now), float(now + max(3, delay)), note[:1200], msg_svr_id),
                )
            self._conn.commit()

    def close(self) -> None:
        with self._lock:
            self._conn.close()


GOOGLE_SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


def normalize_header_text(value: Any) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(value or "")) if not unicodedata.combining(ch)).strip().upper()


def normalize_header_cells(values: list[Any], expected_headers: list[str]) -> list[str]:
    out: list[str] = []
    for value in values[: len(expected_headers)]:
        out.append(normalize_header_text(value))
    while len(out) < len(expected_headers):
        out.append("")
    return out


def extract_google_sheet_ref(value: str) -> tuple[str, Optional[int]]:
    raw = (value or "").strip()
    if not raw:
        raise ValueError("Google Sheets destination is empty")

    if raw.startswith("http://") or raw.startswith("https://"):
        parsed = urlparse(raw)
        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9\-_]+)", parsed.path)
        if not match:
            raise ValueError(f"Unable to extract spreadsheet id from URL: {raw}")
        gid_value = parse_qs(parsed.query).get("gid", [None])[0]
        gid = int(gid_value) if gid_value and str(gid_value).isdigit() else None
        return match.group(1), gid

    if re.fullmatch(r"[a-zA-Z0-9\-_]{20,}", raw):
        return raw, None

    raise ValueError(f"Invalid Google Sheets reference: {raw}")


class RowSink:
    def append(self, row_payload: dict[str, Any], review_needed: bool) -> tuple[str, int]:
        raise NotImplementedError

    def update_row(self, sheet_name: str, row_idx: int, row_payload: dict[str, Any], review_needed: bool) -> None:
        raise NotImplementedError


class ExcelSink(RowSink):
    def __init__(self, excel_path: Path, verification_column_name: str = DEFAULT_VERIFICATION_COLUMN_NAME) -> None:
        self.excel_path = excel_path
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()
        self.headers = build_lanc_headers(verification_column_name)
        self.expected_header_normalized = [normalize_header_text(header) for header in self.headers]
        self._init_workbook()

    def _sheet_has_expected_header(self, wb: Any, sheet_name: str) -> bool:
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        first_row = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())]
        return normalize_header_cells(first_row, self.headers) == self.expected_header_normalized

    def _create_fresh_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lancamentos"
        ws.append(self.headers)
        ws2 = wb.create_sheet("Revisar")
        ws2.append(self.headers)
        wb.save(self.excel_path)
        wb.close()

    def _init_workbook(self) -> None:
        if not self.excel_path.exists():
            self._create_fresh_workbook()
            return

        wb = load_workbook(self.excel_path)
        ok_layout = self._sheet_has_expected_header(wb, "Lancamentos") and self._sheet_has_expected_header(wb, "Revisar")
        wb.close()

        if ok_layout:
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        legacy = self.excel_path.with_name(f"{self.excel_path.stem}_legacy_{ts}{self.excel_path.suffix}")
        try:
            self.excel_path.replace(legacy)
        except Exception:
            pass
        self._create_fresh_workbook()

    def append(self, row_payload: dict[str, Any], review_needed: bool) -> tuple[str, int]:
        with self._lock:
            wb = load_workbook(self.excel_path)
            sheet = "Revisar" if review_needed else "Lancamentos"
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
                ws.append(self.headers)
            ws = wb[sheet]
            ws.append(build_sink_row_values(row_payload))
            row_idx = ws.max_row
            wb.save(self.excel_path)
            wb.close()
            return (sheet, row_idx)

    def update_row(self, sheet_name: str, row_idx: int, row_payload: dict[str, Any], review_needed: bool) -> None:
        target_sheet = str(sheet_name or "").strip() or ("Revisar" if review_needed else "Lancamentos")
        with self._lock:
            wb = load_workbook(self.excel_path)
            if target_sheet not in wb.sheetnames:
                ws = wb.create_sheet(target_sheet)
                ws.append(self.headers)
            ws = wb[target_sheet]
            for col_idx, value in enumerate(build_sink_row_values(row_payload), start=1):
                ws.cell(row=max(2, int(row_idx)), column=col_idx, value=value)
            wb.save(self.excel_path)
            wb.close()


class GoogleSheetsSink(RowSink):
    def __init__(
        self,
        spreadsheet_ref: str,
        credentials_path: Optional[Path],
        main_worksheet: Optional[str] = None,
        review_worksheet: Optional[str] = "Revisar",
        verification_column_name: str = DEFAULT_VERIFICATION_COLUMN_NAME,
    ) -> None:
        self.spreadsheet_id, self.preferred_gid = extract_google_sheet_ref(spreadsheet_ref)
        self.credentials_path = credentials_path
        self.main_worksheet = (main_worksheet or "").strip() or None
        self.review_worksheet = (review_worksheet or "").strip() or None
        self._lock = threading.Lock()
        self.auth_mode = "unknown"
        self.headers = build_lanc_headers(verification_column_name)
        self.expected_header_normalized = [normalize_header_text(header) for header in self.headers]
        self._client = self._build_client()
        self._spreadsheet = self._client.open_by_key(self.spreadsheet_id)
        self._worksheets_by_title: dict[str, Any] = {}
        self._sheet_order: list[str] = []
        self._headers_verified: set[str] = set()
        self._main_sheet_title: Optional[str] = None
        self.spreadsheet_title: Optional[str] = getattr(self._spreadsheet, "title", None)
        self._refresh_sheet_cache()
        self._main_sheet_title = self._resolve_main_sheet_title()
        self._ensure_header(self._main_sheet_title)
        if self.review_worksheet and self.review_worksheet != self._main_sheet_title:
            self._ensure_sheet_exists(self.review_worksheet)
            self._ensure_header(self.review_worksheet)

    def _build_client(self) -> Any:
        import gspread  # type: ignore

        client = None
        errors: list[str] = []

        if self.credentials_path:
            if self.credentials_path.exists():
                try:
                    client = gspread.service_account(filename=str(self.credentials_path))
                    self.auth_mode = f"service_account:{self.credentials_path.name}"
                except Exception as exc:
                    errors.append(f"service account credentials failed: {exc}")
            else:
                errors.append(f"credential file not found: {self.credentials_path}")

        if client is None:
            try:
                import google.auth  # type: ignore

                creds, _project = google.auth.default(scopes=GOOGLE_SHEETS_SCOPES)
                client = gspread.authorize(creds)
                self.auth_mode = "application_default_credentials"
            except Exception as exc:
                errors.append(f"application default credentials failed: {exc}")

        if client is None:
            joined = " | ".join(errors) if errors else "no credentials configured"
            raise RuntimeError(
                "Google Sheets auth is not configured. "
                "Add a service account JSON at the configured path or login with "
                "`gcloud auth application-default login`. "
                f"Details: {joined}"
            )

        return client

    def _refresh_sheet_cache(self) -> None:
        worksheets = self._spreadsheet.worksheets()
        self._worksheets_by_title = {ws.title: ws for ws in worksheets}
        self._sheet_order = [ws.title for ws in worksheets]

    def _resolve_main_sheet_title(self) -> str:
        if self.main_worksheet:
            return self._ensure_sheet_exists(self.main_worksheet)

        if self.preferred_gid is not None:
            for title, worksheet in self._worksheets_by_title.items():
                sheet_id = int(getattr(worksheet, "id", 0) or 0)
                if sheet_id == self.preferred_gid:
                    return title

        if not self._sheet_order:
            raise RuntimeError("Google spreadsheet has no worksheets")
        return self._sheet_order[0]

    def _ensure_sheet_exists(self, title: str) -> str:
        if title in self._worksheets_by_title:
            return title

        try:
            self._spreadsheet.add_worksheet(title=title, rows=1000, cols=max(6, len(self.headers)))
        except Exception:
            self._refresh_sheet_cache()
            if title in self._worksheets_by_title:
                return title
            raise

        self._refresh_sheet_cache()
        return title

    def _ensure_header(self, title: str) -> None:
        if title in self._headers_verified:
            return

        worksheet = self._worksheets_by_title[title]
        values = worksheet.row_values(1)
        if not values or not any(str(v).strip() for v in values):
            worksheet.update(range_name=sheet_header_range(self.headers), values=[self.headers])
        elif normalize_header_cells(values, self.headers) != self.expected_header_normalized:
            legacy_headers = BASE_LANC_HEADERS
            legacy_expected = [normalize_header_text(header) for header in legacy_headers]
            if normalize_header_cells(values, legacy_headers) == legacy_expected:
                worksheet.update(range_name=sheet_header_range(self.headers), values=[self.headers])
            else:
                print(f"[WARN] worksheet_header_unexpected | worksheet={title} | found={values}")
        self._headers_verified.add(title)

    def _target_sheet(self, review_needed: bool) -> str:
        main_title = self._main_sheet_title or self._resolve_main_sheet_title()
        if review_needed and self.review_worksheet and self.review_worksheet != main_title:
            return self.review_worksheet
        return main_title

    def append(self, row_payload: dict[str, Any], review_needed: bool) -> tuple[str, int]:
        with self._lock:
            title = self._target_sheet(review_needed)
            self._ensure_header(title)
            worksheet = self._worksheets_by_title[title]
            worksheet.append_row(
                build_sink_row_values(row_payload),
                value_input_option="USER_ENTERED",
                table_range=sheet_table_range(self.headers),
            )
            row_idx = len(worksheet.col_values(1))
            return title, row_idx

    def update_row(self, sheet_name: str, row_idx: int, row_payload: dict[str, Any], review_needed: bool) -> None:
        with self._lock:
            title = str(sheet_name or "").strip() or self._target_sheet(review_needed)
            if title not in self._worksheets_by_title:
                title = self._ensure_sheet_exists(title)
            self._ensure_header(title)
            worksheet = self._worksheets_by_title[title]
            worksheet.update(
                range_name=sheet_row_range(self.headers, max(2, int(row_idx))),
                values=[build_sink_row_values(row_payload)],
                value_input_option="USER_ENTERED",
            )


class IngestEventHandler(FileSystemEventHandler):  # type: ignore[misc]
    def __init__(self, db: StateDB, cfg: "Config", media_resolver: Optional[WeChatDBResolver]) -> None:
        self.db = db
        self.cfg = cfg
        self.media_resolver = media_resolver

    def on_created(self, event: Any) -> None:
        if event.is_directory:
            return
        file_id = self.db.upsert_candidate(
            Path(event.src_path),
            self.cfg.settle_seconds,
            "created",
            thumb_candidates_enabled=self.cfg.thumb_candidates_enabled,
        )
        if file_id is not None:
            preregister_manual_order_candidate(self.db, self.media_resolver, self.cfg, file_id, Path(event.src_path), "created")

    def on_modified(self, event: Any) -> None:
        if event.is_directory:
            return
        file_id = self.db.upsert_candidate(
            Path(event.src_path),
            self.cfg.settle_seconds,
            "modified",
            thumb_candidates_enabled=self.cfg.thumb_candidates_enabled,
        )
        if file_id is not None:
            preregister_manual_order_candidate(self.db, self.media_resolver, self.cfg, file_id, Path(event.src_path), "modified")


@dataclass
class Config:
    watch_roots: list[Path]
    db_path: Path
    db_merge_path: Path
    excel_path: Path
    sink_mode: str
    gsheet_ref: Optional[str]
    gsheet_worksheet: Optional[str]
    gsheet_review_worksheet: Optional[str]
    google_credentials_path: Optional[Path]
    verification_column_name: str
    client_map_path: Path
    resolution_mode: str
    settle_seconds: int
    reconcile_seconds: int
    recent_files_hours: int
    idle_sleep_seconds: float
    retry_base_seconds: int
    min_confidence: float
    max_retries: int
    original_wait_seconds: int
    temp_correlation_seconds: int
    thumb_candidates_enabled: bool
    manual_order_guard_enabled: bool
    manual_burst_gap_seconds: int
    manual_burst_max_seconds: int
    ui_force_download_enabled: bool
    ui_force_delay_seconds: int
    ui_force_scope: str
    ui_focus_policy: str
    ui_batch_mode: str
    ui_item_timeout_seconds: int
    ui_retry_backoff_seconds: list[int]
    ui_window_backends: list[str]
    ui_window_classes: list[str]
    sheet_order_scope: str
    sheet_materialization_order: str
    sheet_commit_order: str
    disable_watchdog: bool


def reconcile_scan(cfg: Config, db: StateDB) -> int:
    count = 0
    now = time.time()
    overlap_sec = max(60.0, float(cfg.settle_seconds * 3))
    fallback_floor = now - max(1, cfg.recent_files_hours) * 3600
    previous_watermark = db.get_meta_float("reconcile_watermark")
    scan_floor = max(0.0, fallback_floor if previous_watermark is None else previous_watermark - overlap_sec)
    newest_mtime = previous_watermark or 0.0
    for root in cfg.watch_roots:
        if not root.exists():
            continue
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if not is_candidate(p, thumb_candidates_enabled=cfg.thumb_candidates_enabled):
                continue
            try:
                st = p.stat()
            except FileNotFoundError:
                continue
            if float(st.st_mtime) < scan_floor:
                continue
            if db.upsert_candidate(
                p,
                cfg.settle_seconds,
                "reconcile",
                thumb_candidates_enabled=cfg.thumb_candidates_enabled,
            ):
                count += 1
            newest_mtime = max(newest_mtime, float(st.st_mtime))
    db.set_meta("reconcile_watermark", f"{max(newest_mtime, now - overlap_sec):.6f}")
    return count


def resolve_message_context_for_candidate(
    db: StateDB,
    media_resolver: Optional[WeChatDBResolver],
    cfg: Config,
    path: Path,
    source_kind: str,
    mtime: float,
) -> tuple[Optional[WeChatMessageRef], Path, Optional[Path]]:
    if media_resolver is None:
        return None, path, None

    if source_kind == "temp_image":
        context_path_str = db.find_recent_msgattach_context_path(
            mtime,
            lookback_sec=cfg.temp_correlation_seconds,
            lookahead_sec=15,
            limit=48,
        )
        if context_path_str is None:
            context_path_str = db.find_recent_unresolved_msgattach_context_path(
                max_age_sec=max(cfg.original_wait_seconds * 4, 1800),
                limit=60,
            )
        if context_path_str is None:
            return None, path, None
        client_source_path = Path(context_path_str)
        group_hash = extract_group_id_from_path(client_source_path)
        context_row = db.get_latest_file_row_by_path(client_source_path)
        context_mtime = float(context_row["mtime"]) if context_row is not None else mtime
        msg_ref = media_resolver.find_message_for_path(client_source_path, context_mtime)
        if msg_ref is None and group_hash:
            msg_ref = media_resolver.find_unique_message_for_group(
                group_hash,
                context_mtime,
                max(cfg.original_wait_seconds * 4, 1800),
            )
        resolved_client_source = msg_ref.preferred_context_path() if msg_ref is not None else None
        return msg_ref, resolved_client_source or client_source_path, msg_ref.thumb_abs_path if msg_ref is not None else None

    msg_ref = media_resolver.find_message_for_path(path, mtime)
    preferred_context = msg_ref.preferred_context_path() if msg_ref is not None else None
    thumb_path = path if source_kind == "msgattach_thumb_dat" else msg_ref.thumb_abs_path if msg_ref is not None else None
    return msg_ref, preferred_context or path, thumb_path


def preregister_manual_order_candidate(
    db: StateDB,
    media_resolver: Optional[WeChatDBResolver],
    cfg: Config,
    file_id: Optional[str],
    path: Path,
    source_event: str,
) -> None:
    if not cfg.manual_order_guard_enabled:
        return
    source_kind = detect_source_kind(path)
    if not should_refresh_manual_session(source_kind, source_event):
        return
    if media_resolver is None:
        return

    file_row = db.get_file(file_id)
    if file_row is None:
        return

    msg_ref, client_source_path, thumb_path = resolve_message_context_for_candidate(
        db=db,
        media_resolver=media_resolver,
        cfg=cfg,
        path=path,
        source_kind=source_kind,
        mtime=float(file_row["mtime"] or 0.0),
    )
    if msg_ref is None or not msg_ref.msg_svr_id or not msg_ref.talker:
        return

    existing_job = db.get_message_job(msg_ref.msg_svr_id)
    preferred_session_id = str(existing_job["manual_session_id"] or "").strip() if existing_job is not None else None
    session_row = db.start_or_extend_manual_order_session(
        talker=msg_ref.talker,
        create_time=float(msg_ref.create_time or 0.0),
        event_ts=time.time(),
        burst_gap_seconds=cfg.manual_burst_gap_seconds,
        burst_max_seconds=cfg.manual_burst_max_seconds,
        preferred_session_id=preferred_session_id,
    )
    if session_row is None:
        return

    expected_image_path = msg_ref.image_abs_path or expected_full_image_from_thumb_path(thumb_path or client_source_path)
    talker_display = media_resolver.resolve_talker_display_name(msg_ref.talker) if media_resolver is not None else msg_ref.talker
    db.ensure_message_job(
        msg_svr_id=msg_ref.msg_svr_id,
        talker=msg_ref.talker,
        talker_display=talker_display or msg_ref.talker,
        thumb_path=thumb_path,
        expected_image_path=expected_image_path,
        create_time=msg_ref.create_time,
        first_seen_at=float(file_row["first_seen"] or time.time()),
        manual_session_id=str(session_row["session_id"]),
        state="NEW",
        activation_seen_at=float(file_row["first_seen"] or time.time()),
    )
    db.update_file_message_context(
        file_id,
        msg_svr_id=msg_ref.msg_svr_id,
        talker=msg_ref.talker,
        msg_create_time=float(msg_ref.create_time or 0.0),
        manual_session_id=str(session_row["session_id"]),
        session_release_at=float(session_row["release_at"] or 0.0),
    )
    db.set_msg_cursor(msg_ref.create_time, msg_ref.msg_svr_id)


def seed_ready_manual_session_placeholders(
    db: StateDB,
    media_resolver: Optional[WeChatDBResolver],
    cfg: Config,
) -> int:
    if not cfg.manual_order_guard_enabled or media_resolver is None:
        return 0

    seeded = 0
    for session in db.list_manual_sessions_ready_for_seed():
        session_id = str(session["session_id"] or "").strip()
        talker = str(session["talker"] or "").strip()
        min_create_time = float(session["min_create_time"] or 0.0)
        max_create_time = float(session["max_create_time"] or 0.0)
        if not session_id or not talker or min_create_time <= 0 or max_create_time <= 0 or max_create_time < min_create_time:
            db.mark_manual_session_seeded(session_id)
            continue

        for msg_ref in media_resolver.list_image_messages_for_talker(talker, min_create_time, max_create_time):
            if msg_ref.msg_svr_id is None or not msg_ref.talker:
                continue
            expected_image_path = msg_ref.image_abs_path or expected_full_image_from_thumb_path(msg_ref.thumb_abs_path or Path(""))
            existing_job = db.get_message_job(msg_ref.msg_svr_id)
            if existing_job is None:
                talker_display = media_resolver.resolve_talker_display_name(msg_ref.talker) if media_resolver is not None else msg_ref.talker
                db.ensure_message_job(
                    msg_svr_id=msg_ref.msg_svr_id,
                    talker=msg_ref.talker,
                    talker_display=talker_display or msg_ref.talker,
                    thumb_path=msg_ref.thumb_abs_path,
                    expected_image_path=expected_image_path,
                    create_time=msg_ref.create_time,
                    first_seen_at=float(session["started_at"] or time.time()),
                    manual_session_id=session_id,
                    state=SESSION_PENDING_OPEN_STATE,
                    activation_seen_at=0.0,
                )
                seeded += 1
            elif not str(existing_job["manual_session_id"] or "").strip():
                db.ensure_message_job(
                    msg_svr_id=msg_ref.msg_svr_id,
                    talker=msg_ref.talker,
                    talker_display=str(existing_job["talker_display"] or "").strip() or msg_ref.talker,
                    thumb_path=msg_ref.thumb_abs_path,
                    expected_image_path=expected_image_path,
                    create_time=msg_ref.create_time,
                    first_seen_at=float(existing_job["first_seen_at"] or session["started_at"] or time.time()),
                    manual_session_id=session_id,
                    state=str(existing_job["state"] or "NEW"),
                    activation_seen_at=float(existing_job["activation_seen_at"] or 0.0),
                )
        db.mark_manual_session_seeded(session_id)
    return seeded


def ensure_message_job_tracking(
    db: StateDB,
    resolver: ClientResolver,
    media_resolver: Optional[WeChatDBResolver],
    cfg: Config,
    msg_ref: Optional[WeChatMessageRef],
    thumb_path: Optional[Path],
    client_source_path: Path,
    first_seen_at: float,
    manual_session_id: Optional[str] = None,
) -> Optional[sqlite3.Row]:
    if msg_ref is None or not msg_ref.msg_svr_id or not msg_ref.talker:
        return None
    if cfg.ui_force_scope == "mapped-groups" and not resolver.resolve(client_source_path):
        return None

    expected_image_path = msg_ref.image_abs_path or expected_full_image_from_thumb_path(thumb_path or client_source_path)
    if expected_image_path is None:
        return None

    talker_display = media_resolver.resolve_talker_display_name(msg_ref.talker) if media_resolver is not None else msg_ref.talker
    db.ensure_message_job(
        msg_svr_id=msg_ref.msg_svr_id,
        talker=msg_ref.talker,
        talker_display=talker_display or msg_ref.talker,
        thumb_path=thumb_path,
        expected_image_path=expected_image_path,
        create_time=msg_ref.create_time,
        first_seen_at=first_seen_at,
        manual_session_id=manual_session_id,
    )
    db.set_msg_cursor(msg_ref.create_time, msg_ref.msg_svr_id)
    return db.get_message_job(msg_ref.msg_svr_id)


class UIForceDownloadWorker(threading.Thread):
    def __init__(
        self,
        db: StateDB,
        cfg: Config,
        stop_event: threading.Event,
        media_resolver: Optional[WeChatDBResolver] = None,
    ) -> None:
        super().__init__(name="wechat-ui-force-download", daemon=True)
        self.db = db
        self.cfg = cfg
        self.stop_event = stop_event
        self.media_resolver = media_resolver
        self.available = UI_FORCE_DOWNLOADER_AVAILABLE and WeChatUIForceDownloader is not None
        self.unavailable_reason = None if self.available else (UI_FORCE_DOWNLOADER_IMPORT_ERROR or "ui_downloader_unavailable")
        self._downloader = (
            WeChatUIForceDownloader(
                focus_policy=cfg.ui_focus_policy,
                item_timeout_seconds=cfg.ui_item_timeout_seconds,
                window_backends=cfg.ui_window_backends,
                window_class_candidates=cfg.ui_window_classes,
            )
            if self.available
            else None
        )

    def run(self) -> None:
        if not self.available or self._downloader is None:
            return

        runtime_disabled = False
        while not self.stop_event.is_set():
            runtime_enabled = self.db.is_ui_force_runtime_enabled(default_enabled=self.cfg.ui_force_download_enabled)
            if not runtime_enabled:
                if not runtime_disabled:
                    moved_jobs, requeued_files = self.db.set_ui_force_runtime_enabled(False, release_waiting=True)
                    self.db.set_meta("last_ui_result", "ui_force_disabled_runtime_manual_mode")
                    print(
                        f"[UI] runtime_disabled | moved_jobs={moved_jobs} | "
                        f"requeued_files={requeued_files}"
                    )
                    runtime_disabled = True
                self.stop_event.wait(max(0.7, self.cfg.idle_sleep_seconds))
                continue
            if runtime_disabled:
                self.db.set_meta("last_ui_result", "ui_force_reenabled_runtime")
                print("[UI] runtime_reenabled")
                runtime_disabled = False

            try:
                batch_id, jobs = self.db.claim_ui_batch(materialization_order=self.cfg.sheet_materialization_order)
            except Exception as exc:
                note = f"ui_claim_failed:{type(exc).__name__}:{exc}"
                self.db.set_meta("last_ui_result", note)
                print(f"[UI] claim_failed | err={note}")
                self.stop_event.wait(max(1.0, self.cfg.idle_sleep_seconds))
                continue
            if not batch_id or not jobs:
                self.stop_event.wait(max(0.5, self.cfg.idle_sleep_seconds))
                continue

            last_talker = str(jobs[0]["talker_display"])
            self.db.set_meta("last_ui_talker", last_talker)
            candidates = [
                UIMessageCandidate(
                    msg_svr_id=str(job["msg_svr_id"]),
                    talker=str(job["talker"]),
                    talker_display=str(job["talker_display"]),
                    expected_image_path=Path(str(job["expected_image_path"])),
                    thumb_path=Path(str(job["thumb_path"])),
                    create_time=float(job["create_time"]),
                )
                for job in jobs
            ]

            try:
                print(f"[UI] forcing_download | talker={last_talker} | items={len(candidates)}")
                result = self._downloader.force_download_batch(candidates)
                self.db.set_meta("last_ui_result", result.note)
                print(f"[UI] result | talker={last_talker} | ok={result.ok} | note={result.note} | resolved={len(result.resolved_msg_ids)}")
                self.db.finish_ui_batch(
                    batch_id,
                    result.resolved_msg_ids,
                    result.note,
                    self.cfg.ui_retry_backoff_seconds,
                    resolved_notes_by_msg_id=getattr(result, "resolved_sources", {}),
                )
                resolved_ids = set(result.resolved_msg_ids)
                for candidate in candidates:
                    if candidate.msg_svr_id in resolved_ids:
                        resolved_path_str = getattr(result, "resolved_media_paths", {}).get(candidate.msg_svr_id)
                        resolved_path = Path(resolved_path_str) if resolved_path_str else candidate.expected_image_path
                        if resolved_path.exists():
                            file_id = self.db.upsert_candidate(
                                resolved_path,
                                settle_seconds=1,
                                source_event="ui-force",
                                thumb_candidates_enabled=self.cfg.thumb_candidates_enabled,
                            )
                            if file_id is not None:
                                preregister_manual_order_candidate(self.db, self.media_resolver, self.cfg, file_id, resolved_path, "ui-force")
                        elif candidate.expected_image_path.exists():
                            file_id = self.db.upsert_candidate(
                                candidate.expected_image_path,
                                settle_seconds=1,
                                source_event="ui-force",
                                thumb_candidates_enabled=self.cfg.thumb_candidates_enabled,
                            )
                            if file_id is not None:
                                preregister_manual_order_candidate(self.db, self.media_resolver, self.cfg, file_id, candidate.expected_image_path, "ui-force")
            except Exception as exc:
                note = f"ui_worker_failed:{type(exc).__name__}:{exc}"
                self.db.set_meta("last_ui_result", note)
                print(f"[UI] failed | talker={last_talker} | err={note}")
                self.db.finish_ui_batch(batch_id, [], note, self.cfg.ui_retry_backoff_seconds)


def has_core_signal(fields: dict[str, Any], bank: Optional[str]) -> bool:
    if fields.get("amount") is not None:
        return True
    if bank is not None and str(bank).strip():
        return True
    if fields.get("txn_date_source") == "parsed":
        return True
    return fields.get("txn_time_source") == "parsed"


def is_manual_materialization_mode(db: StateDB, cfg: Config) -> bool:
    return not db.is_ui_force_runtime_enabled(default_enabled=cfg.ui_force_download_enabled)


def get_prior_message_order_blocker(
    db: StateDB,
    msg_ref: Optional[WeChatMessageRef],
    manual_session_started_at: Optional[float] = None,
    manual_session_id: Optional[str] = None,
) -> Optional[sqlite3.Row]:
    if msg_ref is None:
        return None
    return db.find_prior_pending_message_job(
        talker=msg_ref.talker,
        create_time=float(msg_ref.create_time or 0.0),
        msg_svr_id=msg_ref.msg_svr_id,
        manual_session_started_at=manual_session_started_at,
        manual_session_id=manual_session_id,
    )


def resolve_media_candidate(
    item: QueueItem,
    db: StateDB,
    resolver: ClientResolver,
    media_resolver: Optional[WeChatDBResolver],
    cfg: Config,
) -> Optional[MediaResolution]:
    original_path = Path(item.path)
    original_source_kind = item.source_kind
    now = time.time()
    wait_deadline = item.first_seen + float(cfg.original_wait_seconds)
    ui_force_deadline = item.first_seen + float(cfg.ui_force_delay_seconds)
    ui_force_runtime_enabled = db.is_ui_force_runtime_enabled(default_enabled=cfg.ui_force_download_enabled)
    manual_materialization_mode = not ui_force_runtime_enabled
    manual_session_started_at = db.get_manual_session_started_at() if manual_materialization_mode else None
    manual_session_id = item.manual_session_id

    if original_source_kind == "temp_image":
        context_path_str = db.find_recent_msgattach_context_path(
            item.mtime,
            lookback_sec=cfg.temp_correlation_seconds,
            lookahead_sec=15,
            limit=48,
        )
        if context_path_str is None:
            context_path_str = db.find_recent_unresolved_msgattach_context_path(
                max_age_sec=max(cfg.original_wait_seconds * 4, 1800),
                limit=60,
            )
        if context_path_str is None:
            if now < wait_deadline:
                db.mark_hold(
                    item.file_id,
                    reason="WAITING_TEMP_CONTEXT",
                    delay_sec=hold_retry_delay_seconds(now, wait_deadline),
                )
                print(f"[HOLD] {original_path.name} | waiting_temp_context")
            else:
                db.mark_done(item.file_id, sha256="", processed_at=now)
                print(f"[SKIP] {original_path.name} | temp_without_unique_context")
            return None

        client_source_path = Path(context_path_str)
        group_hash = extract_group_id_from_path(client_source_path)
        context_row = db.get_latest_file_row_by_path(client_source_path)
        context_mtime = float(context_row["mtime"]) if context_row is not None else item.mtime
        msg_ref = None
        if media_resolver is not None:
            msg_ref = media_resolver.find_message_for_path(client_source_path, context_mtime)
            if msg_ref is None and group_hash:
                msg_ref = media_resolver.find_unique_message_for_group(
                    group_hash,
                    context_mtime,
                    max(cfg.original_wait_seconds * 4, 1800),
                )
        resolved_client_source = msg_ref.preferred_context_path() if msg_ref is not None else None
        tracked_job = ensure_message_job_tracking(
            db=db,
            resolver=resolver,
            media_resolver=media_resolver,
            cfg=cfg,
            msg_ref=msg_ref,
            thumb_path=msg_ref.thumb_abs_path if msg_ref is not None else None,
            client_source_path=resolved_client_source or client_source_path,
            first_seen_at=item.first_seen,
            manual_session_id=manual_session_id,
        )
        order_blocker = get_prior_message_order_blocker(
            db,
            msg_ref,
            manual_session_started_at=manual_session_started_at,
            manual_session_id=manual_session_id,
        )
        if order_blocker is not None:
            blocker_id = str(order_blocker["msg_svr_id"])
            db.mark_hold(item.file_id, reason=f"WAITING_SESSION_PRIOR_MESSAGE_ORDER:{blocker_id}", delay_sec=10)
            print(f"[HOLD] {original_path.name} | waiting_session_prior_message_order | blocker={blocker_id}")
            return None
        if tracked_job is not None and msg_ref is not None and msg_ref.image_abs_path is not None and msg_ref.image_abs_path.exists():
            return MediaResolution(
                original_source_path=original_path,
                original_source_kind=original_source_kind,
                resolved_path=msg_ref.image_abs_path,
                resolved_source_kind=detect_source_kind(msg_ref.image_abs_path),
                client_source_path=resolved_client_source or client_source_path,
                resolution_source="db_image",
                verification_status="CONFIRMADO",
                msg_ref=msg_ref,
                using_thumb_fallback=False,
            )

        if manual_materialization_mode and cfg.thumb_candidates_enabled:
            if tracked_job is not None:
                db.set_message_job_state(
                    msg_ref.msg_svr_id if msg_ref is not None else None,
                    "WAITING_ORIGINAL",
                    note="MANUAL_WAIT_ORIGINAL",
                    next_ui_attempt_at=0.0,
                    reset_batch=True,
                )
            db.mark_hold(
                item.file_id,
                reason="MANUAL_WAIT_ORIGINAL",
                delay_sec=hold_retry_delay_seconds(now, wait_deadline),
            )
            print(f"[HOLD] {original_path.name} | manual_wait_original_from_temp")
            return None

        return MediaResolution(
            original_source_path=original_path,
            original_source_kind=original_source_kind,
            resolved_path=original_path,
            resolved_source_kind=detect_source_kind(original_path),
            client_source_path=resolved_client_source or client_source_path,
            resolution_source="temp_preview",
            verification_status="TEMP_PREVIEW",
            msg_ref=msg_ref,
            using_thumb_fallback=False,
        )

    msg_ref = media_resolver.find_message_for_path(original_path, item.mtime) if media_resolver is not None else None
    preferred_context = msg_ref.preferred_context_path() if msg_ref is not None else None
    client_source_path = preferred_context if preferred_context is not None else original_path
    tracked_job = ensure_message_job_tracking(
        db=db,
        resolver=resolver,
        media_resolver=media_resolver,
        cfg=cfg,
        msg_ref=msg_ref,
        thumb_path=original_path if original_source_kind == "msgattach_thumb_dat" else msg_ref.thumb_abs_path if msg_ref is not None else None,
        client_source_path=client_source_path,
        first_seen_at=item.first_seen,
        manual_session_id=manual_session_id,
    )
    order_blocker = get_prior_message_order_blocker(
        db,
        msg_ref,
        manual_session_started_at=manual_session_started_at,
        manual_session_id=manual_session_id,
    )
    if order_blocker is not None:
        blocker_id = str(order_blocker["msg_svr_id"])
        db.mark_hold(item.file_id, reason=f"WAITING_SESSION_PRIOR_MESSAGE_ORDER:{blocker_id}", delay_sec=10)
        print(f"[HOLD] {original_path.name} | waiting_session_prior_message_order | blocker={blocker_id}")
        return None

    if original_source_kind == "msgattach_thumb_dat":
        if msg_ref is not None and msg_ref.image_abs_path is not None and msg_ref.image_abs_path.exists():
            resolved = msg_ref.image_abs_path
            print(f"[INFO] {original_path.name} -> using_db_image={resolved.name}")
            return MediaResolution(
                original_source_path=original_path,
                original_source_kind=original_source_kind,
                resolved_path=resolved,
                resolved_source_kind=detect_source_kind(resolved),
                client_source_path=client_source_path,
                resolution_source="db_image",
                verification_status="CONFIRMADO",
                msg_ref=msg_ref,
                using_thumb_fallback=False,
            )

        sibling_image = resolve_full_image_from_thumb_path(original_path)
        if sibling_image is not None and sibling_image.exists():
            print(f"[INFO] {original_path.name} -> using_sibling_image={sibling_image.name}")
            return MediaResolution(
                original_source_path=original_path,
                original_source_kind=original_source_kind,
                resolved_path=sibling_image,
                resolved_source_kind=detect_source_kind(sibling_image),
                client_source_path=client_source_path,
                resolution_source="path_sibling_image",
                verification_status="CONFIRMADO",
                msg_ref=msg_ref,
                using_thumb_fallback=False,
            )

        if manual_materialization_mode:
            if tracked_job is not None:
                db.set_message_job_state(
                    msg_ref.msg_svr_id if msg_ref is not None else None,
                    "WAITING_ORIGINAL",
                    note="MANUAL_WAIT_ORIGINAL",
                    next_ui_attempt_at=0.0,
                    reset_batch=True,
                )
            db.mark_hold(
                item.file_id,
                reason="MANUAL_WAIT_ORIGINAL",
                delay_sec=hold_retry_delay_seconds(now, wait_deadline, minimum=2, maximum=5) if now < wait_deadline else 8,
            )
            print(f"[HOLD] {original_path.name} | manual_wait_original")
            return None

        tracked_state = str(tracked_job["state"] or "") if tracked_job is not None else ""
        if tracked_job is not None and tracked_state == "UI_FORCE_RUNNING" and now < wait_deadline:
            if ui_force_runtime_enabled:
                db.mark_hold(item.file_id, reason="WAITING_UI_FORCE_DOWNLOAD", delay_sec=10)
                print(f"[HOLD] {original_path.name} | waiting_ui_force_download_running")
                return None
            db.set_message_job_state(
                msg_ref.msg_svr_id if msg_ref is not None else None,
                "WAITING_ORIGINAL",
                note="WAITING_ORIGINAL_MEDIA",
                next_ui_attempt_at=0.0,
                reset_batch=True,
            )

        if tracked_job is not None and ui_force_runtime_enabled and now < ui_force_deadline:
            db.set_message_job_state(msg_ref.msg_svr_id if msg_ref is not None else None, "WAITING_ORIGINAL", note="WAITING_ORIGINAL_MEDIA", next_ui_attempt_at=0.0, reset_batch=True)
            db.mark_hold(
                item.file_id,
                reason="WAITING_ORIGINAL_MEDIA",
                delay_sec=hold_retry_delay_seconds(now, ui_force_deadline),
            )
            print(f"[HOLD] {original_path.name} | waiting_original_media")
            return None

        if tracked_job is not None and ui_force_runtime_enabled and now < wait_deadline:
            remaining = max(5, min(15, int(wait_deadline - now)))
            db.set_message_job_state(msg_ref.msg_svr_id if msg_ref is not None else None, "UI_FORCE_PENDING", note="WAITING_UI_FORCE_DOWNLOAD", next_ui_attempt_at=0.0, reset_batch=True)
            db.mark_hold(item.file_id, reason="WAITING_UI_FORCE_DOWNLOAD", delay_sec=remaining)
            print(f"[HOLD] {original_path.name} | waiting_ui_force_download")
            return None

        if now < wait_deadline:
            db.mark_hold(
                item.file_id,
                reason="WAITING_ORIGINAL_MEDIA",
                delay_sec=hold_retry_delay_seconds(now, wait_deadline),
            )
            print(f"[HOLD] {original_path.name} | waiting_original_media")
            return None

        if tracked_job is not None and ui_force_runtime_enabled and tracked_state != "UI_FORCE_RUNNING":
            db.set_message_job_state(
                msg_ref.msg_svr_id if msg_ref is not None else None,
                "UI_FORCE_PENDING",
                note="WAITING_ORIGINAL_MEDIA",
                next_ui_attempt_at=0.0,
                reset_batch=True,
            )
        elif tracked_job is not None:
            db.set_message_job_state(
                msg_ref.msg_svr_id if msg_ref is not None else None,
                "WAITING_ORIGINAL",
                note="WAITING_ORIGINAL_MEDIA",
                next_ui_attempt_at=0.0,
                reset_batch=True,
            )
        db.mark_hold(item.file_id, reason="WAITING_ORIGINAL_MEDIA", delay_sec=8)
        print(f"[HOLD] {original_path.name} | waiting_original_media_no_thumb_fallback")
        return None

    resolved_path = original_path
    resolved_source_kind = original_source_kind
    resolution_source = "direct_image"
    if msg_ref is not None and msg_ref.image_abs_path is not None and msg_ref.image_abs_path.exists():
        resolved_path = msg_ref.image_abs_path
        resolved_source_kind = detect_source_kind(resolved_path)
        resolution_source = "db_image"

    return MediaResolution(
        original_source_path=original_path,
        original_source_kind=original_source_kind,
        resolved_path=resolved_path,
        resolved_source_kind=resolved_source_kind,
        client_source_path=client_source_path,
        resolution_source=resolution_source,
        verification_status="CONFIRMADO",
        msg_ref=msg_ref,
        using_thumb_fallback=False,
    )


def process_item(
    item: QueueItem,
    db: StateDB,
    sink: RowSink,
    ocr: OCREngine,
    resolver: ClientResolver,
    media_resolver: Optional[WeChatDBResolver],
    cfg: Config,
) -> None:
    if db.receipt_exists(item.file_id):
        db.mark_done(item.file_id, sha256="", processed_at=time.time())
        return

    claim_started_at = time.time()
    manual_open_to_claim_ms = wall_duration_ms(item.first_seen, claim_started_at)
    path = Path(item.path)
    resolution: Optional[MediaResolution] = None
    msg_svr_id: Optional[str] = None
    active_media_resolver = runtime_media_resolver(media_resolver)
    try:
        resolution = resolve_media_candidate(
            item=item,
            db=db,
            resolver=resolver,
            media_resolver=active_media_resolver,
            cfg=cfg,
        )
        if resolution is None:
            return

        msg_svr_id = resolution.msg_ref.msg_svr_id if resolution.msg_ref is not None else None
        if db.receipt_msg_exists(msg_svr_id):
            db.mark_done(item.file_id, sha256="", processed_at=time.time(), note="RESOLVED_BY_LATER_SUCCESS")
            db.resolve_message_job_paths(msg_svr_id, exclude_file_id=item.file_id, sha256="")
            db.mark_message_job_resolved(msg_svr_id, note="DUPLICATE_MSG_SVR_ID")
            path = resolution.resolved_path
            print(f"[SKIP] {path.name} | duplicate_msg_svr_id={msg_svr_id}")
            return

        path = resolution.resolved_path
        if not path.exists():
            if item.source_kind == "temp_image" and db.message_job_is_terminal(msg_svr_id):
                db.mark_done(item.file_id, sha256="", processed_at=time.time(), note="RESOLVED_BY_LATER_SUCCESS")
                print(f"[SKIP] {path.name} | temp_resolved_by_later_success")
                return
            raise FileNotFoundError(f"Resolved file disappeared: {path}")

        if should_ignore_sender(resolution.msg_ref):
            sender_label = (
                resolution.msg_ref.sender_display
                or resolution.msg_ref.sender_user_name
                or resolution.msg_ref.talker
                or "unknown"
            )
            db.mark_done(item.file_id, sha256="", processed_at=time.time(), note=f"IGNORED_SENDER:{sender_label}")
            db.mark_message_job_resolved(msg_svr_id, note=f"IGNORED_SENDER:{sender_label}")
            print(f"[SKIP] {path.name} | ignored_sender={sender_label}")
            return

        ignore_reason = resolver.ignore_reason(resolution.client_source_path)
        if ignore_reason:
            db.mark_done(item.file_id, sha256="", processed_at=time.time(), note=ignore_reason)
            db.mark_message_job_resolved(msg_svr_id, note=ignore_reason)
            print(f"[SKIP] {path.name} | {ignore_reason}")
            return

        client = resolver.resolve(resolution.client_source_path)
        if not client:
            gid = extract_group_id_from_path(resolution.client_source_path) or "SEM_GRUPO"
            db.mark_hold(item.file_id, reason=f"MISSING_CLIENT_MAP:{gid}", delay_sec=120)
            print(f"[HOLD] {path.name} | grupo_sem_mapa={gid}")
            return

        open_started_at = time.perf_counter()
        img, img_bytes, _ext, _key = open_image_from_file(path)
        open_ms = perf_duration_ms(open_started_at)
        digest = sha256_bytes(img_bytes)
        # The client may legitimately send the same receipt image more than once.
        # Keep binary hashes for traceability, but do not use them to suppress ingest.
        q_score = quality_score(img)

        prep_started_at = time.perf_counter()
        img_for_ocr = prepare_image_for_ocr(img, resolution.resolved_source_kind)
        prep_ms = perf_duration_ms(prep_started_at)
        ocr_started_at = time.perf_counter()
        text, ocr_conf = ocr.extract(img_for_ocr)
        ocr_ms = perf_duration_ms(ocr_started_at)
        ocr_chars = len(text)
        is_receipt, receipt_reason = looks_like_single_receipt(text)
        if not is_receipt:
            db.mark_done(item.file_id, sha256=digest, processed_at=time.time())
            db.mark_message_job_resolved(msg_svr_id, note=f"NOT_RECEIPT:{receipt_reason}")
            print(f"[SKIP] {path.name} | not_receipt={receipt_reason}")
            return

        parse_started_at = time.perf_counter()
        fields = parse_receipt_fields(text, ocr_conf=ocr_conf, q_score=q_score)
        bank = fields.get("bank")
        if bank is None:
            bank = detect_bank(f"{text}\n{client}", fields.get("beneficiary"))
            fields["bank"] = bank
        if not has_core_signal(fields, bank):
            db.mark_exception(item.file_id, reason="EXCEPTION_MISSING_CORE_FIELDS")
            db.set_meta("last_exception_reason", "EXCEPTION_MISSING_CORE_FIELDS")
            db.mark_message_job_exception(msg_svr_id, note="EXCEPTION_MISSING_CORE_FIELDS")
            print(f"[EXCEPTION] {path.name} | missing_core_fields | verification={resolution.verification_status}")
            return

        review_needed = compute_review_needed(
            fields=fields,
            bank=bank,
            quality_score_value=q_score,
            verification_status=resolution.verification_status,
            min_confidence=cfg.min_confidence,
            resolution_source=resolution.resolution_source,
        )
        parse_ms = perf_duration_ms(parse_started_at)
        ingested_at = time.time()

        payload: dict[str, Any] = {
            "file_id": item.file_id,
            "source_path": str(resolution.original_source_path),
            "source_kind": resolution.original_source_kind,
            "ingested_at": ingested_at,
            "sha256": digest,
            "txn_date": fields["txn_date"],
            "txn_time": fields["txn_time"],
            "txn_date_source": fields.get("txn_date_source"),
            "txn_time_source": fields.get("txn_time_source"),
            "client": client,
            "bank": bank,
            "beneficiary": fields["beneficiary"],
            "amount": fields["amount"],
            "amount_raw": fields.get("amount_raw"),
            "amount_rounded": fields.get("amount_rounded"),
            "amount_source": fields.get("amount_source"),
            "currency": fields["currency"],
            "parse_conf": fields["parse_conf"],
            "quality_score": q_score,
            "ocr_engine": ocr.name,
            "ocr_conf": ocr_conf,
            "ocr_chars": ocr_chars,
            "review_needed": review_needed,
            "ocr_text": text[:25000] if text else "",
            "parser_json": json.dumps(fields, ensure_ascii=False),
            "msg_svr_id": msg_svr_id,
            "talker": resolution.msg_ref.talker if resolution.msg_ref is not None else None,
            "msg_create_time": resolution.msg_ref.create_time if resolution.msg_ref is not None else None,
            "manual_session_id": item.manual_session_id,
            "resolved_media_path": str(path),
            "resolution_source": resolution.resolution_source,
            "verification_status": resolution.verification_status,
        }

        row_payload = build_sheet_payload_from_receipt(payload)
        payload["sheet_status"] = "SINK_PENDING"
        payload["sheet_payload_json"] = json.dumps(row_payload, ensure_ascii=False)
        payload["sheet_next_attempt"] = 0.0
        payload["sheet_last_error"] = None
        payload["sheet_committed_at"] = None
        payload["excel_sheet"] = None
        payload["excel_row"] = None
        db.insert_receipt(payload)
        db.mark_done(item.file_id, sha256=digest, processed_at=time.time())
        db.resolve_related_file_paths(
            source_path=resolution.original_source_path,
            exclude_file_id=item.file_id,
            sha256=digest,
        )
        db.resolve_message_job_paths(msg_svr_id, exclude_file_id=item.file_id, sha256=digest)
        if resolution.verification_status == "THUMB_FALLBACK":
            db.mark_message_job_thumb_fallback(msg_svr_id, note=resolution.resolution_source)
        else:
            db.mark_message_job_resolved(msg_svr_id, note=resolution.resolution_source)
        db.set_meta("last_resolution_source", resolution.resolution_source)
        db.set_meta("last_verification_status", resolution.verification_status)

        print(
            f"[OK] {path.name} | cliente={client} | banco={bank} | valor={fields['amount']} "
            f"| data={fields['txn_date']} {fields['txn_time']} | sink=staged "
            f"| resolution={resolution.resolution_source} | verification={resolution.verification_status} "
            f"| manual_open_to_claim_ms={format_ms(manual_open_to_claim_ms)} "
            f"| open_ms={format_ms(open_ms)} | prep_ms={format_ms(prep_ms)} "
            f"| ocr_ms={format_ms(ocr_ms)} | parse_ms={format_ms(parse_ms)} "
            f"| manual_open_to_ingest_ms={format_ms(wall_duration_ms(item.first_seen, ingested_at))}"
        )

    except Exception as exc:
        if isinstance(exc, FileNotFoundError) and item.source_kind == "temp_image" and db.message_job_is_terminal(msg_svr_id):
            db.mark_done(item.file_id, sha256="", processed_at=time.time(), note="RESOLVED_BY_LATER_SUCCESS")
            print(f"[SKIP] {path.name} | temp_missing_after_resolution")
            return
        if isinstance(exc, FileNotFoundError) and item.source_kind == "temp_image" and item.attempts >= 3:
            db.mark_done(item.file_id, sha256="", processed_at=time.time(), note="STALE_TEMP_ORPHAN")
            print(f"[SKIP] {path.name} | temp_file_disappeared_after_{item.attempts}_attempts")
            return
        fast_retry = 5 if isinstance(exc, PermissionError) else None
        db.mark_retry(
            file_id=item.file_id,
            attempts=item.attempts,
            retry_base_sec=cfg.retry_base_seconds,
            err=f"{type(exc).__name__}: {exc}",
            max_retries=cfg.max_retries,
            delay_override_sec=fast_retry,
        )
        print(f"[RETRY] {path.name} | attempt={item.attempts} | err={type(exc).__name__}: {exc}")


def flush_ready_sink_rows(
    db: StateDB,
    sink: RowSink,
    cfg: Config,
    media_resolver: Optional[WeChatDBResolver] = None,
    max_rows: int = 50,
) -> int:
    committed = 0
    limit = max(1, int(max_rows))
    seeded_placeholders = seed_ready_manual_session_placeholders(db, media_resolver, cfg)
    if seeded_placeholders:
        print(f"[SESSION] seeded_placeholders={seeded_placeholders}")
    manual_session_started_at = db.get_manual_session_started_at() if is_manual_materialization_mode(db, cfg) else None
    manual_session_id = db.get_current_manual_session_id() if is_manual_materialization_mode(db, cfg) else None
    for _ in range(limit):
        claimed = db.claim_next_sink_receipt(
            sheet_order_scope=cfg.sheet_order_scope,
            commit_order=cfg.sheet_commit_order,
            manual_session_started_at=manual_session_started_at,
            manual_session_id=manual_session_id,
        )
        if claimed is None:
            break

        file_id = str(claimed["file_id"])
        msg_svr_id = str(claimed.get("msg_svr_id") or "").strip() or "-"
        talker = str(claimed.get("talker") or "").strip() or "-"
        msg_create_time = float(claimed.get("msg_create_time") or 0.0)
        ingested_at = float(claimed.get("ingested_at") or 0.0)
        source_first_seen = float(claimed.get("source_first_seen") or 0.0)
        try:
            sheet, row = sink.append(claimed["row_payload"], review_needed=bool(claimed["review_needed"]))
        except Exception as exc:
            err = f"{type(exc).__name__}: {exc}"
            db.mark_receipt_sink_retry(file_id, err, delay_sec=cfg.retry_base_seconds)
            print(f"[SINK] retry | file_id={file_id} | msg={msg_svr_id} | talker={talker} | err={err}")
            break

        committed_at = time.time()
        db.mark_receipt_sink_committed(file_id, sheet, row, committed_at=committed_at)
        committed += 1
        print(
            f"[SINK] committed | file_id={file_id} | msg={msg_svr_id} | talker={talker} "
            f"| create_time={msg_create_time:.0f} | sheet={sheet} | row={row} "
            f"| ingest_to_commit_ms={format_ms(wall_duration_ms(ingested_at, committed_at))} "
            f"| manual_open_to_commit_ms={format_ms(wall_duration_ms(source_first_seen, committed_at))}"
        )
    return committed


def backfill_missing_receipt_fields(db: StateDB, sink: RowSink, cfg: Config, limit: int = 5000) -> tuple[int, int, int]:
    rows = db.list_receipts_needing_parser_backfill(limit=limit)
    if not rows:
        return (0, 0, 0)

    updated = 0
    sheet_updated = 0
    sheet_failed = 0
    for row in rows:
        ocr_text = str(row["ocr_text"] or "")
        quality_score_value = float(row["quality_score"] or 0.0)
        parsed = parse_receipt_fields(
            ocr_text,
            ocr_conf=float(row["ocr_conf"] or 0.0),
            q_score=quality_score_value,
        )

        client = row["client"]
        bank = row["bank"] or parsed.get("bank")
        if not bank:
            bank = detect_bank(f"{ocr_text}\n{client or ''}", row["beneficiary"])

        amount = parsed.get("amount")
        if amount is None and row["amount"] is not None:
            amount = float(row["amount"])
        amount_raw = parsed.get("amount_raw") or row["amount_raw"]
        amount_rounded = parsed.get("amount_rounded")
        if amount_rounded is None:
            if row["amount_rounded"] is not None:
                amount_rounded = float(row["amount_rounded"])
            else:
                amount_rounded = round_amount_for_output(amount)
        amount_source = parsed.get("amount_source")
        if not amount_source or amount_source == "missing":
            amount_source = str(row["amount_source"] or "").strip() or "missing"
        currency = parsed.get("currency") or row["currency"]
        if currency is None and amount is not None:
            currency = "BRL"

        merged_fields = dict(parsed)
        merged_fields["bank"] = bank
        merged_fields["amount"] = amount
        merged_fields["amount_raw"] = amount_raw
        merged_fields["amount_rounded"] = amount_rounded
        merged_fields["amount_source"] = amount_source
        merged_fields["currency"] = currency

        review_needed = bool(row["review_needed"]) or compute_review_needed(
            fields=merged_fields,
            bank=bank,
            quality_score_value=quality_score_value,
            verification_status=row["verification_status"],
            min_confidence=cfg.min_confidence,
            resolution_source=row["resolution_source"],
        )

        existing_sheet_payload: dict[str, Any] = {}
        payload_json = str(row["sheet_payload_json"] or "").strip()
        if payload_json:
            try:
                decoded = json.loads(payload_json)
                if isinstance(decoded, dict):
                    existing_sheet_payload = decoded
            except Exception:
                existing_sheet_payload = {}

        receipt_payload = {
            "file_id": row["file_id"],
            "client": client,
            "txn_date": merged_fields.get("txn_date"),
            "txn_time": merged_fields.get("txn_time"),
            "bank": bank,
            "amount": amount,
            "amount_rounded": amount_rounded,
            "verification_status": row["verification_status"],
            "msg_svr_id": row["msg_svr_id"],
            "talker": row["talker"],
        }
        sheet_payload = build_sheet_payload_from_receipt(receipt_payload, existing_sheet_payload)

        db.update_receipt_parser_backfill(
            str(row["file_id"]),
            txn_date=merged_fields.get("txn_date"),
            txn_time=merged_fields.get("txn_time"),
            txn_date_source=merged_fields.get("txn_date_source"),
            txn_time_source=merged_fields.get("txn_time_source"),
            amount=amount,
            amount_raw=amount_raw,
            amount_rounded=amount_rounded,
            amount_source=amount_source,
            currency=currency,
            bank=bank,
            parse_conf=merged_fields.get("parse_conf"),
            review_needed=review_needed,
            parser_json=json.dumps(merged_fields, ensure_ascii=False),
            sheet_payload_json=json.dumps(sheet_payload, ensure_ascii=False),
        )
        updated += 1

        # We intentionally do NOT update the sheet during backfill if it's already SINK_COMMITTED.
        # This prevents overwriting manual user corrections or changing the 'Banco' field 
        # unexpectedly during startup when re-parsing happens.
        # Original logic removed:
        # row_idx = row["excel_row"]
        # if str(row["sheet_status"] or "").strip() == "SINK_COMMITTED" and row_idx is not None:
        #    try:
        #        sink.update_row(...)


    return (updated, sheet_updated, sheet_failed)


def default_watch_roots() -> list[Path]:
    home = Path(os.environ.get("USERPROFILE", str(Path.home())))
    base = home / "Documents" / "WeChat Files"
    roots: list[Path] = []
    if base.exists():
        for sub in base.iterdir():
            fs = sub / "FileStorage"
            if fs.exists():
                roots.append(fs)
    return roots


def parse_boolish(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def parse_retry_backoff_seconds(value: Any) -> list[int]:
    if isinstance(value, list):
        raw_parts = value
    else:
        raw_parts = re.split(r"[\s,;]+", str(value or "").strip())

    out: list[int] = []
    for part in raw_parts:
        try:
            ivalue = int(part)
        except Exception:
            continue
        if ivalue > 0:
            out.append(ivalue)
    return out or [5, 10, 20, 40]


def parse_token_list(value: Any, default: list[str]) -> list[str]:
    if isinstance(value, list):
        raw_parts = value
    else:
        raw_parts = re.split(r"[\s,;]+", str(value or "").strip())

    out: list[str] = []
    seen: set[str] = set()
    for part in raw_parts:
        token = str(part or "").strip()
        if not token:
            continue
        key = token.lower()
        if key in seen:
            continue
        out.append(token)
        seen.add(key)
    return out or list(default)


def ensure_client_map_file(map_path: Path, watch_roots: list[Path]) -> None:
    if map_path.exists():
        return
    map_path.parent.mkdir(parents=True, exist_ok=True)

    discovered: dict[str, str] = {}
    for root in watch_roots:
        msgattach = root / "MsgAttach"
        if not msgattach.exists():
            continue
        for sub in msgattach.iterdir():
            if not sub.is_dir():
                continue
            gid = sub.name.strip()
            if gid and gid.lower() not in discovered:
                discovered[gid.lower()] = ""
            if len(discovered) >= 30:
                break
        if len(discovered) >= 30:
            break

    template: dict[str, str] = {
        "COLE_AQUI_ID_DO_GRUPO": "NOME_DO_CLIENTE",
    }
    for gid in sorted(discovered.keys()):
        template[gid] = ""
    map_path.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="WeChat receipt ingestion daemon")
    p.add_argument("--watch-root", action="append", default=[], help="Root directory to monitor (repeatable)")
    p.add_argument("--db-path", default=str(Path.cwd() / "wechat_receipt_state.db"))
    p.add_argument("--excel-path", default=str(Path.cwd() / "pagamentos_wechat.xlsx"))
    p.add_argument("--sink-mode", choices=("excel", "google-sheets"), default=os.getenv("WECHAT_SINK_MODE", "excel"))
    p.add_argument("--gsheet-ref", default=os.getenv("WECHAT_GSHEET_REF", ""))
    p.add_argument("--gsheet-worksheet", default=os.getenv("WECHAT_GSHEET_WORKSHEET", ""))
    p.add_argument("--gsheet-review-worksheet", default=os.getenv("WECHAT_GSHEET_REVIEW_WORKSHEET", ""))
    p.add_argument("--google-credentials-path", default=os.getenv("WECHAT_GOOGLE_CREDENTIALS_PATH", ""))
    p.add_argument("--verification-column-name", default=os.getenv("WECHAT_VERIFICATION_COLUMN_NAME", DEFAULT_VERIFICATION_COLUMN_NAME))
    p.add_argument("--client-map-path", default=str(Path.cwd() / "clientes_grupos.json"))
    p.add_argument("--resolution-mode", choices=("path-only", "db-first"), default=os.getenv("WECHAT_RESOLUTION_MODE", "db-first"))
    p.add_argument("--db-merge-path", default=os.getenv("WECHAT_DB_MERGE_PATH", str(Path.cwd() / ".runtime" / "wechat_merge.db")))
    p.add_argument("--settle-seconds", type=int, default=1)
    p.add_argument("--reconcile-seconds", type=int, default=90)
    p.add_argument("--recent-files-hours", type=int, default=int(os.getenv("WECHAT_RECENT_FILES_HOURS", "24")))
    p.add_argument("--idle-sleep-seconds", type=float, default=1.2)
    p.add_argument("--retry-base-seconds", type=int, default=30)
    p.add_argument("--min-confidence", type=float, default=0.55)
    p.add_argument("--max-retries", type=int, default=0, help="0 means infinite retries")
    p.add_argument("--original-wait-seconds", type=int, default=int(os.getenv("WECHAT_ORIGINAL_WAIT_SECONDS", "90")))
    p.add_argument("--temp-correlation-seconds", type=int, default=int(os.getenv("WECHAT_TEMP_CORRELATION_SECONDS", "30")))
    p.add_argument("--thumb-candidates-enabled", default=os.getenv("WECHAT_THUMB_CANDIDATES_ENABLED", "false"))
    p.add_argument("--manual-order-guard-enabled", default=os.getenv("WECHAT_MANUAL_ORDER_GUARD_ENABLED", "true"))
    p.add_argument("--manual-burst-gap-seconds", type=int, default=int(os.getenv("WECHAT_MANUAL_BURST_GAP_SECONDS", "2")))
    p.add_argument("--manual-burst-max-seconds", type=int, default=int(os.getenv("WECHAT_MANUAL_BURST_MAX_SECONDS", "8")))
    p.add_argument("--ui-force-download-enabled", default=os.getenv("WECHAT_UI_FORCE_DOWNLOAD_ENABLED", "false"))
    p.add_argument("--ui-force-delay-seconds", type=int, default=int(os.getenv("WECHAT_UI_FORCE_DELAY_SECONDS", "15")))
    p.add_argument("--ui-force-scope", default=os.getenv("WECHAT_UI_FORCE_SCOPE", "mapped-groups"))
    p.add_argument("--ui-focus-policy", default=os.getenv("WECHAT_UI_FOCUS_POLICY", "immediate"))
    p.add_argument("--ui-batch-mode", default=os.getenv("WECHAT_UI_BATCH_MODE", "group-sequential"))
    p.add_argument("--ui-item-timeout-seconds", type=int, default=int(os.getenv("WECHAT_UI_ITEM_TIMEOUT_SECONDS", "5")))
    p.add_argument("--ui-retry-backoff-seconds", default=os.getenv("WECHAT_UI_RETRY_BACKOFF_SECONDS", "5,10,20,40"))
    p.add_argument("--ui-window-backends", default=os.getenv("WECHAT_UI_WINDOW_BACKENDS", "win32,uia"))
    p.add_argument("--ui-window-classes", default=os.getenv("WECHAT_UI_WINDOW_CLASSES", "WeChatMainWndForPC,Base_PowerMessageWindow,Chrome_WidgetWin_0"))
    p.add_argument("--sheet-order-scope", default=os.getenv("WECHAT_SHEET_ORDER_SCOPE", "per_talker"))
    p.add_argument("--sheet-materialization-order", default=os.getenv("WECHAT_SHEET_MATERIALIZATION_ORDER", "desc"))
    p.add_argument("--sheet-commit-order", default=os.getenv("WECHAT_SHEET_COMMIT_ORDER", "asc"))
    p.add_argument("--disable-watchdog", action="store_true")
    return p.parse_args()


def build_config(args: argparse.Namespace) -> Config:
    roots = [Path(r) for r in args.watch_root] if args.watch_root else default_watch_roots()
    return Config(
        watch_roots=roots,
        db_path=Path(args.db_path),
        db_merge_path=Path(args.db_merge_path),
        excel_path=Path(args.excel_path),
        sink_mode=str(args.sink_mode).strip().lower(),
        gsheet_ref=(str(args.gsheet_ref).strip() or None),
        gsheet_worksheet=(str(args.gsheet_worksheet).strip() or None),
        gsheet_review_worksheet=(str(args.gsheet_review_worksheet).strip() or None),
        google_credentials_path=(Path(args.google_credentials_path) if str(args.google_credentials_path).strip() else None),
        verification_column_name=(str(args.verification_column_name).strip() or DEFAULT_VERIFICATION_COLUMN_NAME),
        client_map_path=Path(args.client_map_path),
        resolution_mode=(str(args.resolution_mode).strip().lower() or "db-first"),
        settle_seconds=max(1, args.settle_seconds),
        reconcile_seconds=max(20, args.reconcile_seconds),
        recent_files_hours=max(1, int(args.recent_files_hours)),
        idle_sleep_seconds=max(0.2, args.idle_sleep_seconds),
        retry_base_seconds=max(10, args.retry_base_seconds),
        min_confidence=max(0.0, min(1.0, args.min_confidence)),
        max_retries=max(0, args.max_retries),
        original_wait_seconds=max(5, int(args.original_wait_seconds)),
        temp_correlation_seconds=max(5, int(args.temp_correlation_seconds)),
        thumb_candidates_enabled=parse_boolish(args.thumb_candidates_enabled, default=False),
        manual_order_guard_enabled=parse_boolish(args.manual_order_guard_enabled, default=True),
        manual_burst_gap_seconds=max(1, int(args.manual_burst_gap_seconds)),
        manual_burst_max_seconds=max(max(1, int(args.manual_burst_gap_seconds)), int(args.manual_burst_max_seconds)),
        ui_force_download_enabled=parse_boolish(args.ui_force_download_enabled, default=False),
        ui_force_delay_seconds=max(5, int(args.ui_force_delay_seconds)),
        ui_force_scope=(str(args.ui_force_scope).strip().lower() or "mapped-groups"),
        ui_focus_policy=(str(args.ui_focus_policy).strip().lower() or "immediate"),
        ui_batch_mode=(str(args.ui_batch_mode).strip().lower() or "group-sequential"),
        ui_item_timeout_seconds=max(1, int(args.ui_item_timeout_seconds)),
        ui_retry_backoff_seconds=parse_retry_backoff_seconds(args.ui_retry_backoff_seconds),
        ui_window_backends=[token.lower() for token in parse_token_list(args.ui_window_backends, ["win32", "uia"])],
        ui_window_classes=parse_token_list(
            args.ui_window_classes,
            ["WeChatMainWndForPC", "Base_PowerMessageWindow", "Chrome_WidgetWin_0"],
        ),
        sheet_order_scope=(str(args.sheet_order_scope).strip().lower() or "per_talker"),
        sheet_materialization_order=(str(args.sheet_materialization_order).strip().lower() or "desc"),
        sheet_commit_order=(str(args.sheet_commit_order).strip().lower() or "asc"),
        disable_watchdog=bool(args.disable_watchdog),
    )


def build_sink(cfg: Config) -> RowSink:
    if cfg.sink_mode == "google-sheets":
        if not cfg.gsheet_ref:
            raise RuntimeError("Google Sheets mode requires --gsheet-ref with the sheet URL or spreadsheet id")
        return GoogleSheetsSink(
            spreadsheet_ref=cfg.gsheet_ref,
            credentials_path=cfg.google_credentials_path,
            main_worksheet=cfg.gsheet_worksheet,
            review_worksheet=cfg.gsheet_review_worksheet,
            verification_column_name=cfg.verification_column_name,
        )
    return ExcelSink(cfg.excel_path, verification_column_name=cfg.verification_column_name)


def main() -> int:
    args = parse_args()
    cfg = build_config(args)

    # Avoid Windows cp1252 crashes when group names contain non-Latin chars.
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not cfg.watch_roots:
        print("No watch roots found. Pass --watch-root explicitly.")
        return 2

    print("Watch roots:")
    for r in cfg.watch_roots:
        print(f" - {r}")
    print(f"DB: {cfg.db_path}")
    print(f"Sink mode: {cfg.sink_mode}")
    if cfg.sink_mode == "excel":
        print(f"Excel: {cfg.excel_path}")
    else:
        print(f"Google Sheets ref: {cfg.gsheet_ref}")
        if cfg.google_credentials_path:
            print(f"Google credentials: {cfg.google_credentials_path}")
    print(f"Client map: {cfg.client_map_path}")
    print(f"Recent files window (hours): {cfg.recent_files_hours}")
    print(f"Resolution mode: {cfg.resolution_mode}")
    print(f"DB merge path: {cfg.db_merge_path}")
    print(f"Settle (seconds): {cfg.settle_seconds}")
    print(f"Original wait (seconds): {cfg.original_wait_seconds}")
    print(f"Temp correlation (seconds): {cfg.temp_correlation_seconds}")
    print(f"Thumb candidates enabled: {cfg.thumb_candidates_enabled}")
    print(f"Manual order guard: {cfg.manual_order_guard_enabled}")
    print(f"Manual burst gap (seconds): {cfg.manual_burst_gap_seconds}")
    print(f"Manual burst max (seconds): {cfg.manual_burst_max_seconds}")
    print(f"Verification column: {cfg.verification_column_name}")
    print(f"UI force download: {cfg.ui_force_download_enabled}")
    print(f"UI force delay (seconds): {cfg.ui_force_delay_seconds}")
    print(f"UI force scope: {cfg.ui_force_scope}")
    print(f"UI focus policy: {cfg.ui_focus_policy}")
    print(f"UI batch mode: {cfg.ui_batch_mode}")
    print(f"UI item timeout (seconds): {cfg.ui_item_timeout_seconds}")
    print(f"UI retry backoff (seconds): {cfg.ui_retry_backoff_seconds}")
    print(f"UI window backends: {cfg.ui_window_backends}")
    print(f"UI window classes: {cfg.ui_window_classes}")
    print(f"Sheet order scope: {cfg.sheet_order_scope}")
    print(f"Sheet materialization order: {cfg.sheet_materialization_order}")
    print(f"Sheet commit order: {cfg.sheet_commit_order}")
    if cfg.ui_force_download_enabled:
        if not UI_FORCE_DOWNLOADER_AVAILABLE or WeChatUIForceDownloader is None:
            err = UI_FORCE_DOWNLOADER_IMPORT_ERROR or "ui_downloader_unavailable"
            print(f"[WARN] UI probe failed | err={err}")
        else:
            try:
                ui_probe = WeChatUIForceDownloader(
                    focus_policy=cfg.ui_focus_policy,
                    item_timeout_seconds=cfg.ui_item_timeout_seconds,
                    window_backends=cfg.ui_window_backends,
                    window_class_candidates=cfg.ui_window_classes,
                )
                probe_ok, probe_note = ui_probe.probe_main_window()
                if probe_ok:
                    print("UI probe: ok")
                else:
                    print(f"[WARN] UI probe failed | err={probe_note}")
            except Exception as exc:
                print(f"[WARN] UI probe failed | err={type(exc).__name__}: {exc}")

    ensure_client_map_file(cfg.client_map_path, cfg.watch_roots)
    resolver = ClientResolver(cfg.client_map_path)
    media_resolver: Optional[WeChatDBResolver] = None
    if cfg.resolution_mode == "db-first":
        media_resolver = WeChatDBResolver(cfg.watch_roots, cfg.db_merge_path, refresh_seconds=10)

    db = StateDB(cfg.db_path)
    if cfg.ui_force_download_enabled:
        if db.get_meta(UI_FORCE_RUNTIME_META_KEY) is None:
            db.set_ui_force_runtime_enabled(True, release_waiting=False)
    else:
        db.set_ui_force_runtime_enabled(False, release_waiting=False)
        db.start_manual_session()
    if not cfg.thumb_candidates_enabled:
        ignored_manual_open_only = db.ignore_manual_open_only_waits()
        if ignored_manual_open_only:
            print(f"[RECOVER] ignored_manual_open_only_waits={ignored_manual_open_only}")
    if cfg.manual_order_guard_enabled:
        stale_placeholders, stale_released = db.ignore_stale_manual_sessions(
            max_age_sec=max(1800, cfg.original_wait_seconds * 8),
        )
        if stale_placeholders or stale_released:
            print(
                f"[RECOVER] stale_manual_sessions_placeholders={stale_placeholders} "
                f"| released_files={stale_released}"
            )
    ignored_old = db.ignore_stale_queue(time.time() - max(1, cfg.recent_files_hours) * 3600)
    if ignored_old:
        print(f"[RECOVER] ignored_old_queue={ignored_old} | older_than_hours={cfg.recent_files_hours}")
    cleaned_temp_orphans = db.cleanup_stale_temp_orphans(max_age_sec=max(600, cfg.original_wait_seconds * 4))
    if cleaned_temp_orphans:
        print(f"[RECOVER] stale_temp_orphans={cleaned_temp_orphans}")
    recovered_processing_retry, recovered_processing_done = db.recover_stale_processing(
        max_age_sec=max(180, cfg.original_wait_seconds * 2)
    )
    if recovered_processing_retry or recovered_processing_done:
        print(
            f"[RECOVER] stale_processing_retry={recovered_processing_retry} "
            f"| stale_processing_done={recovered_processing_done}"
        )
    try:
        sink = build_sink(cfg)
    except Exception as exc:
        print(str(exc))
        return 4

    if isinstance(sink, GoogleSheetsSink):
        main_sheet = sink._main_sheet_title or sink._resolve_main_sheet_title()
        review_sheet = sink.review_worksheet or main_sheet
        label = sink.spreadsheet_title or sink.spreadsheet_id
        print(
            f"Google Sheets: {label} | main={main_sheet} | review={review_sheet} | auth={sink.auth_mode}"
        )
    if media_resolver is not None:
        if media_resolver.refresh_if_due(force=True):
            print(f"WeChat DB resolver: enabled | wx_dir={media_resolver.selected_wx_dir} | merge={cfg.db_merge_path}")
        else:
            print(f"WeChat DB resolver: degraded_to_path_only | err={media_resolver.last_error or 'unknown'}")
    print(f"UI force runtime enabled: {db.is_ui_force_runtime_enabled(default_enabled=cfg.ui_force_download_enabled)}")

    requeued = db.requeue_mapped_missing_client(resolver, max_age_hours=3, limit=1200)
    if requeued:
        print(f"[RECOVER] requeued_missing_client={requeued}")
    backfilled = db.backfill_receipt_context(resolver, limit=8000)
    if backfilled:
        print(f"[RECOVER] backfilled_receipt_context={backfilled}")
    parser_backfilled, sheet_backfilled, sheet_backfill_failed = backfill_missing_receipt_fields(db, sink, cfg, limit=8000)
    if parser_backfilled:
        print(
            f"[RECOVER] backfilled_receipt_fields={parser_backfilled}"
            f" | sheet_updated={sheet_backfilled}"
            f" | sheet_failed={sheet_backfill_failed}"
        )
    try:
        ocr = build_ocr_engine()
    except Exception as exc:
        print(str(exc))
        return 3
    print(f"OCR engine: {ocr.name}")
    warmup_ocr_engine(ocr)

    observer: Optional[Observer] = None
    if WATCHDOG_AVAILABLE and not cfg.disable_watchdog:
        observer = Observer()
        handler = IngestEventHandler(db, cfg, media_resolver)
        for root in cfg.watch_roots:
            if root.exists():
                observer.schedule(handler, str(root), recursive=True)
        observer.start()
        print("Watchdog: enabled")
    else:
        print("Watchdog: disabled (using reconcile polling only)")

    stop_event = threading.Event()
    ui_worker: Optional[UIForceDownloadWorker] = None
    if cfg.ui_force_download_enabled and cfg.resolution_mode == "db-first":
        ui_worker = UIForceDownloadWorker(db=db, cfg=cfg, stop_event=stop_event, media_resolver=media_resolver)
        if ui_worker.available:
            ui_worker.start()
            print("UI force download worker: enabled")
        else:
            print(f"UI force download worker: unavailable | err={ui_worker.unavailable_reason}")
    else:
        print("UI force download worker: disabled")

    last_reconcile = 0.0
    try:
        while True:
            now = time.time()
            if now - last_reconcile >= cfg.reconcile_seconds:
                added = reconcile_scan(cfg, db)
                last_reconcile = now
                print(f"[SCAN] reconcile complete | queued_or_refreshed={added}")

            flush_ready_sink_rows(db, sink, cfg, media_resolver=media_resolver, max_rows=50)
            manual_session_started_at = db.get_manual_session_started_at() if is_manual_materialization_mode(db, cfg) else None
            manual_session_id = db.get_current_manual_session_id() if is_manual_materialization_mode(db, cfg) else None
            seed_ready_manual_session_placeholders(db, media_resolver, cfg)
            item = db.claim_next(
                manual_session_started_at=manual_session_started_at,
                manual_session_id=manual_session_id,
            )
            if item is None:
                flush_ready_sink_rows(db, sink, cfg, media_resolver=media_resolver, max_rows=50)
                time.sleep(cfg.idle_sleep_seconds)
                continue
            process_item(
                item=item,
                db=db,
                sink=sink,
                ocr=ocr,
                resolver=resolver,
                media_resolver=media_resolver,
                cfg=cfg,
            )
            flush_ready_sink_rows(db, sink, cfg, media_resolver=media_resolver, max_rows=50)
    except KeyboardInterrupt:
        print("Stopping daemon...")
    finally:
        stop_event.set()
        if ui_worker is not None and ui_worker.is_alive():
            ui_worker.join(timeout=5)
        if observer is not None:
            observer.stop()
            observer.join(timeout=5)
        db.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
